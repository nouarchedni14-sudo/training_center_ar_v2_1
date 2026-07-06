from __future__ import annotations

import calendar
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import content_disposition_header
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

try:
    import arabic_reshaper
except Exception:  # pragma: no cover
    arabic_reshaper = None

try:
    from bidi.algorithm import get_display
except Exception:  # pragma: no cover
    get_display = None

from .models import (
    UserAttendanceSummaryArchive,
    خليةغياب,
    كشفغياب,
    دفعة,
    تمهين,
)
from .permissions import require_program_permission
from .program_split_utils import exclude_inactive_trainees

PAGE_TITLE = "متابعة الحضور والغيابات لدى المستخدم"
PROGRAM = "apprentice"

MONTH_CHOICES = [
    (1, "جانفي 01"),
    (2, "فيفري 02"),
    (3, "مارس 03"),
    (4, "أفريل 04"),
    (5, "ماي 05"),
    (6, "جوان 06"),
    (7, "جويلية 07"),
    (8, "أوت 08"),
    (9, "سبتمبر 09"),
    (10, "أكتوبر 10"),
    (11, "نوفمبر 11"),
    (12, "ديسمبر 12"),
]

STATUS_CHOICES = [
    ("current", "الحاليون"),
    ("graduated", "المتخرجون"),
    ("removed", "المشطوبون / المفصولون"),
    ("all", "الكل"),
]

# ترقيم Python: الإثنين=0 ... الأحد=6. في هذه الصفحة نسمح من السبت إلى الخميس فقط.
WEEKDAY_CHOICES = [
    (5, "السبت"),
    (6, "الأحد"),
    (0, "الإثنين"),
    (1, "الثلاثاء"),
    (2, "الأربعاء"),
    (3, "الخميس"),
]
WEEKDAY_LABELS = dict(WEEKDAY_CHOICES)
STATUS_INPUT_CHOICES = [
    ("", "—"),
    ("present", "ح"),
    ("absent", "غ"),
]


def _normalize_attendance_status(value) -> str:
    """توحيد حالات هذه الصفحة إلى حاضر/غائب فقط.

    إذا وجدت قيم قديمة من صفحات أخرى مثل غائب بعذر أو متأخر،
    نعرضها هنا ضمن حاضر/غائب حتى تبقى الصفحة مبسطة كما طلب المستخدم.
    """
    value = str(value or "").strip()
    if value == "present":
        return "present"
    if value == "absent":
        return "absent"
    if value == "excused":
        return "absent"
    if value == "late":
        return "present"
    return ""


def _month_label(month: int) -> str:
    return dict(MONTH_CHOICES).get(int(month), str(month))


def _parse_int(value, default=None):
    try:
        if value in (None, ""):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_date(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _request_value(request, key, default=""):
    if request.method == "POST":
        return request.POST.get(key, request.GET.get(key, default))
    return request.GET.get(key, default)


def _default_period(month: int, year: int) -> tuple[date, date]:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _safe_date(value):
    return value.strftime("%Y-%m-%d") if value else ""


def _status_short(value) -> str:
    value = _normalize_attendance_status(value)
    return {"present": "ح", "absent": "غ"}.get(value, "")


def _status_css(value) -> str:
    value = _normalize_attendance_status(value)
    return {"present": "present", "absent": "absent"}.get(value, "empty")


def _status_long(value) -> str:
    value = _normalize_attendance_status(value)
    return {"present": "حاضر", "absent": "غائب"}.get(value, "غير مسجل")


def _date_column_label(value: date) -> str:
    return value.strftime("%d/%m")


def _training_duration(obj) -> str:
    start = getattr(obj, "تاريخ_بداية_التكوين", None)
    end = getattr(obj, "تاريخ_نهاية_التكوين", None)
    if not start and not end:
        return ""
    if start and end:
        months = (end.year - start.year) * 12 + (end.month - start.month)
        if end.day >= start.day:
            months += 1
        if months >= 12:
            years = months // 12
            rest = months % 12
            if rest:
                return f"{years} سنة و {rest} شهر"
            return f"{years} سنة"
        return f"{months} شهر"
    return f"{_safe_date(start)} إلى {_safe_date(end)}"


def _trainee_name(obj) -> str:
    return " ".join([str(getattr(obj, "اللقب", "") or "").strip(), str(getattr(obj, "الاسم", "") or "").strip()]).strip()


def _selected_weekdays_from_request(request):
    weekdays = []
    for idx in range(1, 6):
        raw = (_request_value(request, f"weekday{idx}") or "").strip()
        if raw == "":
            continue
        try:
            value = int(raw)
        except (TypeError, ValueError):
            continue
        if value not in WEEKDAY_LABELS:
            continue
        if value not in weekdays:
            weekdays.append(value)
    return weekdays


def _month_dates_for_weekdays(start_date: date, end_date: date, weekdays):
    if not weekdays:
        return []
    result = []
    current = start_date
    from datetime import timedelta
    selected = set(int(x) for x in weekdays)
    while current <= end_date:
        if current.weekday() in selected:
            result.append({
                "date": current,
                "iso": current.isoformat(),
                "label": current.strftime("%Y-%m-%d"),
                "day_label": WEEKDAY_LABELS.get(current.weekday(), ""),
            })
        current += timedelta(days=1)
    return result


def _build_filter_context(request):
    today = timezone.localdate()
    month = _parse_int(_request_value(request, "month"), today.month)
    year = _parse_int(_request_value(request, "year"), today.year)
    period_start, period_end = _default_period(month, year)
    date_from = _parse_date(_request_value(request, "date_from")) or period_start
    date_to = _parse_date(_request_value(request, "date_to")) or period_end
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    selected_weekdays = _selected_weekdays_from_request(request)
    filters = {
        "month": month,
        "year": year,
        "date_from": _safe_date(date_from),
        "date_to": _safe_date(date_to),
        "promotion": (_request_value(request, "promotion") or "").strip(),
        "specialty": (_request_value(request, "specialty") or "").strip(),
        "semester": (_request_value(request, "semester") or "").strip(),
        # هذه الصفحة خاصة بالمتكونين الحاليين فقط؛ لا نسمح بإدراج المشطوبين أو المتخرجين هنا.
        "status": "current",
        "user": (_request_value(request, "user") or "").strip(),
        "q": (_request_value(request, "q") or "").strip(),
        "weekday1": str(selected_weekdays[0]) if len(selected_weekdays) > 0 else "",
        "weekday2": str(selected_weekdays[1]) if len(selected_weekdays) > 1 else "",
        "weekday3": str(selected_weekdays[2]) if len(selected_weekdays) > 2 else "",
        "weekday4": str(selected_weekdays[3]) if len(selected_weekdays) > 3 else "",
        "weekday5": str(selected_weekdays[4]) if len(selected_weekdays) > 4 else "",
    }
    return filters, date_from, date_to, selected_weekdays

def _apply_status_filter(qs, status: str):
    today = timezone.localdate()
    inactive_q = (
        Q(الحالة__icontains="مشطوب") |
        Q(الحالة__icontains="شطب") |
        Q(الحالة__icontains="مفصول") |
        Q(الحالة__icontains="فصل") |
        Q(الحالة__icontains="منقطع") |
        Q(الحالة__icontains="متوقف") |
        Q(الحالة__icontains="موقوف") |
        Q(الحالة__icontains="انسحب")
    )
    if status == "current":
        return exclude_inactive_trainees(qs).filter(Q(تاريخ_نهاية_التكوين__isnull=True) | Q(تاريخ_نهاية_التكوين__gt=today))
    if status == "graduated":
        return exclude_inactive_trainees(qs).filter(تاريخ_نهاية_التكوين__lte=today)
    if status == "removed":
        return qs.filter(inactive_q)
    return qs


def _base_trainee_queryset(filters):
    qs = تمهين.objects.select_related("الدفعة").all()
    qs = _apply_status_filter(qs, filters.get("status") or "current")
    if filters.get("promotion"):
        qs = qs.filter(الدفعة_id=filters["promotion"])
    if filters.get("specialty"):
        qs = qs.filter(التخصص=filters["specialty"])
    if filters.get("semester"):
        qs = qs.filter(السداسي=filters["semester"])
    if filters.get("user"):
        qs = qs.filter(المستخدم=filters["user"])
    q = filters.get("q") or ""
    if q:
        qs = qs.filter(
            Q(اللقب__icontains=q) |
            Q(الاسم__icontains=q) |
            Q(رقم_التسجيل__icontains=q) |
            Q(التخصص__icontains=q) |
            Q(المستخدم__icontains=q)
        )
    return qs.order_by("المستخدم", "التخصص", "اللقب", "الاسم")


def _current_active_queryset():
    today = timezone.localdate()
    return exclude_inactive_trainees(تمهين.objects.all()).filter(
        Q(تاريخ_نهاية_التكوين__isnull=True) | Q(تاريخ_نهاية_التكوين__gt=today)
    )


def _apply_option_filters(qs, filters, exclude_key: str = ""):
    """فلترة خيارات القوائم حسب الاختيارات الحالية، مع استثناء القائمة نفسها.

    مثال: عند اختيار دفعة، قائمة التخصصات والسداسيات والمستخدمين تعرض فقط
    ما يخص تلك الدفعة. وعند اختيار سداسي، تتقلص التخصصات والمستخدمون تلقائيًا.
    """
    if exclude_key != "promotion" and filters.get("promotion"):
        qs = qs.filter(الدفعة_id=filters["promotion"])
    if exclude_key != "specialty" and filters.get("specialty"):
        qs = qs.filter(التخصص=filters["specialty"])
    if exclude_key != "semester" and filters.get("semester"):
        qs = qs.filter(السداسي=filters["semester"])
    if exclude_key != "user" and filters.get("user"):
        qs = qs.filter(المستخدم=filters["user"])
    q = filters.get("q") or ""
    if q:
        qs = qs.filter(
            Q(اللقب__icontains=q) |
            Q(الاسم__icontains=q) |
            Q(رقم_التسجيل__icontains=q) |
            Q(التخصص__icontains=q) |
            Q(المستخدم__icontains=q)
        )
    return qs


def _options_context(filters=None):
    filters = filters or {}
    active_qs = _current_active_queryset()

    promotion_qs = _apply_option_filters(active_qs, filters, exclude_key="promotion")
    promotion_ids = promotion_qs.exclude(الدفعة__isnull=True).values_list("الدفعة_id", flat=True).distinct()
    promotion_options = دفعة.objects.filter(id__in=promotion_ids, مفعلة=True).order_by("-السنة", "-رقم_الدورة")

    specialty_qs = _apply_option_filters(active_qs, filters, exclude_key="specialty")
    specialty_options = sorted({x for x in specialty_qs.exclude(التخصص__isnull=True).exclude(التخصص="").values_list("التخصص", flat=True) if str(x or "").strip()})

    semester_qs = _apply_option_filters(active_qs, filters, exclude_key="semester")
    semester_options = sorted({x for x in semester_qs.exclude(السداسي__isnull=True).exclude(السداسي="").values_list("السداسي", flat=True) if str(x or "").strip()})

    user_qs = _apply_option_filters(active_qs, filters, exclude_key="user")
    user_options = sorted({x for x in user_qs.exclude(المستخدم__isnull=True).exclude(المستخدم="").values_list("المستخدم", flat=True) if str(x or "").strip()})

    return {
        "month_choices": MONTH_CHOICES,
        "status_choices": STATUS_CHOICES,
        "weekday_choices": WEEKDAY_CHOICES,
        "status_input_choices": STATUS_INPUT_CHOICES,
        "promotion_options": promotion_options,
        "specialty_options": specialty_options,
        "semester_options": semester_options,
        "user_options": user_options,
    }


def _count_entries(trainee_ids, start_date, end_date):
    result = defaultdict(Counter)
    if not trainee_ids:
        return result
    qs = خليةغياب.objects.filter(
        الكشف__البرنامج=PROGRAM,
        trainee_id__in=trainee_ids,
        التاريخ__gte=start_date,
        التاريخ__lte=end_date,
    ).values("trainee_id", "الحالة")
    for item in qs.iterator(chunk_size=2000):
        status = _normalize_attendance_status(item.get("الحالة"))
        if not status:
            continue
        result[item["trainee_id"]][status] += 1
    return result


def _present_count(counter: Counter) -> int:
    return int(counter.get("present", 0))


def _absent_count(counter: Counter) -> int:
    return int(counter.get("absent", 0))


def _get_or_create_summary_sheet(filters, selected_weekdays, user=None):
    promotion = None
    if filters.get("promotion"):
        try:
            promotion = دفعة.objects.get(pk=filters["promotion"])
        except دفعة.DoesNotExist:
            promotion = None
    defaults = {
        "created_by": user if getattr(user, "is_authenticated", False) else None,
        "يوم_الدراسة_1": selected_weekdays[0] if len(selected_weekdays) > 0 else None,
        "يوم_الدراسة_2": selected_weekdays[1] if len(selected_weekdays) > 1 else None,
        "يوم_الدراسة_3": selected_weekdays[2] if len(selected_weekdays) > 2 else None,
        "يوم_الدراسة_4": selected_weekdays[3] if len(selected_weekdays) > 3 else None,
        "يوم_الدراسة_5": selected_weekdays[4] if len(selected_weekdays) > 4 else None,
    }
    sheet, _created = كشفغياب.objects.get_or_create(
        البرنامج=PROGRAM,
        الدفعة=promotion,
        التخصص=filters.get("specialty") or "",
        الشهر=filters["month"],
        السنة=filters["year"],
        defaults=defaults,
    )
    update_fields = []
    for field_name, value in defaults.items():
        if field_name == "created_by":
            continue
        if getattr(sheet, field_name) != value:
            setattr(sheet, field_name, value)
            update_fields.append(field_name)
    if update_fields:
        update_fields.append("updated_at")
        sheet.save(update_fields=update_fields)
    return sheet


def _birth_place_only(obj) -> str:
    """عرض مكان الازدياد فقط دون الولاية.

    بعض الصفحات القديمة كانت تعرض البلدية والولاية معًا مثل:
    "تيسمسيلت - تيسمسيلت". في هذه الصفحة نحتاج البلدية فقط.
    """
    for field_name in ("بلدية_الميلاد", "مكان_الميلاد", "مكان_الازدياد", "البلدية"):
        value = getattr(obj, field_name, None)
        text = str(value or "").replace("—", "-").strip()
        if text:
            if "-" in text:
                text = text.split("-", 1)[0].strip()
            return text
    return ""


def _attendance_detail_columns_and_map(trainee_ids, columns, sheet=None):
    """Load saved statuses for the selected days so the page can be used for entry and summary."""
    if not trainee_ids or not columns:
        return {}
    date_values = [col["date"] for col in columns]
    qs = خليةغياب.objects.filter(
        الكشف__البرنامج=PROGRAM,
        trainee_id__in=trainee_ids,
        التاريخ__in=date_values,
        رقم_الخانة=1,
    )
    if sheet is not None:
        qs = qs.filter(الكشف=sheet)
    detail_map = {}
    for entry in qs.only("trainee_id", "التاريخ", "الحالة"):
        status = _normalize_attendance_status(entry.الحالة)
        detail_map[(entry.trainee_id, entry.التاريخ.isoformat())] = {
            "value": status,
            "display": _status_short(status),
            "title": _status_long(status),
            "css": _status_css(status),
        }
    return detail_map

def _build_rows(filters, date_from, date_to, selected_weekdays=None, sheet=None):
    selected_weekdays = selected_weekdays or []
    qs = _base_trainee_queryset(filters)
    trainees = list(qs[:10000])
    trainee_ids = [obj.pk for obj in trainees]
    detail_columns = _month_dates_for_weekdays(date_from, date_to, selected_weekdays)
    detail_map = _attendance_detail_columns_and_map(trainee_ids, detail_columns, sheet=sheet)

    # أعداد الفترة من الأيام المختارة فقط ومن الجدول الحالي حتى لا تختلط مع جداول أخرى.
    period_counts = defaultdict(Counter)
    for (trainee_id, _iso), cell in detail_map.items():
        status = _normalize_attendance_status(cell.get("value"))
        if status:
            period_counts[trainee_id][status] += 1

    # المجموع التراكمي يبقى من كل جداول الغياب اليومية للتمهين إلى غاية تاريخ النهاية.
    cumulative_counts = _count_entries(trainee_ids, date(1900, 1, 1), date_to)

    rows = []
    totals = Counter()
    for idx, obj in enumerate(trainees, start=1):
        period_counter = period_counts.get(obj.pk, Counter())
        cumulative_counter = cumulative_counts.get(obj.pk, Counter())
        start_training = getattr(obj, "تاريخ_بداية_التكوين", None)
        end_training = getattr(obj, "تاريخ_نهاية_التكوين", None)
        period_cells = []
        for col in detail_columns:
            cell = detail_map.get((obj.pk, col["iso"]), {"value": "", "display": "", "title": "غير مسجل", "css": "empty"})
            period_cells.append({**cell, "iso": col["iso"]})

        row = {
            "index": idx,
            "trainee_id": obj.pk,
            "trainee": obj,
            "name": _trainee_name(obj),
            "birth_date": _safe_date(getattr(obj, "تاريخ_الميلاد", None)),
            "birth_place": _birth_place_only(obj),
            "specialty": getattr(obj, "التخصص", "") or "",
            "registration_number": getattr(obj, "رقم_التسجيل", "") or "",
            "training_start": _safe_date(start_training),
            "training_end": _safe_date(end_training),
            "user": getattr(obj, "المستخدم", "") or "",
            "semester": getattr(obj, "السداسي", "") or "",
            "status": getattr(obj, "الحالة", "") or "",
            "period_present": _present_count(period_counter),
            "period_absent": _absent_count(period_counter),
            "total_present": _present_count(cumulative_counter),
            "total_absent": _absent_count(cumulative_counter),
            "period_cells": period_cells,
        }
        row["base_total_present"] = max(0, row["total_present"] - row["period_present"])
        row["base_total_absent"] = max(0, row["total_absent"] - row["period_absent"])
        total_period = row["period_present"] + row["period_absent"]
        total_cumulative = row["total_present"] + row["total_absent"]
        row["period_attendance_rate"] = round((row["period_present"] / total_period) * 100, 2) if total_period else 0
        row["period_absence_rate"] = round((row["period_absent"] / total_period) * 100, 2) if total_period else 0
        row["total_attendance_rate"] = round((row["total_present"] / total_cumulative) * 100, 2) if total_cumulative else 0
        row["total_absence_rate"] = round((row["total_absent"] / total_cumulative) * 100, 2) if total_cumulative else 0
        rows.append(row)
        totals["period_present"] += row["period_present"]
        totals["period_absent"] += row["period_absent"]
        totals["total_present"] += row["total_present"]
        totals["total_absent"] += row["total_absent"]
    totals["rows_count"] = len(rows)
    return rows, totals, detail_columns

def _current_query(request):
    query = request.GET.copy()
    query.pop("page", None)
    return query.urlencode()


def _summary_context(request):
    filters, date_from, date_to, selected_weekdays = _build_filter_context(request)
    sheet = None
    if selected_weekdays:
        sheet = _get_or_create_summary_sheet(filters, selected_weekdays, user=request.user)
    rows, totals, detail_columns = _build_rows(filters, date_from, date_to, selected_weekdays, sheet=sheet)
    paginator = Paginator(rows, 500)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    context = {
        "title": PAGE_TITLE,
        "page_title": PAGE_TITLE,
        "program_label": "التمهين",
        "filters": filters,
        "date_from": _safe_date(date_from),
        "date_to": _safe_date(date_to),
        "rows": list(page_obj.object_list),
        "all_rows": rows,
        "page_obj": page_obj,
        "totals": totals,
        "detail_columns": detail_columns,
        "selected_weekdays": selected_weekdays,
        "has_selected_weekdays": bool(selected_weekdays),
        "query_string": _current_query(request),
        "print_url": reverse("user_attendance_summary_print") + ("?" + _current_query(request) if _current_query(request) else ""),
        "excel_url": reverse("user_attendance_summary_export", args=["xlsx"]) + ("?" + _current_query(request) if _current_query(request) else ""),
        "pdf_url": reverse("user_attendance_summary_export", args=["pdf"]) + ("?" + _current_query(request) if _current_query(request) else ""),
        "archive_url": reverse("user_attendance_summary_archive_create") + ("?" + _current_query(request) if _current_query(request) else ""),
    }
    context.update(_options_context(filters))
    return context


def _redirect_query_from_post(request):
    keys = ["month", "year", "date_from", "date_to", "promotion", "specialty", "semester", "user", "q", "weekday1", "weekday2", "weekday3", "weekday4", "weekday5"]
    parts = []
    from urllib.parse import urlencode
    data = {key: request.POST.get(key, "") for key in keys if request.POST.get(key, "") not in (None, "")}
    return urlencode(data)


def _save_user_attendance_summary(request):
    filters, date_from, date_to, selected_weekdays = _build_filter_context(request)
    if not selected_weekdays:
        messages.error(request, "اختر يومًا واحدًا على الأقل قبل حفظ الحضور والغيابات.")
        return

    sheet = _get_or_create_summary_sheet(filters, selected_weekdays, user=request.user)
    rows, _totals, detail_columns = _build_rows(filters, date_from, date_to, selected_weekdays, sheet=sheet)
    trainee_ids = [row["trainee_id"] for row in rows]
    dates = [col["date"] for col in detail_columns]
    existing = {
        (entry.trainee_id, entry.التاريخ.isoformat()): entry
        for entry in خليةغياب.objects.filter(الكشف=sheet, trainee_id__in=trainee_ids, التاريخ__in=dates, رقم_الخانة=1)
    }

    valid_statuses = {"present", "absent"}
    date_lookup = {col["iso"]: col["date"] for col in detail_columns}
    visible_trainee_ids = {row["trainee_id"] for row in rows}
    to_create = []
    to_update = []
    to_delete = []

    # نعالج فقط الخانات التي وصلت من الصفحة الحالية، حتى لا يتم حذف صفحات أخرى عند وجود pagination.
    for field_name, posted_value in request.POST.items():
        if not field_name.startswith("status__"):
            continue
        try:
            _prefix, trainee_id_raw, iso_value = field_name.split("__", 2)
            trainee_id = int(trainee_id_raw)
        except (ValueError, TypeError):
            continue
        if trainee_id not in visible_trainee_ids or iso_value not in date_lookup:
            continue
        raw = _normalize_attendance_status(posted_value)
        entry = existing.get((trainee_id, iso_value))
        if raw not in valid_statuses:
            if entry is not None:
                to_delete.append(entry.pk)
            continue
        if entry is None:
            to_create.append(خليةغياب(
                الكشف=sheet,
                trainee_id=trainee_id,
                التاريخ=date_lookup[iso_value],
                رقم_الخانة=1,
                الحالة=raw,
                recorded_by=request.user,
            ))
        elif entry.الحالة != raw or entry.recorded_by_id != request.user.id:
            entry.الحالة = raw
            entry.recorded_by = request.user
            to_update.append(entry)

    with transaction.atomic():
        if to_delete:
            خليةغياب.objects.filter(pk__in=to_delete).delete()
        if to_create:
            خليةغياب.objects.bulk_create(to_create, batch_size=1000)
        if to_update:
            خليةغياب.objects.bulk_update(to_update, ["الحالة", "recorded_by", "updated_at"], batch_size=1000)
    messages.success(request, f"تم حفظ الحضور والغيابات: {len(to_create) + len(to_update)} خانة، وحذف {len(to_delete)} خانة فارغة.")


def _clear_period_attendance(request):
    filters, date_from, date_to, selected_weekdays = _build_filter_context(request)
    if not selected_weekdays:
        messages.error(request, "اختر أيام الدراسة قبل مسح عدد الحضور والغيابات في الفترة.")
        return
    sheet = _get_or_create_summary_sheet(filters, selected_weekdays, user=request.user)
    rows, _totals, detail_columns = _build_rows(filters, date_from, date_to, selected_weekdays, sheet=sheet)
    trainee_ids = [row["trainee_id"] for row in rows]
    dates = [col["date"] for col in detail_columns]
    deleted, _ = خليةغياب.objects.filter(
        الكشف=sheet,
        trainee_id__in=trainee_ids,
        التاريخ__in=dates,
        رقم_الخانة=1,
    ).delete()
    messages.success(request, f"تم مسح عدد الحضور والغيابات للفترة الحالية. عدد الخانات المحذوفة: {deleted}.")


def _clear_cumulative_attendance(request):
    filters, _date_from, date_to, _selected_weekdays = _build_filter_context(request)
    rows, _totals, _detail_columns = _build_rows(filters, date(1900, 1, 1), date_to, [], sheet=None)
    trainee_ids = [row["trainee_id"] for row in rows]
    if not trainee_ids:
        messages.warning(request, "لا توجد سجلات مطابقة للفلاتر الحالية لمسح مجموع الأعمدة.")
        return
    deleted, _ = خليةغياب.objects.filter(
        الكشف__البرنامج=PROGRAM,
        trainee_id__in=trainee_ids,
        التاريخ__lte=date_to,
        رقم_الخانة=1,
    ).delete()
    messages.success(request, f"تم مسح مجموع أعمدة الحضور والغيابات للمتكونين المعروضين إلى غاية {date_to:%Y-%m-%d}. عدد الخانات المحذوفة: {deleted}.")

def _row_to_export_values(row, detail_columns=None):
    values = [
        row["name"], row["birth_date"], row["birth_place"], row["specialty"], row["registration_number"],
        row.get("training_start", ""), row.get("training_end", ""),
    ]
    if detail_columns:
        values.extend([cell.get("display", "") for cell in row.get("period_cells", [])])
    values.extend([
        row["period_present"], row["period_absent"], row["total_present"], row["total_absent"],
        row["total_attendance_rate"], row["total_absence_rate"], row["user"], row["semester"], row["status"],
    ])
    return values


def _export_headers(detail_columns=None):
    headers = [
        "اسم المتكون", "تاريخ الازدياد", "مكان الازدياد", "التخصص", "رقم التسجيل",
        "تاريخ بداية التكوين", "تاريخ نهاية التكوين",
    ]
    if detail_columns:
        headers.extend([f"{col['label']}" for col in detail_columns])
    headers.extend([
        "عدد الحضور", "عدد الغيابات", "مجموع الحضور", "مجموع الغيابات",
        "نسبة الحضور", "نسبة الغياب", "المستخدم", "السداسي", "الحالة",
    ])
    return headers


@login_required
def user_attendance_summary(request):
    require_program_permission(request, PROGRAM, "view")
    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action in {"save_attendance", "clear_period", "clear_cumulative"}:
            require_program_permission(request, PROGRAM, "change")
            if action == "save_attendance":
                _save_user_attendance_summary(request)
            elif action == "clear_period":
                _clear_period_attendance(request)
            elif action == "clear_cumulative":
                _clear_cumulative_attendance(request)
            query = _redirect_query_from_post(request)
            url = reverse("user_attendance_summary")
            return redirect(f"{url}?{query}" if query else url)
    context = _summary_context(request)
    return render(request, "trainees/user_attendance_summary.html", context)


@login_required
def user_attendance_summary_print(request):
    require_program_permission(request, PROGRAM, "view")
    context = _summary_context(request)
    context["print_mode"] = True
    return render(request, "trainees/user_attendance_summary_print.html", context)


@login_required
def user_attendance_summary_export(request, fmt):
    require_program_permission(request, PROGRAM, "view")
    fmt = (fmt or "").lower().strip()
    context = _summary_context(request)
    rows = context["all_rows"]
    if fmt == "xlsx":
        return _export_excel(rows, context)
    if fmt == "pdf":
        return _export_pdf(rows, context)
    raise Http404()


def _export_excel(rows, context):
    wb = Workbook()
    ws = wb.active
    ws.title = "متابعة الحضور"
    detail_columns = context.get("detail_columns") or []
    headers = _export_headers(detail_columns)
    ws.append([PAGE_TITLE])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([f"الفترة: {context['date_from']} إلى {context['date_to']}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="B7C6D8")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in rows:
        ws.append(_row_to_export_values(row, detail_columns))
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [28, 14, 22, 28, 16, 16, 16] + ([8] * len(detail_columns)) + [12, 12, 14, 14, 12, 12, 22, 12, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = content_disposition_header(True, "متابعة_الحضور_والغيابات_لدى_المستخدم.xlsx")
    return response


def _register_pdf_font():
    candidates = [
        Path(settings.BASE_DIR) / "fonts" / "Amiri-Regular.ttf",
        Path(settings.BASE_DIR) / "fonts" / "NotoNaskhArabic-Regular.ttf",
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/tahoma.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for candidate in candidates:
        try:
            if candidate.exists():
                if "ArabicUI" not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont("ArabicUI", str(candidate)))
                return "ArabicUI"
        except Exception:
            continue
    return "Helvetica"


def _pdf_text(value):
    text = str(value or "")
    try:
        if arabic_reshaper is not None and get_display is not None and any("\u0600" <= ch <= "\u06FF" for ch in text):
            return get_display(arabic_reshaper.reshape(text))
    except Exception:
        pass
    return text


def _export_pdf(rows, context):
    buffer = BytesIO()
    font_name = _register_pdf_font()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name
    styles["Normal"].fontName = font_name
    story = [Paragraph(_pdf_text(PAGE_TITLE), styles["Title"]), Spacer(1, 8), Paragraph(_pdf_text(f"الفترة: {context['date_from']} إلى {context['date_to']}"), styles["Normal"]), Spacer(1, 8)]
    headers = ["الاسم", "التخصص", "رقم التسجيل", "بداية التكوين", "نهاية التكوين", "الحضور", "الغياب", "مجموع الحضور", "مجموع الغياب", "المستخدم"]
    data = [[_pdf_text(h) for h in headers]]
    for row in rows[:1200]:
        data.append([_pdf_text(x) for x in [row["name"], row["specialty"], row["registration_number"], row.get("training_start", ""), row.get("training_end", ""), row["period_present"], row["period_absent"], row["total_present"], row["total_absent"], row["user"]]])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C6D8")),
    ]))
    story.append(table)
    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = content_disposition_header(True, "متابعة_الحضور_والغيابات_لدى_المستخدم.pdf")
    return response


@login_required
def user_attendance_summary_archive_create(request):
    require_program_permission(request, PROGRAM, "change")
    context = _summary_context(request)
    rows_json = []
    for row in context["all_rows"]:
        item = {k: v for k, v in row.items() if k not in {"trainee"}}
        rows_json.append(item)
    title = f"{PAGE_TITLE} - {_month_label(context['filters']['month'])} {context['filters']['year']}"
    archive = UserAttendanceSummaryArchive.objects.create(
        title=title,
        program=PROGRAM,
        filters_json={**context["filters"], "detail_columns": [{"iso": col.get("iso"), "label": col.get("label")} for col in context.get("detail_columns", [])]},
        rows_json=rows_json,
        row_count=len(rows_json),
        total_present=context["totals"].get("period_present", 0),
        total_absent=context["totals"].get("period_absent", 0),
        created_by=request.user,
    )
    messages.success(request, f"تمت أرشفة التقرير بنجاح: {archive.title}")
    return redirect("user_attendance_summary_archive_detail", pk=archive.pk)


@login_required
def user_attendance_summary_archive(request):
    require_program_permission(request, PROGRAM, "view")
    qs = UserAttendanceSummaryArchive.objects.filter(program=PROGRAM).select_related("created_by").order_by("-created_at")
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(created_by__username__icontains=q))
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    return render(request, "trainees/user_attendance_summary_archive.html", {
        "title": "أرشيف متابعة الحضور والغيابات لدى المستخدم",
        "page_obj": page_obj,
        "q": q,
    })


@login_required
def user_attendance_summary_archive_detail(request, pk):
    require_program_permission(request, PROGRAM, "view")
    archive = get_object_or_404(UserAttendanceSummaryArchive.objects.select_related("created_by"), pk=pk, program=PROGRAM)
    rows = archive.rows_json or []
    archive_filters = archive.filters_json or {}
    return render(request, "trainees/user_attendance_summary_print.html", {
        "title": archive.title,
        "page_title": archive.title,
        "filters": archive_filters,
        "date_from": archive_filters.get("date_from", ""),
        "date_to": archive_filters.get("date_to", ""),
        "detail_columns": archive_filters.get("detail_columns", []),
        "rows": rows,
        "all_rows": rows,
        "totals": {"period_present": archive.total_present, "period_absent": archive.total_absent, "rows_count": archive.row_count},
        "archive": archive,
        "print_mode": True,
    })
