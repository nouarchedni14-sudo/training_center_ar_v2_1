from __future__ import annotations  # استيراد عناصر محددة من مكتبة/وحدة

import os  # استيراد مكتبة/وحدة بايثون
import uuid  # استيراد مكتبة/وحدة بايثون
from dataclasses import dataclass  # استيراد عناصر محددة من مكتبة/وحدة
from typing import Iterable, List, Tuple, Dict, Any  # استيراد عناصر محددة من مكتبة/وحدة

from django.conf import settings  # استيراد عناصر محددة من مكتبة/وحدة
from django.contrib import messages  # استيراد عناصر محددة من مكتبة/وحدة
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, HttpResponseForbidden  # استيراد عناصر محددة من مكتبة/وحدة
from django.shortcuts import render  # استيراد عناصر محددة من مكتبة/وحدة
from django.urls import path, reverse  # استيراد عناصر محددة من مكتبة/وحدة
from django.db.models import DateField  # استيراد عناصر محددة من مكتبة/وحدة

from openpyxl import load_workbook, Workbook  # استيراد عناصر محددة من مكتبة/وحدة
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter  # استيراد عناصر محددة من مكتبة/وحدة

from .admin_filters import status_group  # استيراد عناصر محددة من مكتبة/وحدة
from .semester_utils import compute_semester_with_repeater, compute_semester_for_trainee, resolve_session_year, add_months  # استيراد عناصر محددة من مكتبة/وحدة
from .models import دفعة, refresh_all_promotion_semester_starts, cohort_start_dates_for_model  # استيراد عناصر محددة من مكتبة/وحدة
from .program_columns import excel_columns_for_model, IGNORED_IMPORT_FIELDS
from .evening_training_type import (
    clean_crossing_specialty_label,
    detect_evening_training_type,
    clamp_semester_for_evening_type,
)

SUPPORTED_EXCEL_IMPORT_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_EXCEL_IMPORT_EXTENSIONS_TEXT = "XLSX / XLSM / XLTX / XLTM"

try:
    from .name_aliases import SURNAME_ALIAS_MAP
except Exception:
    SURNAME_ALIAS_MAP = {}
from .audit import audit_view_event, audit_error_event
from .services.import_reporting_service import (
    build_import_issue,
    summarize_import_issues,
    group_import_issues,
    categorize_import_message,
)


import re  # استيراد مكتبة/وحدة بايثون
from datetime import date, datetime  # استيراد عناصر محددة من مكتبة/وحدة
import calendar  # استيراد مكتبة/وحدة بايثون
from zipfile import BadZipFile

def _row_identity_display(last_name, first_name, reg_no):
    parts = []
    full_name = " ".join([p for p in [str(first_name or "").strip(), str(last_name or "").strip()] if p])
    if full_name:
        parts.append(full_name)
    if str(reg_no or "").strip():
        parts.append(f"رقم التسجيل: {str(reg_no).strip()}")
    return " — ".join(parts)


def _strip_nbsp(value):  # تعريف دالة (Function)
    return str(value).replace("\u00A0", " ").strip()  # إرجاع قيمة من الدالة

def clean_date(value):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Return a python date or None. Accepts Excel date (datetime/date) or strings."""
    if value is None:  # شرط (If)
        return None  # إرجاع قيمة من الدالة
    if isinstance(value, datetime):  # شرط (If)
        return value.date()  # إرجاع قيمة من الدالة
    if isinstance(value, date):  # شرط (If)
        return value  # إرجاع قيمة من الدالة

    s = _strip_nbsp(value)  # تعيين قيمة لمتغير/إعداد
    s = re.sub(r"\s+", "", s)  # تعيين قيمة لمتغير/إعداد
    if not s or s.lower() in {"nan", "none", "null"}:  # شرط (If)
        return None  # إرجاع قيمة من الدالة

    # 00-00-YYYY is treated as invalid/unknown for real dates
    if re.fullmatch(r"00[-/ ]00[-/ ]\d{4}", s):  # شرط (If)
        return None  # إرجاع قيمة من الدالة

    # YYYY-MM-DD
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)  # تعيين قيمة لمتغير/إعداد
    if m:  # شرط (If)
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))  # تعيين قيمة لمتغير/إعداد
        try:  # سطر كود لتنفيذ منطق/إعداد
            return date(y, mo, d)  # إرجاع قيمة من الدالة
        except ValueError:  # سطر كود لتنفيذ منطق/إعداد
            return None  # إرجاع قيمة من الدالة

    # DD-MM-YYYY or DD/MM/YYYY
    m = re.fullmatch(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)  # تعيين قيمة لمتغير/إعداد
    if m:  # شرط (If)
        dd, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))  # تعيين قيمة لمتغير/إعداد
        try:  # سطر كود لتنفيذ منطق/إعداد
            return date(y, mo, dd)  # إرجاع قيمة من الدالة
        except ValueError:  # سطر كود لتنفيذ منطق/إعداد
            return None  # إرجاع قيمة من الدالة

    # YYYY/MM/DD
    m = re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)  # تعيين قيمة لمتغير/إعداد
    if m:  # شرط (If)
        y, mo, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))  # تعيين قيمة لمتغير/إعداد
        try:  # سطر كود لتنفيذ منطق/إعداد
            return date(y, mo, dd)  # إرجاع قيمة من الدالة
        except ValueError:  # سطر كود لتنفيذ منطق/إعداد
            return None  # إرجاع قيمة من الدالة

    return None  # إرجاع قيمة من الدالة

def normalize_registration_number(value):
    if not value:
        return ""

    s = str(value)

    # إزالة المسافات (العادية وغير المرئية)
    s = s.replace("\u00A0", "").strip()
    s = "".join(s.split())

    import re

    parts = re.findall(r'\d+|[A-Za-z]+', s)

    numbers = "".join([p for p in parts if p.isdigit()])
    letters = "".join([p for p in parts if p.isalpha()])

    return numbers + letters

def parse_birthdate_with_assumed(value):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """
    Birthdate parsing with an "assumed" flag for partial/unknown dates.  # سطر كود لتنفيذ منطق/إعداد

    Important: the DB can't store month/day=00, so we map partial dates to the *end* of the  # تعيين قيمة لمتغير/إعداد
    known period (as requested):  # سطر كود لتنفيذ منطق/إعداد
    - 00-00-YYYY                => (YYYY-12-31, True)  (year only)  # تعيين قيمة لمتغير/إعداد
    - YYYY-00-00 / YYYY/00/00   => (YYYY-12-31, True)  (year only)  # تعيين قيمة لمتغير/إعداد
    - YYYY-MM-00                => (YYYY-MM-last_day, True) (day unknown)  # تعيين قيمة لمتغير/إعداد
    - valid full date           => (date, False)  # تعيين قيمة لمتغير/إعداد
    - empty/invalid             => (None, False)  # تعيين قيمة لمتغير/إعداد
    # سطر كود لتنفيذ منطق/إعداد
    """
    if value is None:  # شرط (If)
        return None, False  # إرجاع قيمة من الدالة

    if isinstance(value, datetime):  # شرط (If)
        return value.date(), False  # إرجاع قيمة من الدالة
    if isinstance(value, date):  # شرط (If)
        return value, False  # إرجاع قيمة من الدالة

    s = _strip_nbsp(value)  # تعيين قيمة لمتغير/إعداد
    s = re.sub(r"\s+", "", s)  # تعيين قيمة لمتغير/إعداد
    if not s or s.lower() in {"nan", "none", "null"}:  # شرط (If)
        return None, False  # إرجاع قيمة من الدالة

    # Year-only formats (day+month unknown) => end of year
    m = re.fullmatch(r"00[-/ ]00[-/ ](\d{4})", s)  # تعيين قيمة لمتغير/إعداد
    if m:  # شرط (If)
        y = int(m.group(1))  # تعيين قيمة لمتغير/إعداد
        return date(y, 12, 31), True  # إرجاع قيمة من الدالة

    m = re.fullmatch(r"(\d{4})[-/ ]00[-/ ]00", s)  # تعيين قيمة لمتغير/إعداد
    if m:  # شرط (If)
        y = int(m.group(1))  # تعيين قيمة لمتغير/إعداد
        return date(y, 12, 31), True  # إرجاع قيمة من الدالة

    # Year + month, unknown day => last day of that month
    m = re.fullmatch(r"(\d{4})[-/ ](\d{1,2})[-/ ]00", s)  # تعيين قيمة لمتغير/إعداد
    if m:  # شرط (If)
        y = int(m.group(1))  # تعيين قيمة لمتغير/إعداد
        mo = int(m.group(2))  # تعيين قيمة لمتغير/إعداد
        if 1 <= mo <= 12:  # شرط (If)
            last_day = calendar.monthrange(y, mo)[1]  # تعيين قيمة لمتغير/إعداد
            return date(y, mo, last_day), True  # إرجاع قيمة من الدالة
        return None, False  # إرجاع قيمة من الدالة

    d = clean_date(value)  # تعيين قيمة لمتغير/إعداد
    return (d, False) if d else (None, False)  # إرجاع قيمة من الدالة

def parse_repeater_flag(value):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Excel column 'معيد' contains the word 'معيد' for repeaters; otherwise empty."""
    if value is None:  # شرط (If)
        return False  # إرجاع قيمة من الدالة
    s = _strip_nbsp(value)  # تعيين قيمة لمتغير/إعداد
    if not s:  # شرط (If)
        return False  # إرجاع قيمة من الدالة
    s2 = re.sub(r"\s+", "", s)  # تعيين قيمة لمتغير/إعداد
    return (s in {"معيد", "نعم", "oui"} or s2 in {"1", "true", "True"} or ("معيد" in s))  # إرجاع قيمة من الدالة


def parse_assumed_flag(value):
    """حوّل قيمة عمود مفترض إلى Boolean آمن.

    مهم جداً: PostgreSQL لا يقبل NULL في هذا الحقل، لذلك أي خلية فارغة
    أو قيمة غير واضحة تُحفظ False بدل None.
    """
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    s = _strip_nbsp(value)
    if not s:
        return False
    s2 = re.sub(r"\s+", "", s).lower()
    if s2 in {"1", "true", "yes", "oui", "نعم", "صح", "مفترض"}:
        return True
    if s2 in {"0", "false", "no", "non", "لا", "خطأ"}:
        return False
    return False


def _ensure_assumed_default(data: Dict[str, Any], has_assumed_field: bool) -> None:
    """اضمن أن حقل مفترض لا يبقى None قبل المعاينة أو الحفظ."""
    if not has_assumed_field:
        return
    data["مفترض"] = parse_assumed_flag(data.get("مفترض"))


def _normalize_evening_training_type_for_import(model_cls, data: Dict[str, Any]) -> None:
    """صنّف كل سطر مسائي/معابر تلقائياً أثناء الاستيراد المختلط."""
    if getattr(model_cls, "__name__", "") != "مسائي_ومعابر":
        return
    training_type = detect_evening_training_type(data)
    data["نوع_التكوين"] = training_type
    if data.get("التخصص"):
        data["التخصص"] = clean_crossing_specialty_label(data.get("التخصص"))
    if data.get("السداسي"):
        data["السداسي"] = clamp_semester_for_evening_type(data.get("السداسي"), training_type)


def _norm_header(s: str) -> str:  # تعريف دالة (Function)
    # توحيد أسماء رؤوس الأعمدة حتى لا يفشل الاستيراد بسبب همزة/مسافة بسيطة،
    # مثل: الإسم/الاسم، الأجنبية/الاجنبية، رقم/ م-الشطب/رقم/م-الشطب.
    value = str(s or "").replace("\u00A0", " ").replace("ـ", "").strip()
    value = re.sub(r"\s+", "", value)
    value = value.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ى", "ي")
    return value


def _has_user_value(value) -> bool:
    if value is None:
        return False
    return bool(_strip_nbsp(value))


# رؤوس بديلة نقبلها أثناء الاستيراد حتى لا يفشل Excel القديم بسبب اختلاف اسم العمود فقط.
# مثال: العمود القديم "الرقم" يعادل العمود الرسمي الحالي "الرقم التعريفي".
_IMPORT_HEADER_ALIASES = {
    _norm_header("الرقم"): "الرقم_التعريفي",
    _norm_header("رقم"): "الرقم_التعريفي",
    _norm_header("#"): "الرقم_التعريفي",
    _norm_header("نوع التكوين"): "نوع_التكوين",
    _norm_header("نوع_التكوين"): "نوع_التكوين",
    _norm_header("نمط التكوين"): "نوع_التكوين",
}

# هذه العناوين تعني أن اللقب والاسم موجودان في عمود واحد، وسنقسمهما تلقائياً.
_COMBINED_NAME_HEADER_KEYS = {
    _norm_header("اللقب والاسم"),
    _norm_header("اللقب والإسم"),
    _norm_header("اللقب و الإسم"),
    _norm_header("اللقب و الاسم"),
    _norm_header("اللقب و الاسم الكامل"),
    _norm_header("الاسم واللقب"),
    _norm_header("الإسم واللقب"),
    _norm_header("الاسم و اللقب"),
    _norm_header("الإسم و اللقب"),
    _norm_header("الاسم الكامل"),
}

# حقول جديدة/اختيارية لا يجب أن تمنع الاستيراد عند غيابها من ملفات Excel القديمة.
_OPTIONAL_IMPORT_MISSING_FIELDS = {
    "بلدية_الإقامة_بالعربية",
    "نوع_التكوين",
}

_COMPOUND_SURNAME_PREFIXES = {
    _norm_header("بن"),
    _norm_header("ابن"),
    _norm_header("بنت"),
    _norm_header("ولد"),
    _norm_header("أولاد"),
    _norm_header("اولاد"),
    _norm_header("آل"),
}


def _clean_name_text(value) -> str:
    text = _strip_nbsp(value)
    text = text.replace("،", " ").replace(",", " ").replace("/", " ").replace("\\", " ")
    text = text.replace("-", " ").replace("_", " ")
    return re.sub(r"\s+", " ", text).strip()


def _strip_latin_accents(value: str) -> str:
    try:
        import unicodedata
        return "".join(
            ch for ch in unicodedata.normalize("NFKD", value)
            if not unicodedata.combining(ch)
        )
    except Exception:
        return value


def _norm_name_token_for_match(value: str) -> str:
    value = str(value or "").replace("\u00A0", " ").replace("ـ", " ").strip()
    value = _strip_latin_accents(value)
    value = value.lower()
    value = value.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ٱ", "ا")
    value = value.replace("ى", "ي").replace("ة", "ه")
    value = value.replace("ؤ", "و").replace("ئ", "ي")
    value = re.sub(r"[^0-9a-zA-Z\u0600-\u06FF]+", "", value)
    return value


def _name_tokens(value) -> List[str]:
    text = _clean_name_text(value)
    return [p for p in text.split(" ") if p]


def _build_surname_alias_candidates():
    generic_prefixes = {
        "بن", "ابن", "بنت", "ولد", "اولاد", "أولاد", "آل", "ايت", "أيت",
        "ben", "bent", "ould", "ouled", "ait", "aït",
    }
    generic_norm = {_norm_name_token_for_match(x) for x in generic_prefixes}
    seen = set()
    candidates = []

    for key, aliases in (SURNAME_ALIAS_MAP or {}).items():
        values = [key]
        if isinstance(aliases, (list, tuple, set)):
            values.extend(aliases)
        for alias in values:
            parts = _name_tokens(alias)
            if not parts:
                continue
            norm_parts = [_norm_name_token_for_match(part) for part in parts]
            norm_parts = [part for part in norm_parts if part]
            if not norm_parts:
                continue
            joined = "".join(norm_parts)
            # لا نستعمل البادئات العامة وحدها من القائمة، لأنها تقسم "بن علي محمد" إلى "بن" فقط.
            if len(norm_parts) == 1 and joined in generic_norm:
                continue
            if joined in seen:
                continue
            seen.add(joined)
            candidates.append((joined, len(joined)))

    # الأطول أولاً حتى يطابق "بن علي" قبل "بن"، و"ولد سيدي محمد" قبل "ولد سيدي".
    candidates.sort(key=lambda item: item[1], reverse=True)
    return candidates


_SURNAME_ALIAS_CANDIDATES = _build_surname_alias_candidates()


def _split_by_surname_alias(parts: List[str]) -> Tuple[str, str]:
    if len(parts) < 2 or not _SURNAME_ALIAS_CANDIDATES:
        return "", ""

    normalized_parts = [_norm_name_token_for_match(part) for part in parts]
    normalized_parts = [part for part in normalized_parts if part]
    full_joined = "".join(normalized_parts)
    if not full_joined:
        return "", ""

    for alias_joined, alias_len in _SURNAME_ALIAS_CANDIDATES:
        if not full_joined.startswith(alias_joined):
            continue

        consumed = 0
        matched_count = 0
        for idx, part in enumerate(normalized_parts, start=1):
            consumed += len(part)
            if consumed == alias_len:
                matched_count = idx
                break
            if consumed > alias_len:
                break

        if matched_count and matched_count < len(parts):
            return " ".join(parts[:matched_count]).strip(), " ".join(parts[matched_count:]).strip()

    return "", ""


def split_arabic_full_name(value) -> Tuple[str, str]:
    """
    فصل عمود "اللقب والاسم" عند الاستيراد.

    المنطق الجديد:
    1) يبحث أولاً في قائمة الألقاب المصححة SURNAME_ALIAS_MAP من الأطول إلى الأقصر.
       مثال: "بن علي محمد" => اللقب "بن علي" والاسم "محمد".
    2) إذا لم يجد لقباً مطابقاً، يستعمل قاعدة البادئات العامة بن/بنت/ولد/أولاد/آل/أيت.
    3) إذا لم تنجح القواعد السابقة، يستعمل أول كلمة كلقب والباقي اسماً.
       مثال: "ربحي صدام حسين" => اللقب "ربحي" والاسم "صدام حسين".
    """
    if value is None:
        return "", ""
    text = _clean_name_text(value)
    if not text:
        return "", ""

    parts = text.split(" ")
    if len(parts) == 1:
        return parts[0], ""

    alias_last, alias_first = _split_by_surname_alias(parts)
    if alias_last:
        return alias_last, alias_first

    first_key = _norm_header(parts[0])
    if first_key in _COMPOUND_SURNAME_PREFIXES and len(parts) >= 3:
        return " ".join(parts[:2]).strip(), " ".join(parts[2:]).strip()

    return parts[0].strip(), " ".join(parts[1:]).strip()

def _header_to_field_map(model_cls, include_assumed: bool = False) -> Dict[str, str]:
    mapping = {_norm_header(vn): name for name, vn in expected_columns_for_model(model_cls, include_assumed=include_assumed)}
    allowed_field_names = {name for name, _ in expected_columns_for_model(model_cls, include_assumed=True)}
    model_field_names = {field.name for field in model_cls._meta.fields}

    for alias_key, field_name in _IMPORT_HEADER_ALIASES.items():
        if field_name in allowed_field_names or field_name in model_field_names:
            mapping[alias_key] = field_name

    for key in _COMBINED_NAME_HEADER_KEYS:
        mapping[key] = "__combined_full_name__"

    return mapping


def _apply_combined_name_split(data: Dict[str, Any]) -> None:
    raw_full_name = data.pop("__combined_full_name__", None)
    if not _has_user_value(raw_full_name):
        return

    last_name, first_name = split_arabic_full_name(raw_full_name)
    if last_name and not _has_user_value(data.get("اللقب")):
        data["اللقب"] = last_name
    if first_name and not _has_user_value(data.get("الاسم")):
        data["الاسم"] = first_name


def _present_fields_from_headers(model_cls, headers: List[str]) -> Tuple[set, set]:
    mapping = _header_to_field_map(model_cls, include_assumed=True)
    present_fields = set()
    present_header_keys = set()

    for header in headers:
        key = _norm_header(header)
        if not key:
            continue
        present_header_keys.add(key)
        field_name = mapping.get(key)
        if field_name == "__combined_full_name__":
            present_fields.update({"اللقب", "الاسم"})
        elif field_name:
            present_fields.add(field_name)

    return present_fields, present_header_keys


def _append_row_error(errors: List[str], row_index: int, row_identity: str, message: str) -> None:
    prefix = f"السطر {row_index}" + (f" — {row_identity}" if row_identity else "")
    errors.append(f"{prefix}: {message}")


def _normalize_remaining_date_fields(model_cls, data: Dict[str, Any], row_index: int, row_identity: str, errors: List[str]) -> bool:
    """تحويل أي DateField إضافي في النموذج إلى قيمة تاريخ صحيحة قبل الحفظ الجماعي."""
    valid = True
    handled = {"تاريخ_الميلاد", "تاريخ_بداية_التكوين", "تاريخ_نهاية_التكوين", "تاريخ_التكوين_السابق_للمعيدين"}

    for field in model_cls._meta.fields:
        if not isinstance(field, DateField):
            continue

        field_name = field.name
        if field_name in handled:
            continue
        if field_name not in data:
            continue

        raw = data.get(field_name)
        if not _has_user_value(raw):
            data[field_name] = None
            continue

        parsed = clean_date(raw)
        if parsed is None:
            _append_row_error(errors, row_index, row_identity, f"{getattr(field, 'verbose_name', field_name)} غير صالح: {raw}")
            valid = False
            data[field_name] = None
        else:
            data[field_name] = parsed

    return valid


def _validate_row_dates(data: Dict[str, Any], row_index: int, row_identity: str, errors: List[str]) -> bool:
    valid = True

    raw_birthdate = data.get("تاريخ_الميلاد")
    if _has_user_value(raw_birthdate):
        bd, assumed = parse_birthdate_with_assumed(raw_birthdate)
        if bd is None:
            _append_row_error(errors, row_index, row_identity, f"تاريخ الميلاد غير صالح: {raw_birthdate}")
            valid = False
        else:
            data["تاريخ_الميلاد"] = bd
            if "مفترض" in {f.name for f in data.get('_model_fields', [])}:
                data["مفترض"] = assumed
    else:
        data["تاريخ_الميلاد"] = None

    for field_name, label in [("تاريخ_بداية_التكوين", "تاريخ بداية التكوين"), ("تاريخ_نهاية_التكوين", "تاريخ نهاية التكوين")]:
        raw = data.get(field_name)
        if not _has_user_value(raw):
            data[field_name] = None
            continue
        parsed = clean_date(raw)
        if parsed is None:
            _append_row_error(errors, row_index, row_identity, f"{label} غير صالح: {raw}")
            valid = False
        data[field_name] = parsed

    start_date_value = data.get("تاريخ_بداية_التكوين")
    end_date_value = data.get("تاريخ_نهاية_التكوين")
    if start_date_value and end_date_value and end_date_value < start_date_value:
        _append_row_error(errors, row_index, row_identity, "تاريخ نهاية التكوين أقدم من تاريخ بداية التكوين.")
        valid = False
    raw_prev_training = data.get("تاريخ_التكوين_السابق_للمعيدين")
    if "تاريخ_التكوين_السابق_للمعيدين" in data:
        if not _has_user_value(raw_prev_training):
            data["تاريخ_التكوين_السابق_للمعيدين"] = None
        else:
            parsed_prev_training = clean_date(raw_prev_training)
            if parsed_prev_training is None:
                _append_row_error(
                    errors,
                    row_index,
                    row_identity,
                    f"تاريخ التكوين السابق للمعيدين غير صالح: {raw_prev_training}",
                )
                valid = False
            else:
                data["تاريخ_التكوين_السابق_للمعيدين"] = parsed_prev_training
    return valid


def _required_core_fields(data: Dict[str, Any]) -> List[str]:
    required = []
    checks = [
        ("اللقب", "اللقب"),
        ("الاسم", "الاسم"),
        ("التخصص", "التخصص"),
        ("رقم_التسجيل", "رقم التسجيل"),
        ("تاريخ_بداية_التكوين", "تاريخ بداية التكوين"),
        ("تاريخ_نهاية_التكوين", "تاريخ نهاية التكوين"),
    ]
    for field_name, label in checks:
        value = data.get(field_name)
        if isinstance(value, str):
            value = value.strip()
        if not value:
            required.append(label)
    return required


def expected_columns_for_model(model_cls, include_assumed: bool = False) -> List[Tuple[str, str]]:  # تعريف دالة (Function)
    """
    ترتيب وأسماء أعمدة Excel حسب النمط، مطابق لملف اعمدة البرنامج.xlsx.

    هذه الدالة تُستعمل في:
    - قالب الاستيراد.
    - فحص رؤوس الأعمدة قبل الاستيراد.
    - ربط رؤوس Excel بأسماء حقول قاعدة البيانات.
    """
    return excel_columns_for_model(model_cls, include_assumed=include_assumed)



def validate_headers(model_cls, headers: List[str]) -> Tuple[bool, List[str], List[str]]:  # تعريف دالة (Function)
    expected = expected_columns_for_model(model_cls, include_assumed=False)  # تعيين قيمة لمتغير/إعداد
    present_fields, _present_header_keys = _present_fields_from_headers(model_cls, headers)

    missing = []
    for field_name, label in expected:
        if field_name in _OPTIONAL_IMPORT_MISSING_FIELDS:
            continue
        if field_name not in present_fields:
            missing.append(label)

    official_header_keys = {_norm_header(label) for _, label in expected_columns_for_model(model_cls, include_assumed=True)}
    alias_header_keys = set(_IMPORT_HEADER_ALIASES.keys()) | set(_COMBINED_NAME_HEADER_KEYS)
    allowed_header_keys = official_header_keys | alias_header_keys
    extra = [h for h in headers if _norm_header(h) and _norm_header(h) not in allowed_header_keys]  # تعيين قيمة لمتغير/إعداد

    ok = (len(missing) == 0)  # تعيين قيمة لمتغير/إعداد
    return ok, missing, extra  # إرجاع قيمة من الدالة


def export_template_xlsx(model_cls) -> HttpResponse:  # تعريف دالة (Function)
    wb = Workbook()  # تعيين قيمة لمتغير/إعداد
    ws = wb.active  # تعيين قيمة لمتغير/إعداد
    ws.title = "قالب"  # تعيين قيمة لمتغير/إعداد
    headers = [vn for _, vn in expected_columns_for_model(model_cls, include_assumed=True)]  # تعيين قيمة لمتغير/إعداد
    ws.append(headers)  # سطر كود لتنفيذ منطق/إعداد
    # Make header row a bit wider
    for i, h in enumerate(headers, start=1):  # حلقة تكرار (For)
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(h) + 4))  # تعيين قيمة لمتغير/إعداد
    bio = bytes_io = None  # تعيين قيمة لمتغير/إعداد
    from io import BytesIO  # استيراد عناصر محددة من مكتبة/وحدة
    bytes_io = BytesIO()  # تعيين قيمة لمتغير/إعداد
    wb.save(bytes_io)  # سطر كود لتنفيذ منطق/إعداد
    bytes_io.seek(0)  # سطر كود لتنفيذ منطق/إعداد
    filename = f"template_{model_cls._meta.model_name}.xlsx"  # تعيين قيمة لمتغير/إعداد
    resp = HttpResponse(  # تعيين قيمة لمتغير/إعداد
        bytes_io.getvalue(),  # سطر كود لتنفيذ منطق/إعداد
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # تعيين قيمة لمتغير/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'  # تعيين قيمة لمتغير/إعداد
    return resp  # إرجاع قيمة من الدالة


def _uploaded_excel_extension(uploaded_file) -> str:
    name = str(getattr(uploaded_file, "name", "") or "").strip()
    return os.path.splitext(name)[1].lower()


def _validate_uploaded_excel_file(uploaded_file) -> Tuple[bool, str]:
    ext = _uploaded_excel_extension(uploaded_file)
    if ext in SUPPORTED_EXCEL_IMPORT_EXTENSIONS:
        return True, ""
    if ext in {".xls", ".xlsb", ".csv", ".ods"}:
        return False, (
            f"الصيغة {ext.upper()} غير مدعومة مباشرة في هذا الاستيراد. "
            f"احفظ الملف من Excel بصيغة {SUPPORTED_EXCEL_IMPORT_EXTENSIONS_TEXT} ثم أعد المحاولة."
        )
    return False, f"يرجى اختيار ملف Excel بصيغة {SUPPORTED_EXCEL_IMPORT_EXTENSIONS_TEXT}."


def _save_temp_upload(uploaded_file) -> str:  # تعريف دالة (Function)
    ok, error_message = _validate_uploaded_excel_file(uploaded_file)
    if not ok:
        raise ValueError(error_message)

    base_dir = os.path.join(settings.BASE_DIR, "tmp_imports")  # تعيين قيمة لمتغير/إعداد
    os.makedirs(base_dir, exist_ok=True)  # سطر كود لتنفيذ منطق/إعداد
    ext = _uploaded_excel_extension(uploaded_file) or ".xlsx"
    temp_id = f"{uuid.uuid4().hex}{ext}"  # تعيين قيمة لمتغير/إعداد
    path = os.path.join(base_dir, temp_id)  # تعيين قيمة لمتغير/إعداد
    with open(path, "wb") as f:  # سطر كود لتنفيذ منطق/إعداد
        for chunk in uploaded_file.chunks():  # حلقة تكرار (For)
            f.write(chunk)  # سطر كود لتنفيذ منطق/إعداد
    return temp_id  # إرجاع قيمة من الدالة


def _temp_path(temp_id: str) -> str:  # تعريف دالة (Function)
    safe_temp_id = os.path.basename(str(temp_id or "").strip())
    base_dir = os.path.join(settings.BASE_DIR, "tmp_imports")
    if os.path.splitext(safe_temp_id)[1].lower() in SUPPORTED_EXCEL_IMPORT_EXTENSIONS:
        return os.path.join(base_dir, safe_temp_id)
    # توافق مع الملفات المؤقتة القديمة التي كانت تحفظ دائماً بصيغة .xlsx
    return os.path.join(base_dir, f"{safe_temp_id}.xlsx")  # إرجاع قيمة من الدالة


def _load_import_workbook(temp_id: str):
    try:
        return load_workbook(_temp_path(temp_id), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise ValueError(
            f"تعذر فتح الملف. الصيغ المدعومة هي: {SUPPORTED_EXCEL_IMPORT_EXTENSIONS_TEXT}. "
            "إذا كان الملف بصيغة XLS قديمة، افتحه في Excel ثم احفظه بصيغة XLSX أو XLSM."
        ) from exc


def _read_sheets(temp_id: str) -> List[str]:  # تعريف دالة (Function)
    wb = _load_import_workbook(temp_id)  # تعريف مسار URL في Django
    try:
        return list(wb.sheetnames)  # إرجاع قيمة من الدالة
    finally:
        wb.close()


def _read_headers(temp_id: str, sheet_name: str) -> List[str]:  # تعريف دالة (Function)
    wb = _load_import_workbook(temp_id)  # تعريف مسار URL في Django
    try:
        ws = wb[sheet_name]  # تعيين قيمة لمتغير/إعداد
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):  # حلقة تكرار (For)
            if row and any(v is not None and str(v).strip() for v in row):  # شرط (If)
                return [str(v).strip() if v is not None else "" for v in row]  # إرجاع قيمة من الدالة
        return []  # إرجاع قيمة من الدالة
    finally:
        wb.close()


def _iter_rows(temp_id: str, sheet_name: str) -> Iterable[Dict[str, Any]]:  # تعريف دالة (Function)
    wb = _load_import_workbook(temp_id)  # تعريف مسار URL في Django
    try:
        ws = wb[sheet_name]  # تعيين قيمة لمتغير/إعداد
        header = None  # تعيين قيمة لمتغير/إعداد
        for row in ws.iter_rows(values_only=True):  # حلقة تكرار (For)
            if header is None:  # شرط (If)
                if row and any(v is not None and str(v).strip() for v in row):  # شرط (If)
                    header = [str(v).strip() if v is not None else "" for v in row]  # تعيين قيمة لمتغير/إعداد
                continue  # سطر كود لتنفيذ منطق/إعداد
            if not row or all(v is None or str(v).strip() == "" for v in row):  # شرط (If)
                continue  # سطر كود لتنفيذ منطق/إعداد
            yield {header[i]: row[i] if i < len(row) else None for i in range(len(header))}  # سطر كود لتنفيذ منطق/إعداد
    finally:
        wb.close()


def _map_row_to_fields(model_cls, row: Dict[str, Any]) -> Dict[str, Any]:  # تعريف دالة (Function)
    # map by verbose_name
    vn_to_name = _header_to_field_map(model_cls, include_assumed=False)  # تعيين قيمة لمتغير/إعداد
    data: Dict[str, Any] = {}  # تعيين قيمة لمتغير/إعداد
    for vn, val in row.items():  # حلقة تكرار (For)
        key = _norm_header(vn)  # تعيين قيمة لمتغير/إعداد
        if key in vn_to_name:  # شرط (If)
            field_name = vn_to_name[key]  # تعيين قيمة لمتغير/إعداد
            data[field_name] = val  # تعيين قيمة لمتغير/إعداد
    _apply_combined_name_split(data)

    for ignored_field in IGNORED_IMPORT_FIELDS:
        data.pop(ignored_field, None)

    # unify status if present
    if "الحالة" in data:  # شرط (If)
        data["الحالة"] = status_group(data.get("الحالة"))  # تعيين قيمة لمتغير/إعداد
    return data  # إرجاع قيمة من الدالة


@dataclass
class ImportPreviewRow:
    row_index: int
    display_name: str
    registration_number: str
    specialty: str
    start_date: Any
    end_date: Any


def _build_import_key(data: Dict[str, Any]):
    ident = str(data.get("الرقم_التعريفي") or "").strip()
    reg_no = normalize_registration_number(data.get("رقم_التسجيل"))
    data["رقم_التسجيل"] = reg_no
    last_name = str(data.get("اللقب") or "").strip()
    first_name = str(data.get("الاسم") or "").strip()
    specialty = str(data.get("التخصص") or "").strip()
    start_date_value = data.get("تاريخ_بداية_التكوين")
    end_date_value = data.get("تاريخ_نهاية_التكوين")

    if reg_no:
        return ("reg", reg_no)
    if ident:
        return ("ident", ident)
    return ("fallback", last_name, first_name, specialty, start_date_value, end_date_value)


def preview_sheet(model_cls, temp_id: str, sheet_name: str, max_preview_rows: int = 12) -> Dict[str, Any]:
    wb = _load_import_workbook(temp_id)
    try:
        if sheet_name not in wb.sheetnames:
            return {"ok": False, "errors": ["SHEET_NOT_FOUND"]}

        ws = wb[sheet_name]
        header: List[str] = []
        header_row_num = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row and any(v is not None and str(v).strip() for v in row):
                header = [str(v).strip() if v is not None else "" for v in row]
                header_row_num = r_idx
                break

        if not header or header_row_num is None:
            return {"ok": False, "errors": ["EMPTY_SHEET"]}

        ok, missing, extra = validate_headers(model_cls, header)
        if not ok:
            return {"ok": False, "errors": ["HEADER_MISMATCH"], "missing": missing, "extra": extra, "headers": header}

        vn_to_name = _header_to_field_map(model_cls, include_assumed=True)
        model_field_names = {f.name for f in model_cls._meta.fields}
        has_assumed_field = ("مفترض" in model_field_names)
        preview_rows: List[ImportPreviewRow] = []
        errors: List[str] = []
        seen_keys = set()
        total_rows = 0
        valid_rows = 0
        invalid_rows = 0
        duplicate_rows = 0

        for row_index, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=header_row_num + 1):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            total_rows += 1
            data: Dict[str, Any] = {}
            for i, cell in enumerate(row):
                if i >= len(header):
                    break
                hn = _norm_header(header[i])
                if not hn:
                    continue
                fn = vn_to_name.get(hn)
                if fn:
                    data[fn] = cell

            _apply_combined_name_split(data)

            # تجاهل الحقول المحسوبة/الداخلية حتى لو كانت موجودة في ملف Excel
            for ignored_field in IGNORED_IMPORT_FIELDS:
                data.pop(ignored_field, None)

            if "الحالة" in data:
                data["الحالة"] = status_group(data.get("الحالة"))
            if "معيد" in data:
                data["معيد"] = parse_repeater_flag(data.get("معيد"))
            if "تاريخ_التكوين_السابق_للمعيدين" in data and not data.get("معيد", False):
                data["تاريخ_التكوين_السابق_للمعيدين"] = None

            row_identity = _row_identity_display(data.get("اللقب"), data.get("الاسم"), data.get("رقم_التسجيل"))
            row_valid = _validate_row_dates(data, row_index, row_identity, errors)
            if row_valid:
                row_valid = _normalize_remaining_date_fields(model_cls, data, row_index, row_identity, errors)

            if has_assumed_field:
                if "تاريخ_الميلاد" in data and data.get("تاريخ_الميلاد") is not None:
                    _, assumed = parse_birthdate_with_assumed(data.get("تاريخ_الميلاد"))
                    data["مفترض"] = assumed
                _ensure_assumed_default(data, has_assumed_field)

            _normalize_evening_training_type_for_import(model_cls, data)

            missing_core = _required_core_fields(data)
            if missing_core:
                _append_row_error(errors, row_index, row_identity, "بيانات ناقصة في الحقول الأساسية: " + ", ".join(missing_core))
                row_valid = False

            key = _build_import_key(data)
            if key in seen_keys:
                _append_row_error(errors, row_index, row_identity, "هذا السجل مكرر داخل نفس ملف Excel.")
                duplicate_rows += 1
                row_valid = False
            else:
                seen_keys.add(key)

            if row_valid:
                valid_rows += 1
            else:
                invalid_rows += 1

            if len(preview_rows) < max_preview_rows:
                preview_rows.append(ImportPreviewRow(
                    row_index=row_index,
                    display_name=" ".join([p for p in [str(data.get("الاسم") or "").strip(), str(data.get("اللقب") or "").strip()] if p]),
                    registration_number=str(data.get("رقم_التسجيل") or "").strip(),
                    specialty=str(data.get("التخصص") or "").strip(),
                    start_date=data.get("تاريخ_بداية_التكوين") or "",
                    end_date=data.get("تاريخ_نهاية_التكوين") or "",
                ))

        return {
            "ok": True,
            "headers": header,
            "total_rows": total_rows,
            "valid_rows": valid_rows,
            "invalid_rows": invalid_rows,
            "duplicate_rows": duplicate_rows,
            "preview_rows": preview_rows,
            "errors": errors,
            "error_count": len(errors),
        }
    finally:
        wb.close()

def import_sheet(model_cls, temp_id: str, sheet_name: str) -> Tuple[int, List[str]]:  # تعريف دالة (Function)
    """
    استيراد سريع مع تجهيز الدفعات أثناء الاستيراد فقط.
    """
    from django.db import transaction, connection  # استيراد عناصر محددة من مكتبة/وحدة

    created = 0
    errors: List[str] = []

    try:
        with connection.cursor() as c:
            c.execute("PRAGMA journal_mode=WAL;")
            c.execute("PRAGMA synchronous=NORMAL;")
            c.execute("PRAGMA temp_store=MEMORY;")
            c.execute("PRAGMA cache_size=-200000;")
    except Exception:
        pass

    def _promotion_name(session_no: int) -> str:
        return "فيفري" if session_no == 1 else "سبتمبر"

    def _promotion_sort_key(key):
        year_value, session_no = key
        return (year_value, 0 if session_no == 1 else 1)

    def _max_semesters_for_row(row_data, start_date_value, end_date_value):
        model_name = getattr(model_cls, "__name__", "")
        months = 0
        if start_date_value and end_date_value:
            months = (end_date_value.year - start_date_value.year) * 12 + (end_date_value.month - start_date_value.month)
            if end_date_value.day >= start_date_value.day:
                months += 1
        joined = " ".join([
            str(row_data.get("النظام") or ""),
            str(row_data.get("التخصص") or ""),
            str(row_data.get("الحالة") or ""),
        ])
        if model_name == "مسائي_ومعابر" and ("معابر" in joined or "معبر" in joined or months <= 14):
            return 2
        if model_name == "تمهين" and months >= 34:
            return 6
        return 5

    def _semester_label(number):
        labels = {1: "الأول", 2: "الثاني", 3: "الثالث", 4: "الرابع", 5: "الخامس", 6: "السادس"}
        return labels.get(number, labels.get(min(number, 6), "الأول"))

    wb = _load_import_workbook(temp_id)
    try:
        if sheet_name not in wb.sheetnames:
            return 0, ["SHEET_NOT_FOUND"]
        ws = wb[sheet_name]

        header: List[str] = []
        header_row_num = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row and any(v is not None and str(v).strip() for v in row):
                header = [str(v).strip() if v is not None else "" for v in row]
                header_row_num = r_idx
                break
        if not header or header_row_num is None:
            return 0, ["EMPTY_SHEET"]

        ok, missing, extra = validate_headers(model_cls, header)
        if not ok:
            errors.append("HEADER_MISMATCH")
            return 0, errors

        vn_to_name = _header_to_field_map(model_cls, include_assumed=True)
        field_names = {f.name for f in model_cls._meta.fields}
        has_assumed_field = ("مفترض" in field_names)
        has_promotion_field = ("الدفعة" in field_names)

        rows_cache: List[Tuple[int, Dict[str, Any], tuple]] = []
        promotions_info: Dict[tuple, Dict[str, Any]] = {}

        for row_index, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=header_row_num + 1):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            data: Dict[str, Any] = {}
            for i, cell in enumerate(row):
                if i >= len(header):
                    break
                hn = _norm_header(header[i])
                if not hn:
                    continue
                fn = vn_to_name.get(hn)
                if not fn:
                    continue
                data[fn] = cell

            _apply_combined_name_split(data)

            # تجاهل الحقول المحسوبة/الداخلية حتى لو كانت موجودة في ملف Excel
            for ignored_field in IGNORED_IMPORT_FIELDS:
                data.pop(ignored_field, None)

            if "الحالة" in data:
                data["الحالة"] = status_group(data.get("الحالة"))
            if "معيد" in data:
                data["معيد"] = parse_repeater_flag(data.get("معيد"))
            if "تاريخ_التكوين_السابق_للمعيدين" in data and not data.get("معيد", False):
                data["تاريخ_التكوين_السابق_للمعيدين"] = None
            # لا نترك حقل "مفترض" فارغاً أبداً، حتى إذا كان تاريخ الميلاد فارغاً
            # أو كان عمود "مفترض" موجوداً في Excel وخليته فارغة.
            _ensure_assumed_default(data, has_assumed_field)

            last_name = str(data.get("اللقب") or "").strip()
            first_name = str(data.get("الاسم") or "").strip()
            specialty = str(data.get("التخصص") or "").strip()
            reg_no = normalize_registration_number(data.get("رقم_التسجيل"))
            data["رقم_التسجيل"] = reg_no
            start_date_value = data.get("تاريخ_بداية_التكوين")
            end_date_value = data.get("تاريخ_نهاية_التكوين")
            row_identity = _row_identity_display(last_name, first_name, reg_no)

            data["_model_fields"] = list(model_cls._meta.fields)
            if not _validate_row_dates(data, row_index, row_identity, errors):
                data.pop("_model_fields", None)
                continue
            data.pop("_model_fields", None)
            if not _normalize_remaining_date_fields(model_cls, data, row_index, row_identity, errors):
                continue

            # حماية ثانية بعد تحويل التواريخ، لأن بعض المسارات قد تعيد تعيين القيم.
            _ensure_assumed_default(data, has_assumed_field)
            _normalize_evening_training_type_for_import(model_cls, data)

            start_date_value = data.get("تاريخ_بداية_التكوين")
            end_date_value = data.get("تاريخ_نهاية_التكوين")
            missing_core = _required_core_fields(data)
            if missing_core:
                _append_row_error(errors, row_index, row_identity, "بيانات ناقصة في الحقول الأساسية: " + ", ".join(missing_core))
                continue

            promotion_key = None
            if has_promotion_field:
                session_no, year_value = resolve_session_year(reg_no, start_date_value)
                if session_no and year_value:
                    promotion_key = (year_value, session_no)
                    info = promotions_info.get(promotion_key)
                    if not info:
                        promotions_info[promotion_key] = {
                            "session_no": session_no,
                            "year": year_value,
                            "official_start": start_date_value,
                        }
                    else:
                        existing_start = info.get("official_start")
                        if (not existing_start) or (start_date_value and start_date_value < existing_start):
                            info["official_start"] = start_date_value
                else:
                    prefix = f"السطر {row_index}" + (f" — {row_identity}" if row_identity else "")
                    errors.append(f"{prefix}: تعذر استخراج رقم الدورة والسنة من رقم التسجيل: {reg_no}")

            ident = str(data.get("الرقم_التعريفي") or "").strip()
            if reg_no:
                key = ("reg", reg_no)
            elif ident:
                key = ("ident", ident)
            else:
                key = ("fallback", last_name, first_name, specialty, start_date_value, end_date_value)
            rows_cache.append((row_index, data, key))

        if not rows_cache:
            return 0, errors

        promotion_map: Dict[tuple, Any] = {}
        ordered_keys: List[tuple] = []
        if promotions_info:
            ordered_keys = sorted(promotions_info.keys(), key=_promotion_sort_key)
            existing_promotions = {
                (obj.السنة, obj.رقم_الدورة): obj
                for obj in دفعة.objects.filter(السنة__in={k[0] for k in ordered_keys}, رقم_الدورة__in=[1, 2])
            }

            to_create_promotions = []
            for key in ordered_keys:
                if key in existing_promotions:
                    promotion_map[key] = existing_promotions[key]
                    continue
                info = promotions_info[key]
                official_start = info["official_start"]
                obj = دفعة(
                    اسم_الدفعة=_promotion_name(info["session_no"]),
                    رقم_الدورة=info["session_no"],
                    السنة=info["year"],
                    تاريخ_الدخول_الرسمي=official_start,
                    بداية_السداسي_1=official_start,
                    بداية_السداسي_2=add_months(official_start, 6) if official_start else None,
                    بداية_السداسي_3=add_months(official_start, 12) if official_start else None,
                    بداية_السداسي_4=add_months(official_start, 18) if official_start else None,
                    بداية_السداسي_5=add_months(official_start, 24) if official_start else None,
                    مفعلة=True,
                )
                to_create_promotions.append(obj)

            if to_create_promotions:
                دفعة.objects.bulk_create(to_create_promotions, batch_size=500, ignore_conflicts=True)
                existing_promotions = {
                    (obj.السنة, obj.رقم_الدورة): obj
                    for obj in دفعة.objects.filter(السنة__in={k[0] for k in ordered_keys}, رقم_الدورة__in=[1, 2])
                }

            promotions_to_update = []
            for idx, key in enumerate(ordered_keys):
                obj = existing_promotions.get(key)
                if not obj:
                    continue
                info = promotions_info[key]
                starts = [info["official_start"]]
                for offset in range(1, 5):
                    next_idx = idx + offset
                    if next_idx < len(ordered_keys):
                        next_start = promotions_info[ordered_keys[next_idx]]["official_start"]
                    else:
                        next_start = add_months(starts[-1], 6) if starts[-1] else None
                    starts.append(next_start)
                changed = False
                desired_name = _promotion_name(info["session_no"])
                if obj.اسم_الدفعة != desired_name:
                    obj.اسم_الدفعة = desired_name
                    changed = True
                if obj.تاريخ_الدخول_الرسمي != starts[0]:
                    obj.تاريخ_الدخول_الرسمي = starts[0]
                    changed = True
                attrs = ["بداية_السداسي_1", "بداية_السداسي_2", "بداية_السداسي_3", "بداية_السداسي_4", "بداية_السداسي_5"]
                for pos, attr in enumerate(attrs):
                    if getattr(obj, attr) != starts[pos]:
                        setattr(obj, attr, starts[pos])
                        changed = True
                if not obj.مفعلة:
                    obj.مفعلة = True
                    changed = True
                if changed:
                    promotions_to_update.append(obj)
                promotion_map[key] = obj

            if promotions_to_update:
                دفعة.objects.bulk_update(
                    promotions_to_update,
                    ["اسم_الدفعة", "تاريخ_الدخول_الرسمي", "بداية_السداسي_1", "بداية_السداسي_2", "بداية_السداسي_3", "بداية_السداسي_4", "بداية_السداسي_5", "مفعلة"],
                    batch_size=200,
                )

            # بعد إنشاء/تعديل الدفعات، أعد بناء رزنامة السداسيات لكل الدفعات
            # حتى لا يعتمد الاستيراد على حقول قديمة أو ناقصة.
            refresh_all_promotion_semester_starts()
            existing_promotions = {
                (obj.السنة, obj.رقم_الدورة): obj
                for obj in دفعة.objects.filter(السنة__in={k[0] for k in ordered_keys}, رقم_الدورة__in=[1, 2])
            }
            promotion_map = {key: existing_promotions[key] for key in ordered_keys if key in existing_promotions}

        import_start_dates = [data.get("تاريخ_بداية_التكوين") for _, data, _ in rows_cache if data.get("تاريخ_بداية_التكوين")]
        cohort_starts = cohort_start_dates_for_model(model_cls) + import_start_dates
        cohort_starts_by_type = {}
        if getattr(model_cls, "__name__", "") == "مسائي_ومعابر":
            for training_type in {data.get("نوع_التكوين") for _, data, _ in rows_cache if data.get("نوع_التكوين")}:
                type_dates = [data.get("تاريخ_بداية_التكوين") for _, data, _ in rows_cache if data.get("نوع_التكوين") == training_type and data.get("تاريخ_بداية_التكوين")]
                cohort_starts_by_type[training_type] = cohort_start_dates_for_model(model_cls, training_type=training_type) + type_dates

        existing_keys = set()
        # أسرع من جلب كائنات Django كاملة: نقرأ القيم اللازمة فقط وبـ iterator.
        # هذا يحافظ على نفس منطق منع التكرار، لكنه أخف بكثير عند وجود آلاف المتكونين.
        existing_rows = model_cls.objects.values_list(
            "الرقم_التعريفي",
            "رقم_التسجيل",
            "اللقب",
            "الاسم",
            "التخصص",
            "تاريخ_بداية_التكوين",
            "تاريخ_نهاية_التكوين",
        ).iterator(chunk_size=2000)
        for ident_raw, reg_raw, last_raw, first_raw, specialty_raw, start_raw, end_raw in existing_rows:
            reg_no = normalize_registration_number(reg_raw)
            ident = str(ident_raw or "").strip()
            if reg_no:
                existing_keys.add(("reg", reg_no))
            elif ident:
                existing_keys.add(("ident", ident))
            else:
                existing_keys.add((
                    "fallback",
                    str(last_raw or "").strip(),
                    str(first_raw or "").strip(),
                    str(specialty_raw or "").strip(),
                    start_raw,
                    end_raw,
                ))

        to_create = []
        prepared_rows: List[Tuple[int, str, Dict[str, Any]]] = []
        seen_file_keys = {}
        for idx, data, key in rows_cache:
            row_identity = _row_identity_display(data.get("اللقب"), data.get("الاسم"), data.get("رقم_التسجيل"))
            prefix = f"السطر {idx}" + (f" — {row_identity}" if row_identity else "")

            if key in seen_file_keys:
                first_row = seen_file_keys[key]
                errors.append(f"{prefix}: هذا السجل مكرر داخل نفس الملف (ظهر أولاً في السطر {first_row}) وتم تجاهله.")
                continue

            if key in existing_keys:
                errors.append(f"{prefix}: هذا السجل موجود من قبل، تم تجاهله.")
                continue

            try:
                reg_no = str(data.get("رقم_التسجيل") or "").strip()
                start_date_value = data.get("تاريخ_بداية_التكوين")
                end_date_value = data.get("تاريخ_نهاية_التكوين")
                is_repeater = bool(data.get("معيد", False))
                session_no, year_value = resolve_session_year(reg_no, start_date_value)
                promotion_key = (year_value, session_no) if session_no and year_value else None
                if has_promotion_field and promotion_key in promotion_map:
                    data["الدفعة"] = promotion_map[promotion_key]
                if has_promotion_field:
                    promotion_obj = data.get("الدفعة") if isinstance(data.get("الدفعة"), دفعة) else None
                    row_cohort_starts = cohort_starts_by_type.get(data.get("نوع_التكوين"), cohort_starts)
                    raw = compute_semester_for_trainee(
                        promotion_obj,
                        start_date_value,
                        end_date_value,
                        is_repeater=is_repeater,
                        cohort_starts=row_cohort_starts,
                        original_end_date=data.get("تاريخ_التكوين_السابق_للمعيدين"),
                    )
                    data["السداسي"] = raw or _semester_label(1)
                    if getattr(model_cls, "__name__", "") == "مسائي_ومعابر":
                        data["السداسي"] = clamp_semester_for_evening_type(data.get("السداسي"), data.get("نوع_التكوين"))
                if getattr(model_cls, "__name__", "") == "تمهين" and not is_repeater:
                    data["تاريخ_التكوين_السابق_للمعيدين"] = None

                # هذه الحقول تُدار داخليًا ولا يجب أخذها خامًا من Excel
                if has_promotion_field and not isinstance(data.get("الدفعة"), دفعة):
                    data.pop("الدفعة", None)
                if "السداسي" in data and not isinstance(data.get("السداسي"), str):
                    data.pop("السداسي", None)

                # آخر حماية قبل إنشاء كائن Django، حتى لا يصل None إلى PostgreSQL.
                _ensure_assumed_default(data, has_assumed_field)

                row_payload = dict(data)
                to_create.append(model_cls(**row_payload))
                prepared_rows.append((idx, row_identity, row_payload))
                existing_keys.add(key)
                seen_file_keys[key] = idx
            except Exception as e:
                errors.append(f"{prefix}: خطأ أثناء التحضير: {e}")

        if not to_create:
            return 0, errors

        try:
            with transaction.atomic():
                model_cls.objects.bulk_create(to_create, batch_size=2000, ignore_conflicts=True)
                created = len(to_create)
        except Exception as e:
            errors.append(f"خطأ أثناء الحفظ الجماعي: {e}")
            created = 0
            for row_index, row_identity, payload in prepared_rows:
                prefix = f"السطر {row_index}" + (f" — {row_identity}" if row_identity else "")
                try:
                    _ensure_assumed_default(payload, has_assumed_field)
                    model_cls.objects.create(**payload)
                    created += 1
                except Exception as row_exc:
                    errors.append(f"{prefix}: تعذر الحفظ الفردي: {row_exc}")

        return created, errors
    finally:
        wb.close()


@dataclass  # سطر كود لتنفيذ منطق/إعداد
class ImportContext:  # تعريف كلاس (Class)
    model_label: str  # سطر كود لتنفيذ منطق/إعداد
    model_verbose: str  # سطر كود لتنفيذ منطق/إعداد
    import_url: str  # سطر كود لتنفيذ منطق/إعداد
    template_url: str  # سطر كود لتنفيذ منطق/إعداد
    delete_all_url: str  # سطر كود لتنفيذ منطق/إعداد


class ExcelImportAdminMixin:  # تعريف كلاس (Class)
    # سطر كود لتنفيذ منطق/إعداد
    """
    Add 'استيراد Excel' view to any ModelAdmin.  # سطر كود لتنفيذ منطق/إعداد
    # سطر كود لتنفيذ منطق/إعداد
    """
    change_list_template = "admin/change_list_with_import.html"  # تعيين قيمة لمتغير/إعداد

    def get_import_context(self, request: HttpRequest) -> ImportContext:  # تعريف دالة (Function)
        model_label = f"{self.model._meta.app_label}_{self.model._meta.model_name}"  # تعيين قيمة لمتغير/إعداد
        import_url = reverse(f"admin:{model_label}_import_excel")  # تعيين قيمة لمتغير/إعداد
        template_url = reverse(f"admin:{model_label}_export_template")  # تعيين قيمة لمتغير/إعداد
        delete_all_url = reverse(f"admin:{model_label}_delete_all")  # تعيين قيمة لمتغير/إعداد

        return ImportContext(  # إرجاع قيمة من الدالة
            model_label=model_label,  # تعيين قيمة لمتغير/إعداد
            model_verbose=str(self.model._meta.verbose_name_plural),  # تعيين قيمة لمتغير/إعداد
            import_url=import_url,  # تعيين قيمة لمتغير/إعداد
            template_url=template_url,  # تعيين قيمة لمتغير/إعداد
            delete_all_url=delete_all_url,  # تعيين قيمة لمتغير/إعداد
        )  # سطر كود لتنفيذ منطق/إعداد


    def changelist_view(self, request, extra_context=None):  # تعريف دالة (Function)
        extra_context = extra_context or {}  # تعيين قيمة لمتغير/إعداد
        ic = self.get_import_context(request)  # تعيين قيمة لمتغير/إعداد
        extra_context.update({"import_url": ic.import_url, "delete_all_url": ic.delete_all_url})  # سطر كود لتنفيذ منطق/إعداد
        return super().changelist_view(request, extra_context=extra_context)  # إرجاع قيمة من الدالة

    def get_urls(self):  # تعريف دالة (Function)
        urls = super().get_urls()  # تعيين قيمة لمتغير/إعداد
        model_label = f"{self.model._meta.app_label}_{self.model._meta.model_name}"  # تعيين قيمة لمتغير/إعداد
        custom = [  # تعيين قيمة لمتغير/إعداد
            path("import-excel/", self.admin_site.admin_view(self.import_excel_view), name=f"{model_label}_import_excel"),  # تعريف مسار URL في Django
            path("export-template/", self.admin_site.admin_view(self.export_template_view), name=f"{model_label}_export_template"),  # تعريف مسار URL في Django
            path("delete-all/", self.admin_site.admin_view(self.delete_all_view), name=f"{model_label}_delete_all"),  # تعريف مسار URL في Django
        ]  # سطر كود لتنفيذ منطق/إعداد
        return custom + urls  # إرجاع قيمة من الدالة

    def delete_all_view(self, request: HttpRequest) -> HttpResponse:  # تعريف دالة (Function)
        # سطر كود لتنفيذ منطق/إعداد
        """حذف كل البيانات من هذا القسم مع شاشة تأكيد."""
        if not self.has_delete_permission(request):  # شرط (If)
            return HttpResponseForbidden("ليس لديك صلاحية الحذف.")  # إرجاع قيمة من الدالة

        ctx = self.get_import_context(request).__dict__  # تعيين قيمة لمتغير/إعداد
        changelist_url = reverse(f"admin:{ctx['model_label']}_changelist")  # تعيين قيمة لمتغير/إعداد

        if request.method == "POST":  # شرط (If)
            if request.POST.get("confirm") == "yes":  # شرط (If)
                count = self.model.objects.count()  # تعيين قيمة لمتغير/إعداد
                self.model.objects.all().delete()  # سطر كود لتنفيذ منطق/إعداد
                audit_view_event(
                    request,
                    event_type="admin",
                    action="delete",
                    target_model=self.model._meta.label,
                    object_repr=self.model._meta.verbose_name_plural,
                    before_data={"count": count},
                    after_data={"count": 0},
                    changed_fields=["all_records"],
                    details=f"حذف جميع سجلات {ctx['model_verbose']} دفعة واحدة.",
                    program=self.model.__name__,
                )
                messages.success(request, f"تم حذف جميع السجلات ({count} سجل) من: {ctx['model_verbose']}.")  # إظهار رسالة للمستخدم (نجاح/خطأ) في Django
                return HttpResponseRedirect(changelist_url)  # إرجاع قيمة من الدالة

            messages.info(request, "تم إلغاء العملية.")  # إظهار رسالة للمستخدم (نجاح/خطأ) في Django
            return HttpResponseRedirect(changelist_url)  # إرجاع قيمة من الدالة

        return render(  # إرجاع قيمة من الدالة
            request,  # سطر كود لتنفيذ منطق/إعداد
            "admin/delete_all_confirm.html",  # سطر كود لتنفيذ منطق/إعداد
            {**ctx, "changelist_url": changelist_url, "count": self.model.objects.count()},  # سطر كود لتنفيذ منطق/إعداد
        )  # سطر كود لتنفيذ منطق/إعداد

    def export_template_view(self, request: HttpRequest) -> HttpResponse:  # تعريف دالة (Function)
        response = export_template_xlsx(self.model)  # إرجاع قيمة من الدالة
        audit_view_event(
            request,
            event_type="admin",
            action="request",
            target_model=self.model._meta.label,
            object_repr=self.model._meta.verbose_name_plural,
            details="تنزيل قالب Excel فارغ للاستيراد",
            program=self.model.__name__,
        )
        return response

    def import_excel_view(self, request: HttpRequest) -> HttpResponse:  # تعريف دالة (Function)
        ctx = self.get_import_context(request).__dict__  # تعيين قيمة لمتغير/إعداد

        if request.method == "GET":  # شرط (If)
            audit_view_event(
                request,
                event_type="screen",
                action="view",
                target_model=self.model._meta.label,
                object_repr=self.model._meta.verbose_name_plural,
                details="فتح شاشة استيراد Excel",
                program=self.model.__name__,
            )
            return render(request, "admin/import_excel.html", {**ctx, "step": 1})  # إرجاع قيمة من الدالة

        temp_id = request.POST.get("temp_id", "").strip()
        sheet_name = request.POST.get("sheet_name", "").strip()
        action = request.POST.get("action", "").strip()

        if action == "download_template":
            response = export_template_xlsx(self.model)
            audit_view_event(
                request,
                event_type="admin",
                action="request",
                target_model=self.model._meta.label,
                object_repr=self.model._meta.verbose_name_plural,
                details="تنزيل قالب Excel من شاشة الاستيراد",
                program=self.model.__name__,
            )
            return response

        if not temp_id:
            uploaded = request.FILES.get("excel_file")
            if not uploaded:
                messages.error(request, "يرجى اختيار ملف Excel.")
                return render(request, "admin/import_excel.html", {**ctx, "step": 1})
            try:
                temp_id = _save_temp_upload(uploaded)
                sheets = _read_sheets(temp_id)
            except ValueError as exc:
                messages.error(request, str(exc))
                return render(request, "admin/import_excel.html", {**ctx, "step": 1})
            audit_view_event(
                request,
                event_type="admin",
                action="request",
                target_model=self.model._meta.label,
                object_repr=self.model._meta.verbose_name_plural,
                after_data={"temp_id": temp_id, "uploaded_name": getattr(uploaded, "name", ""), "sheet_count": len(sheets)},
                changed_fields=["excel_upload"],
                details="رفع ملف Excel وتجهيز معاينة الأوراق",
                program=self.model.__name__,
            )
            return render(request, "admin/import_excel.html", {**ctx, "step": 2, "temp_id": temp_id, "sheets": sheets})

        sheets = _read_sheets(temp_id)
        if not sheet_name:
            return render(request, "admin/import_excel.html", {**ctx, "step": 2, "temp_id": temp_id, "sheets": sheets})

        if sheet_name not in sheets:
            messages.error(request, "الورقة المختارة غير موجودة داخل الملف.")
            return render(request, "admin/import_excel.html", {**ctx, "step": 2, "temp_id": temp_id, "sheets": sheets})

        headers = _read_headers(temp_id, sheet_name)
        ok, missing, extra = validate_headers(self.model, headers)
        if not ok:
            return render(
                request,
                "admin/import_excel_mismatch.html",
                {**ctx, "temp_id": temp_id, "sheet_name": sheet_name, "missing": missing, "extra": extra},
            )

        if action != "confirm_import":
            preview = preview_sheet(self.model, temp_id, sheet_name)
            audit_view_event(
                request,
                event_type="screen",
                action="view",
                target_model=self.model._meta.label,
                object_repr=self.model._meta.verbose_name_plural,
                after_data={"sheet_name": sheet_name, "preview_rows": len(preview.get("rows", []) if isinstance(preview, dict) else [])},
                changed_fields=["sheet_name"],
                details="معاينة بيانات Excel قبل الاستيراد",
                program=self.model.__name__,
            )
            return render(
                request,
                "admin/import_excel_preview.html",
                {**ctx, "step": 3, "temp_id": temp_id, "sheet_name": sheet_name, "preview": preview},
            )

        try:
            created, errors = import_sheet(self.model, temp_id, sheet_name)
        except Exception as exc:
            audit_error_event(
                request,
                event_type="admin",
                target_model=self.model._meta.label,
                object_repr=self.model._meta.verbose_name_plural,
                details=f"فشل استيراد Excel من الورقة {sheet_name}: {exc.__class__.__name__}: {exc}",
                program=self.model.__name__,
            )
            raise
        audit_view_event(
            request,
            event_type="admin",
            action="create",
            target_model=self.model._meta.label,
            object_repr=self.model._meta.verbose_name_plural,
            after_data={"sheet_name": sheet_name, "created": created, "error_count": len(errors)},
            changed_fields=["excel_import"],
            details=f"تنفيذ استيراد Excel من الورقة {sheet_name}",
            program=self.model.__name__,
            success=len(errors) == 0,
        )
        issues = [
            build_import_issue(categorize_import_message(message), message)
            for message in errors
        ]
        summary_rows = summarize_import_issues(issues)
        summary = {
            "created": created,
            "error_count": len(errors),
            "duplicate_count": sum(row["count"] for row in summary_rows if row["category"] in {"duplicate_in_file", "duplicate_existing"}),
            "validation_count": sum(row["count"] for row in summary_rows if row["category"] in {"header_mismatch", "missing_required", "invalid_date", "invalid_birthdate", "invalid_registration", "save_error", "generic"}),
            "by_category": summary_rows,
        }

        return render(
            request,
            "admin/import_excel_report.html",
            {
                **ctx,
                "title": f"نتيجة استيراد بيانات Excel — {ctx['model_verbose']}",
                "sheet_name": sheet_name,
                "created": created,
                "messages_list": errors,
                "issue_groups": group_import_issues(issues),
                "summary": summary,
                "back_url": reverse(f"admin:{ctx['model_label']}_changelist"),
                "import_again_url": reverse(f"admin:{ctx['model_label']}_import_excel"),
            },
        )
