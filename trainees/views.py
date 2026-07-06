from contextlib import contextmanager
from collections import Counter
import json
import calendar
from datetime import date

from django.contrib.auth import authenticate, login, logout  # استيراد عناصر محددة من مكتبة/وحدة
from django.shortcuts import render, redirect, get_object_or_404  # استيراد عناصر محددة من مكتبة/وحدة
from django.contrib.auth.decorators import login_required  # استيراد عناصر محددة من مكتبة/وحدة
from django.contrib.admin.views.decorators import staff_member_required  # استيراد عناصر محددة من مكتبة/وحدة
from django.contrib.contenttypes.models import ContentType
from django.urls import reverse  # استيراد عناصر محددة من مكتبة/وحدة
import os
import re
import shutil
from pathlib import Path
from django.conf import settings
from django.http import HttpResponseBadRequest, HttpResponse, Http404, QueryDict
from django.utils.http import content_disposition_header
from django.template.loader import render_to_string
from django.core.exceptions import PermissionDenied
from django.contrib import messages  # استيراد عناصر محددة من مكتبة/وحدة
from django.views.decorators.http import require_POST  # استيراد عناصر محددة من مكتبة/وحدة
from urllib.parse import urlencode, quote  # استيراد عناصر محددة من مكتبة/وحدة
from email.header import Header
from django.db import transaction
from django.db.models import Case, When, Value, IntegerField, F, Q, Count, Max, Avg  # استيراد عناصر محددة من مكتبة/وحدة
from django.utils import timezone  # استيراد عناصر محددة من مكتبة/وحدة
from django.core.paginator import Paginator
from django.core.management import call_command
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.utils import simpleSplit
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from .services.attendance_action_document_service import (
    attendance_action_document_context,
    build_attendance_action_word_response,
    build_attendance_action_pdf_response,
)

try:
    import arabic_reshaper
except Exception:
    arabic_reshaper = None

try:
    from bidi.algorithm import get_display
except Exception:
    get_display = None

from .models import حضوري_أولي, تمهين, مسائي_ومعابر, دفعة, refresh_all_promotion_semester_starts, cohort_start_dates_for_model, ActivityLog, كشفغياب, خليةغياب, AttendanceStatSnapshot, AttendanceAction, AttendanceActionDeletion, DismissalDecision, SanctionRecord, CustomField, CustomFieldValue, ComprehensiveAuditLog
from trainees.status_utils import unified_status_code  # استيراد عناصر محددة من مكتبة/وحدة
from .forms import FORM_BY_PROGRAM, MODEL_BY_PROGRAM, AttendanceActionForm, DismissalDecisionForm, SanctionRecordForm  # استيراد عناصر محددة من مكتبة/وحدة
from .forms import DATE_INPUT_FORMATS
from .semester_utils import compute_semester_for_trainee, normalize_repeater_training_dates, resolve_session_year  # استيراد عناصر محددة من مكتبة/وحدة
from .permissions import has_program_permission, visible_programs, can_access_admin_panel, require_program_permission, is_access_within_schedule, get_access_denied_message, build_access_summary
from .program_columns import DISPLAY_PROGRAM_COLUMNS
from .evening_training_type import (
    EVENING_TRAINING_TYPE_EVENING,
    EVENING_TRAINING_TYPE_CROSSING,
    clean_crossing_specialty_label,
    detect_evening_training_type,
    clamp_semester_for_evening_type,
)
from .program_split_utils import (
    exclude_inactive_trainees,
    filter_evening_trainee_queryset_by_program,
    filter_generic_records_to_active_trainees,
    filter_records_by_split_program,
    filter_records_by_split_program_for_active_trainees,
)


from .services.attendance_action_sync_service import (
    build_monthly_action_payload,
    next_attendance_document_number,
    sync_attendance_actions,
)
from .services.attendance_action_management_service import (
    attendance_action_base_query,
    attendance_action_source,
    attendance_actions_qs,
    clear_attendance_action_deletion,
    parse_bulk_action_date,
    register_attendance_action_deletion,
    selected_action_ids_from_request,
    summarize_attendance_actions,
)

from .attendance_slots_common import SLOT_PROGRAMS as ATTENDANCE_SLOT_PROGRAMS, _table_payload as build_attendance_slot_payload
from .services.attendance_slot_action_sync_service import (
    AUTO_ARCHIVE_MARK,
    MANUAL_ARCHIVE_MARK,
    sync_slot_attendance_actions,
)
from .services.saved_attendance_stats_service import build_saved_attendance_stats_archive_context
from .services.saved_attendance_stats_export_service import (
    build_saved_attendance_stats_excel_response,
    build_saved_attendance_stats_pdf_response,
)


def _append_attendance_action_note_marker(notes: str, marker: str) -> str:
    notes = (notes or "").strip()
    if marker in notes:
        return notes
    return (notes + "\n" + marker).strip() if notes else marker


def _remove_attendance_action_note_marker(notes: str, marker: str) -> str:
    lines = [line.strip() for line in (notes or "").splitlines() if line.strip() and line.strip() != marker]
    return "\n".join(lines).strip()


def _mark_attendance_action_manual_archive(obj) -> bool:
    new_notes = _append_attendance_action_note_marker(
        _remove_attendance_action_note_marker(obj.notes, AUTO_ARCHIVE_MARK),
        MANUAL_ARCHIVE_MARK,
    )
    if obj.notes != new_notes:
        obj.notes = new_notes
        return True
    return False


def _clear_attendance_action_archive_markers(obj) -> bool:
    new_notes = _remove_attendance_action_note_marker(
        _remove_attendance_action_note_marker(obj.notes, AUTO_ARCHIVE_MARK),
        MANUAL_ARCHIVE_MARK,
    )
    if obj.notes != new_notes:
        obj.notes = new_notes
        return True
    return False


from .services.attendance_table_service import (
    build_attendance_changes,
    delete_saved_attendance_entries,
    existing_attendance_entries,
    persist_attendance_changes,
)
from .services.account_dashboard_service import (
    build_account_context,
    build_dashboard_context,
)
from .services.listing_service import (
    apply_advanced_filters as listing_apply_advanced_filters,
    build_program_title,
    build_query_string_without_page,
    build_semester_options,
    build_specialty_options,
    can_export_for_user,
    extract_list_filters,
    normalize_text,
    unique_clean_values,
)
from .services.attendance_view_state_service import (
    build_preserved_query,
    parse_old_stats_cutoff,
    resolve_attendance_post_action,
    should_process_attendance_delete,
    should_process_attendance_save,
    valid_old_stats_cutoff,
)


from .services.media_service import save_uploaded_media
from .services.attendance_navigation_service import build_attendance_home_cards



def _client_ip(request):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return (request.META.get("REMOTE_ADDR") or "").strip() or None


def log_activity(request, action, program="", obj=None, details=""):
    user = request.user if getattr(request, "user", None) and request.user.is_authenticated else None
    object_repr = ""
    if obj is not None:
        object_repr = getattr(obj, "اللقب_والاسم", None) or str(obj)
    ActivityLog.objects.create(
        user=user,
        action=action,
        program=program or "",
        object_repr=object_repr[:255],
        details=(details or "")[:2000],
        path=(request.path or "")[:255],
        ip_address=_client_ip(request),
    )


def deny_with_log(request, program, action):
    log_activity(request, "access_denied", program=program, details=f"محاولة {action}")
    raise PermissionDenied("غير مصرح لك بهذا الإجراء")


def _user_force_password_change(user):
    if not getattr(user, "is_authenticated", False):
        return False
    try:
        return bool(user.access_profile.force_password_change)
    except Exception:
        return False


@login_required
def account_overview(request):
    # صفحة مستقلة توضح للمستخدم حالته وصلاحياته بدل الاكتفاء برسالة مختصرة.
    context = build_account_context(request.user, build_access_summary_func=build_access_summary)
    context["title"] = "حالة الحساب والصلاحيات"
    return render(request, "trainees/account_overview.html", context)

def _admin_changelist_url(model_cls):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Safely build the admin changelist URL for a model (works with Arabic model names)."""
    return reverse(f"admin:{model_cls._meta.app_label}_{model_cls._meta.model_name}_changelist")  # إرجاع قيمة من الدالة


def _admin_changelist_filtered(model_cls, **query_params):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Admin changelist URL + querystring filters."""
    base = _admin_changelist_url(model_cls)  # تعيين قيمة لمتغير/إعداد
    if not query_params:  # شرط (If)
        return base  # إرجاع قيمة من الدالة
    return f"{base}?{urlencode(query_params)}"  # إرجاع قيمة من الدالة



def _recompute_semesters(model_cls):  # تعريف دالة (Function)
    """إعادة حساب السداسي وربط الدفعات بكفاءة أعلى."""
    refresh_all_promotion_semester_starts()
    field_names = {f.name for f in model_cls._meta.get_fields()}
    fields = ["id", "رقم_التسجيل", "الدفعة", "تاريخ_بداية_التكوين", "تاريخ_نهاية_التكوين", "السداسي"]
    if "معيد" in field_names:
        fields.append("معيد")
    if "تاريخ_التكوين_السابق_للمعيدين" in field_names:
        fields.append("تاريخ_التكوين_السابق_للمعيدين")
    if "نوع_التكوين" in field_names:
        fields.append("نوع_التكوين")
    qs = model_cls.objects.select_related("الدفعة").only(*fields)
    cohort_starts = cohort_start_dates_for_model(model_cls)
    cohort_starts_by_type = {}

    promotion_map = {
        (p.رقم_الدورة, p.السنة): p
        for p in دفعة.objects.filter(مفعلة=True).only("id", "رقم_الدورة", "السنة", "بداية_السداسي_1", "بداية_السداسي_2", "بداية_السداسي_3", "بداية_السداسي_4", "بداية_السداسي_5")
    }

    to_update = []
    for obj in qs.iterator(chunk_size=2000):
        original_promotion_id = obj.الدفعة_id
        original_semester = obj.السداسي or ""

        if not obj.الدفعة and getattr(obj, "رقم_التسجيل", None):
            session_no, year_value = resolve_session_year(obj.رقم_التسجيل, getattr(obj, "تاريخ_بداية_التكوين", None))
            if session_no and year_value:
                obj.الدفعة = promotion_map.get((session_no, year_value))

        if bool(getattr(obj, "معيد", False)) and hasattr(obj, "تاريخ_التكوين_السابق_للمعيدين"):
            from .semester_utils import normalize_repeater_training_dates
            normalize_repeater_training_dates(obj)

        is_repeater = bool(getattr(obj, "معيد", False))
        row_cohort_starts = cohort_starts
        if "نوع_التكوين" in field_names:
            training_type = getattr(obj, "نوع_التكوين", None)
            if training_type:
                row_cohort_starts = cohort_starts_by_type.setdefault(
                    training_type,
                    cohort_start_dates_for_model(model_cls, training_type=training_type),
                )
        obj.السداسي = compute_semester_for_trainee(
            obj.الدفعة,
            obj.تاريخ_بداية_التكوين,
            obj.تاريخ_نهاية_التكوين,
            is_repeater=is_repeater,
            cohort_starts=row_cohort_starts,
            original_end_date=getattr(obj, "تاريخ_التكوين_السابق_للمعيدين", None),
        ) or obj.السداسي
        if "نوع_التكوين" in field_names:
            obj.السداسي = clamp_semester_for_evening_type(obj.السداسي, getattr(obj, "نوع_التكوين", None))

        if obj.الدفعة_id != original_promotion_id or (obj.السداسي or "") != original_semester:
            to_update.append(obj)

    if to_update:
        update_fields = ["الدفعة", "السداسي"]
        if "تاريخ_نهاية_التكوين" in field_names:
            update_fields.append("تاريخ_نهاية_التكوين")
        if "تاريخ_التكوين_السابق_للمعيدين" in field_names:
            update_fields.append("تاريخ_التكوين_السابق_للمعيدين")
        model_cls.objects.bulk_update(to_update, update_fields, batch_size=1000)


def _attach_row_flags(rows):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Attach UI-only flags on row objects (no DB writes)."""
    for obj in rows:  # حلقة تكرار (For)
        try:  # سطر كود لتنفيذ منطق/إعداد
            obj.is_removed = (unified_status_code(getattr(obj, "الحالة", "")) == "removed")  # تعيين قيمة لمتغير/إعداد
        except Exception:  # سطر كود لتنفيذ منطق/إعداد
            obj.is_removed = False  # تعيين قيمة لمتغير/إعداد
    return rows  # إرجاع قيمة من الدالة


def _refresh_rows_live_semesters(rows, model_cls):
    """احسب السداسي مباشرةً للصفوف المعروضة وحدّث العمود فوراً.

    هذا يضمن أن عمود السداسي في الواجهة الرئيسية يعكس رزنامة الدفعات
    حتى لو كانت هناك سجلات قديمة لم تُحفظ بعد بمنطق الحساب الجديد.
    """
    rows = list(rows)
    if not rows:
        return rows

    to_update = []
    field_names = {f.name for f in model_cls._meta.get_fields()}
    has_repeater = "معيد" in field_names
    common_cohort_starts = cohort_start_dates_for_model(model_cls)
    cohort_starts_by_type = {}

    for obj in rows:
        training_type = getattr(obj, "نوع_التكوين", None) if getattr(model_cls, "__name__", "") == "مسائي_ومعابر" else None
        if training_type:
            cohort_starts = cohort_starts_by_type.setdefault(
                training_type,
                cohort_start_dates_for_model(model_cls, training_type=training_type),
            )
        else:
            cohort_starts = common_cohort_starts
        new_sem = compute_semester_for_trainee(
            getattr(obj, "الدفعة", None),
            getattr(obj, "تاريخ_بداية_التكوين", None),
            getattr(obj, "تاريخ_نهاية_التكوين", None),
            is_repeater=bool(getattr(obj, "معيد", False)) if has_repeater else False,
            cohort_starts=cohort_starts,
            original_end_date=getattr(obj, "تاريخ_التكوين_السابق_للمعيدين", None),
        )
        if new_sem and getattr(model_cls, "__name__", "") == "مسائي_ومعابر":
            new_sem = clamp_semester_for_evening_type(new_sem, getattr(obj, "نوع_التكوين", None))
        if new_sem:
            if getattr(obj, "السداسي", None) != new_sem:
                obj.السداسي = new_sem
                to_update.append(obj)
            else:
                obj.السداسي = new_sem

    if to_update:
        model_cls.objects.bulk_update(to_update, ["السداسي"], batch_size=1000)

    return rows


def _semester_rank_case():  # تعريف دالة (Function)
    # Map Arabic semester labels to numeric rank (1..5). Unknown/empty goes last (99).
    return Case(  # إرجاع قيمة من الدالة
        When(السداسي="الأول", then=Value(1)),  # تعيين قيمة لمتغير/إعداد
        When(السداسي="الثاني", then=Value(2)),  # تعيين قيمة لمتغير/إعداد
        When(السداسي="الثالث", then=Value(3)),  # تعيين قيمة لمتغير/إعداد
        When(السداسي="الرابع", then=Value(4)),  # تعيين قيمة لمتغير/إعداد
        When(السداسي="الخامس", then=Value(5)),  # تعيين قيمة لمتغير/إعداد
        default=Value(99),  # تعيين قيمة لمتغير/إعداد
        output_field=IntegerField(),  # تعيين قيمة لمتغير/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد




def _removed_rank_case():  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Return Case expression: 1 for removed/مشطوب-like statuses, else 0.
    We intentionally keep this DB-side (no Python normalization) using common keywords.  # سطر كود لتنفيذ منطق/إعداد
    # سطر كود لتنفيذ منطق/إعداد
    """
    return Case(  # إرجاع قيمة من الدالة
        When(Q(الحالة__icontains="مشطوب") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="شطب") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="مفصول") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="فصل") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="متوقف") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="موقوف") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="توقف") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="منقطع") |  # تعيين قيمة لمتغير/إعداد
             Q(الحالة__icontains="انسحب"),  # تعيين قيمة لمتغير/إعداد
             then=Value(1)),  # تعيين قيمة لمتغير/إعداد
        default=Value(0),  # تعيين قيمة لمتغير/إعداد
        output_field=IntegerField(),  # تعيين قيمة لمتغير/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد

def _get_ordered_rows(model_cls, program: str, graduates: bool = False):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Return an ordered queryset for list views.

    Ordering goals:  # سطر كود لتنفيذ منطق/إعداد
    - All programs: order by specialization, then by semester progress (1..5),  # سطر كود لتنفيذ منطق/إعداد
      so trainees closer to the end appear lower in the list.  # سطر كود لتنفيذ منطق/إعداد
    - Apprentice (تمهين): repeaters (معيد=True) must appear below their non-repeater  # تعيين قيمة لمتغير/إعداد
      classmates within the same cohort (same تخصص + نفس تواريخ التكوين الأصلية).  # سطر كود لتنفيذ منطق/إعداد
      We keep them grouped with their peers, but pushed to the bottom.  # سطر كود لتنفيذ منطق/إعداد
    # سطر كود لتنفيذ منطق/إعداد
    """
    qs = model_cls.objects.select_related("الدفعة").all()  # تعيين قيمة لمتغير/إعداد
    qs = _filter_queryset_by_program(qs, program)

    today = timezone.localdate()  # تعيين قيمة لمتغير/إعداد
    # Graduates: end date exists and is in the past/today
    if graduates:  # شرط (If)
        qs = qs.filter(تاريخ_نهاية_التكوين__isnull=False, تاريخ_نهاية_التكوين__lte=today)  # تعيين قيمة لمتغير/إعداد
    else:  # فرع بديل (Else)
        # Current trainees: end date missing or in the future
        qs = qs.filter(Q(تاريخ_نهاية_التكوين__isnull=True) | Q(تاريخ_نهاية_التكوين__gt=today))  # تعيين قيمة لمتغير/إعداد


    # Push records missing key dates to the bottom
    start_is_null = Case(  # تعيين قيمة لمتغير/إعداد
        When(تاريخ_بداية_التكوين__isnull=True, then=Value(1)),  # تعيين قيمة لمتغير/إعداد
        default=Value(0),  # تعيين قيمة لمتغير/إعداد
        output_field=IntegerField(),  # تعيين قيمة لمتغير/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد
    end_is_null = Case(  # تعيين قيمة لمتغير/إعداد
        When(تاريخ_نهاية_التكوين__isnull=True, then=Value(1)),  # تعيين قيمة لمتغير/إعداد
        default=Value(0),  # تعيين قيمة لمتغير/إعداد
        output_field=IntegerField(),  # تعيين قيمة لمتغير/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد

    sem_rank = _semester_rank_case()  # تعيين قيمة لمتغير/إعداد

    removed_rank = _removed_rank_case()  # تعيين قيمة لمتغير/إعداد

    if program == "apprentice":  # شرط (If)
        # Cohort end date: for repeaters we use the original end date (stored in تاريخ_التكوين_السابق_للمعيدين)
        # so they stay grouped with their classmates, not separated by the extended end date.
        field_names = {f.name for f in model_cls._meta.get_fields()}  # تعيين قيمة لمتغير/إعداد
        if "معيد" in field_names and "تاريخ_التكوين_السابق_للمعيدين" in field_names:  # شرط (If)
            cohort_end = Case(  # تعيين قيمة لمتغير/إعداد
                When(معيد=True, then=F("تاريخ_التكوين_السابق_للمعيدين")),  # تعيين قيمة لمتغير/إعداد
                default=F("تاريخ_نهاية_التكوين"),  # تعيين قيمة لمتغير/إعداد
            )  # سطر كود لتنفيذ منطق/إعداد
            repeater_rank = Case(  # تعيين قيمة لمتغير/إعداد
                When(معيد=True, then=Value(1)),  # تعيين قيمة لمتغير/إعداد
                default=Value(0),  # تعيين قيمة لمتغير/إعداد
                output_field=IntegerField(),  # تعيين قيمة لمتغير/إعداد
            )  # سطر كود لتنفيذ منطق/إعداد
            display_sem_rank = Case(  # تعيين قيمة لمتغير/إعداد
                When(معيد=True, then=F("_sem_rank") + Value(1)),  # تعيين قيمة لمتغير/إعداد
                default=F("_sem_rank"),  # تعيين قيمة لمتغير/إعداد
                output_field=IntegerField(),  # تعيين قيمة لمتغير/إعداد
            )  # سطر كود لتنفيذ منطق/إعداد
            qs = qs.annotate(  # تعيين قيمة لمتغير/إعداد
                _start_null=start_is_null,  # تعيين قيمة لمتغير/إعداد
                _end_null=end_is_null,  # تعيين قيمة لمتغير/إعداد
                _sem_rank=sem_rank,  # تعيين قيمة لمتغير/إعداد
                _display_sem_rank=display_sem_rank,  # تعيين قيمة لمتغير/إعداد
                _cohort_end=cohort_end,  # تعيين قيمة لمتغير/إعداد
                _repeater_rank=repeater_rank,  # تعيين قيمة لمتغير/إعداد
                _removed_rank=removed_rank,  # تعيين قيمة لمتغير/إعداد
            ).order_by(  # سطر كود لتنفيذ منطق/إعداد
                "_start_null",  # سطر كود لتنفيذ منطق/إعداد
                # سطر كود لتنفيذ منطق/إعداد
                "_display_sem_rank",        # semester progression (repeaters shifted +1 for ordering)
                "التخصص",  # سطر كود لتنفيذ منطق/إعداد
                "تاريخ_بداية_التكوين",  # سطر كود لتنفيذ منطق/إعداد
                "_end_null", "_cohort_end",  # سطر كود لتنفيذ منطق/إعداد
                "_removed_rank",  # سطر كود لتنفيذ منطق/إعداد
                # سطر كود لتنفيذ منطق/إعداد
                "_repeater_rank",           # non-repeaters first, repeaters last within same cohort
                "اللقب", "الاسم", "id"  # سطر كود لتنفيذ منطق/إعداد
            )  # سطر كود لتنفيذ منطق/إعداد
            return qs  # إرجاع قيمة من الدالة

    # Default ordering for initial/evening (and apprentice if fields missing)
    qs = qs.annotate(  # تعيين قيمة لمتغير/إعداد
        _start_null=start_is_null,  # تعيين قيمة لمتغير/إعداد
        _end_null=end_is_null,  # تعيين قيمة لمتغير/إعداد
        _sem_rank=sem_rank,  # تعيين قيمة لمتغير/إعداد
        _removed_rank=removed_rank,  # تعيين قيمة لمتغير/إعداد
    ).order_by(  # سطر كود لتنفيذ منطق/إعداد
        "_start_null",  # سطر كود لتنفيذ منطق/إعداد
        # سطر كود لتنفيذ منطق/إعداد
        "_sem_rank",                  # semester progression 1..5
        "التخصص",  # سطر كود لتنفيذ منطق/إعداد
        "تاريخ_بداية_التكوين",  # سطر كود لتنفيذ منطق/إعداد
        "_end_null", "تاريخ_نهاية_التكوين",  # سطر كود لتنفيذ منطق/إعداد
        "_removed_rank",  # سطر كود لتنفيذ منطق/إعداد
        "اللقب", "الاسم", "id"  # سطر كود لتنفيذ منطق/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد
    return qs  # إرجاع قيمة من الدالة



@contextmanager
def _temporary_env(overrides: dict[str, str]):
    previous = {}
    missing = object()
    try:
        for key, value in overrides.items():
            previous[key] = os.environ.get(key, missing)
            os.environ[key] = value
        yield
    finally:
        for key, value in previous.items():
            if value is missing:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


def _is_desktop_mode() -> bool:
    return str(getattr(settings, "APP_MODE", "") or "").strip().lower() == "desktop"


def _show_developer_login() -> bool:
    # زر المطور لا يظهر في صفحة دخول المستخدمين العاديين.
    return False


def _developer_login_enabled() -> bool:
    return os.getenv("DEV_LOGIN_ENABLED", "0").strip().lower() in {"1", "true", "yes", "on"}


def _developer_credentials() -> tuple[str, str, str]:
    # لا توجد قيم افتراضية لكلمة مرور المطوّر داخل نسخ الأجهزة.
    # يجب أن تكون موجودة فقط في .env الخاص بجهاز المطوّر/الخادم المركزي.
    username = (os.getenv("DEV_USERNAME", "") or "").strip()
    password = os.getenv("DEV_PASSWORD", "")
    email = (os.getenv("DEV_EMAIL", "") or "").strip()
    return username, password, email


def _ensure_desktop_developer_account() -> tuple[str, str]:
    if not _developer_login_enabled():
        raise PermissionError("Developer login is disabled on this device.")
    username, password, email = _developer_credentials()
    if not username or not password:
        raise PermissionError("Developer credentials are not configured on this device.")
    with _temporary_env({
        "DEV_USERNAME": username,
        "DEV_PASSWORD": password,
        "DEV_EMAIL": email,
    }):
        call_command("ensure_developer", "--reset-password")
    return username, password


def _render_error_navigation(request, *, title: str, message: str, status_code: int):
    return render(request, "error_navigation.html", {
        "title": title,
        "error_title": title,
        "error_message": message,
        "status_code": status_code,
    }, status=status_code)


def error_404(request, exception):
    return _render_error_navigation(
        request,
        title="الصفحة غير موجودة",
        message="تعذر العثور على الصفحة المطلوبة. يمكنك الرجوع إلى صفحة الدخول ومتابعة العمل.",
        status_code=404,
    )


def error_403(request, exception):
    return _render_error_navigation(
        request,
        title="غير مسموح",
        message="ليس لديك إذن للوصول إلى هذه الصفحة. يمكنك الرجوع إلى صفحة الدخول.",
        status_code=403,
    )


def error_500(request):
    return _render_error_navigation(
        request,
        title="حدث خطأ داخلي",
        message="حدث خطأ غير متوقع داخل البرنامج. يمكنك الرجوع إلى صفحة الدخول ثم المحاولة مرة أخرى.",
        status_code=500,
    )

def developer_login_view(request):
    """صفحة دخول خاصة بالمطوّر الأعلى فقط.

    لا تعمل في أجهزة الموظفين إلا إذا فعّلها المطوّر صراحة داخل ملف .env.
    وعند فتح المكتب من لوحة المطوّر المركزية يجب ألّا تدخل جلسة مستخدم المكتب العادي.
    لذلك إذا كان المتصفح يحمل جلسة مستخدم عادي على نفس المنفذ، نخرجها ونطلب حساب المطوّر فقط.
    """
    if not _developer_login_enabled():
        return _render_error_navigation(
            request,
            title="غير متاح",
            message="دخول المطوّر غير مفعّل على هذا الجهاز.",
            status_code=404,
        )
    context = {}
    if request.user.is_authenticated:
        dev_username, _dev_password, _dev_email = _developer_credentials()
        current_username = getattr(request.user, "username", "") or ""
        if request.user.is_superuser and (not dev_username or current_username == dev_username):
            return redirect("dashboard")
        logout(request)
        context["error"] = "هذه الصفحة مخصصة لحساب المطور فقط. تم إخراج حساب المستخدم العادي، أدخل حساب المطور للمتابعة."
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        dev_username, _dev_password, _dev_email = _developer_credentials()
        if username != dev_username:
            context.update({"error": "هذه الصفحة مخصصة لحساب المطور الأعلى فقط.", "username": username})
            return render(request, "trainees/developer_login.html", context)
        try:
            _ensure_desktop_developer_account()
        except Exception as exc:
            ActivityLog.objects.create(
                user=None,
                action="developer_login_failed",
                program="",
                object_repr="فشل دخول المطور",
                details=f"تعذر تجهيز حساب المطور: {exc}",
                path=(request.path or "")[:255],
                ip_address=_client_ip(request),
            )
            context.update({"error": "تعذر تهيئة حساب المطور. راجع السجلات أو أعد المحاولة.", "username": username})
            return render(request, "trainees/developer_login.html", context)
        user = authenticate(request, username=username, password=password)
        if user and user.is_superuser:
            login(request, user)
            log_activity(request, "developer_login", details="دخول المطور الأعلى")
            return redirect("dashboard")
        context.update({"error": "اسم المستخدم أو كلمة المرور غير صحيحة لحساب المطور.", "username": username})
        return render(request, "trainees/developer_login.html", context)
    return render(request, "trainees/developer_login.html", context)


def login_view(request):  # تعريف دالة (Function)
    if request.user.is_authenticated:  # شرط (If)
        return redirect("dashboard")  # إرجاع قيمة من الدالة
    context = {"show_developer_login": _show_developer_login()}
    if request.method == "POST":  # شرط (If)
        action = (request.POST.get("action") or "user_login").strip()
        username = request.POST.get("username","").strip()  # تعيين قيمة لمتغير/إعداد
        password = request.POST.get("password","")  # تعيين قيمة لمتغير/إعداد
        dev_username, dev_password, _dev_email = _developer_credentials()

        if action == "developer_login" and _show_developer_login():
            if username != dev_username:
                context.update({
                    "error": "زر دخول المطور مخصص لحساب المطور فقط.",
                    "username": username,
                    "show_developer_login": _show_developer_login(),
                })
                return render(request, "trainees/login.html", context)
            try:
                _ensure_desktop_developer_account()
            except Exception as exc:
                ActivityLog.objects.create(
                    user=None,
                    action="developer_login_failed",
                    program="",
                    object_repr="فشل دخول المطور",
                    details=f"تعذر تجهيز حساب المطور: {exc}",
                    path=(request.path or "")[:255],
                    ip_address=_client_ip(request),
                )
                context.update({
                    "error": "تعذر تهيئة حساب المطور حاليًا. راجع السجلات أو أعد المحاولة.",
                    "username": username,
                    "show_developer_login": _show_developer_login(),
                })
                return render(request, "trainees/login.html", context)
        elif action == "developer_login":
            context.update({
                "error": "دخول المطور غير متاح في هذا الوضع.",
                "username": username,
                "show_developer_login": _show_developer_login(),
            })
            return render(request, "trainees/login.html", context)
        else:
            if dev_username and username == dev_username:
                context.update({
                    "error": "حساب المطور يجب أن يدخل عبر صفحة دخول المطور فقط.",
                    "username": username,
                    "show_developer_login": _show_developer_login(),
                })
                return render(request, "trainees/login.html", context)

        user = authenticate(request, username=username, password=password)  # تعيين قيمة لمتغير/إعداد
        if user:  # شرط (If)
            if not is_access_within_schedule(user):
                denied_message = get_access_denied_message(user)
                ActivityLog.objects.create(
                    user=user,
                    action="access_denied",
                    program="",
                    object_repr="منع دخول",
                    details=f"تم منع الدخول بعد نجاح التحقق من كلمة المرور. السبب: {denied_message}",
                    path=(request.path or "")[:255],
                    ip_address=_client_ip(request),
                )
                from .models import UserAccountAuditLog, serialize_sensitive_account_state
                profile = getattr(user, "access_profile", None)
                UserAccountAuditLog.objects.create(
                    actor=None,
                    target_user=user,
                    action="login_denied_window",
                    changed_fields=[],
                    before_data=serialize_sensitive_account_state(profile) if profile else {},
                    after_data={},
                    notes=f"تم رفض تسجيل الدخول بسبب نافذة الصلاحية. السبب: {denied_message}",
                    ip_address=_client_ip(request),
                )
                denied_context = build_account_context(user, build_access_summary_func=build_access_summary)
                denied_context.update({
                    "title": "تعذر تسجيل الدخول",
                    "error": denied_message,
                    "show_login_actions": True,
                    "username": username,
                })
                return render(request, "trainees/access_status.html", denied_context)
            login(request, user)
            log_activity(request, "login", details="تسجيل دخول ناجح")
            if _user_force_password_change(user):
                messages.warning(request, "يجب تغيير كلمة المرور من لوحة الإدارة بواسطة المدير أو عبر صفحة الإدارة.")
            return redirect("dashboard")
        ActivityLog.objects.create(
            user=None,
            action="login_failed",
            program="",
            object_repr="محاولة دخول فاشلة",
            details=f"فشل تسجيل الدخول للحساب: {username}",
            path=(request.path or "")[:255],
            ip_address=_client_ip(request),
        )
        context.update({"error": "اسم المستخدم أو كلمة المرور غير صحيحة", "username": username, "show_developer_login": _show_developer_login()})
        return render(request, "trainees/login.html", context)  # إرجاع قيمة من الدالة
    return render(request, "trainees/login.html", context)  # إرجاع قيمة من الدالة

def logout_view(request):  # تعريف دالة (Function)
    if request.user.is_authenticated:
        log_activity(request, "logout", details="تسجيل خروج")
    logout(request)
    return redirect("login")


def _redirect_after_save(request, program, obj_pk, default="list"):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """Redirect based on which submit button was used.
    action values: stay | list | back  # سطر كود لتنفيذ منطق/إعداد
    # سطر كود لتنفيذ منطق/إعداد
    """
    # تعيين قيمة لمتغير/إعداد
    action = request.POST.get("action")  # set by submit buttons
    if not action:  # شرط (If)
        action = default  # تعيين قيمة لمتغير/إعداد
    if action == "stay":  # شرط (If)
        # stay on edit page
        return redirect("trainee_edit", program=program, pk=obj_pk) if _has_named_url("trainee_edit") else redirect(request.path)  # إرجاع قيمة من الدالة
    if action == "back":  # شرط (If)
        # go back to previous page if available, else list
        return redirect(request.META.get("HTTP_REFERER") or _program_list_url(program))  # إرجاع قيمة من الدالة
    # list
    return redirect(_program_list_url(program))  # إرجاع قيمة من الدالة

def _has_named_url(name: str) -> bool:  # تعريف دالة (Function)
    from django.urls import get_resolver  # استيراد عناصر محددة من مكتبة/وحدة
    try:  # سطر كود لتنفيذ منطق/إعداد
        return any(p.name == name for p in get_resolver().url_patterns)  # إرجاع قيمة من الدالة
    except Exception:  # سطر كود لتنفيذ منطق/إعداد
        return False  # إرجاع قيمة من الدالة


@login_required
def dashboard(request):
    allowed_programs = visible_programs(request.user)
    admin_access = can_access_admin_panel(request.user)

    if not allowed_programs and not admin_access:
        messages.error(request, "لم تُمنح لك أي صلاحية بعد. تواصل مع مدير النظام.")

    log_activity(request, "view", details="عرض الرئيسية")
    dashboard_context = build_dashboard_context(
        request.user,
        program_specs=[
            ("initial", "الحضوري الأولي", حضوري_أولي),
            ("apprentice", "التكوين عن طريق التمهين", تمهين),
            ("evening", "الدروس المسائية", مسائي_ومعابر),
            ("crossing", "المعابر", مسائي_ومعابر),
        ],
        today=timezone.localdate(),
        allowed_programs=allowed_programs,
        admin_access=admin_access,
        promotion_count=دفعة.objects.count(),
        get_ordered_rows=_get_ordered_rows,
        refresh_rows_live_semesters=_refresh_rows_live_semesters,
        build_access_summary_func=build_access_summary,
        reverse_func=reverse,
    )
    return render(request, "trainees/dashboard.html", dashboard_context)


PROGRAM_TITLES = {
    "initial": "الحضوري الأولي",
    "apprentice": "عن طريق التمهين",
    "evening": "الدروس المسائية",
    "crossing": "المعابر",
}


def _evening_type_for_program(program: str):
    if program == "crossing":
        return EVENING_TRAINING_TYPE_CROSSING
    if program == "evening":
        return EVENING_TRAINING_TYPE_EVENING
    return None


def _filter_queryset_by_program(qs, program: str):
    return filter_evening_trainee_queryset_by_program(qs, program)


def _force_evening_type_for_program(obj, program: str):
    training_type = _evening_type_for_program(program)
    if training_type and hasattr(obj, "نوع_التكوين"):
        obj.نوع_التكوين = training_type
        if getattr(obj, "التخصص", None):
            obj.التخصص = clean_crossing_specialty_label(obj.التخصص)
    elif getattr(obj, "__class__", None) and obj.__class__.__name__ == "مسائي_ومعابر" and hasattr(obj, "نوع_التكوين"):
        obj.نوع_التكوين = detect_evening_training_type(obj)
        if getattr(obj, "التخصص", None):
            obj.التخصص = clean_crossing_specialty_label(obj.التخصص)
    return obj


def _program_requires_custom_weekdays(program: str, specialty: str = "") -> bool:
    # التمهين، الدروس المسائية، والمعابر لا تكون أيام الدراسة فيها ثابتة دائمًا.
    # لذلك يجب اختيار أيام الدراسة قبل إنشاء الجدول حتى تنعكس على العرض والحفظ
    # والتصدير والإعذارات والنسبة الرسمية.
    return program in {"apprentice", "evening", "crossing"}


def _program_allows_third_weekday(program: str) -> bool:
    # اليوم الثالث أصبح اختياريًا في التمهين أيضًا.
    return program in {"apprentice", "evening", "crossing"}


def _program_max_custom_weekdays(program: str) -> int:
    # التمهين: يومان أساسيان + يوم ثالث اختياري.
    # الدروس المسائية والمعابر: حتى خمسة أيام اختيارية من السبت إلى الخميس.
    if program == "apprentice":
        return 3
    if program in {"evening", "crossing"}:
        return 5
    return len(ATTENDANCE_PROGRAMS.get(program, {}).get("weekday_numbers", []))


def _custom_weekday_keys(program: str):
    max_days = _program_max_custom_weekdays(program)
    return tuple(f"weekday{i}" for i in range(1, max_days + 1))


def _attendance_weekday_choices_for_program(program: str):
    if _program_requires_custom_weekdays(program):
        return [{"value": value, "label": AR_WEEKDAYS[value]} for value in CUSTOM_ATTENDANCE_WEEKDAY_VALUES]
    return [{"value": value, "label": label} for value, label in كشفغياب.WEEKDAY_CHOICES]

PROGRAM_COLUMNS = DISPLAY_PROGRAM_COLUMNS



def _base_cols_for_program(program: str):
    return list(PROGRAM_COLUMNS.get(program, []))


def _attach_program_list_extra_columns(rows, model_cls, program: str, graduates: bool = False):
    """Attach calculated columns used in PROGRAM_COLUMNS.

    - رقم_مقرر_الفصل أصبح حقلاً حقيقياً في BaseTrainee ويُستورد من Excel.
    - إذا وُجد مقرر فصل منشأ في DismissalDecision وله رقم، نعرض رقم المقرر المنشأ
      لأنه أحدث من الرقم المستورد. وإذا لم يوجد، نُبقي الرقم المحفوظ في سجل المتكون.
    - بلدية_الإقامة_بالعربية حقل حقيقي في BaseTrainee/models.py ولا نكتب فوقه هنا.
    """
    if not rows or not model_cls:
        return rows

    scope = "graduated" if graduates else "current"
    try:
        ct = ContentType.objects.get_for_model(model_cls)
        ids = [obj.pk for obj in rows]
        decisions = DismissalDecision.objects.filter(
            program=program,
            decision_scope=scope,
            trainee_content_type=ct,
            trainee_object_id__in=ids,
            is_archived=False,
        ).only("trainee_object_id", "decision_number").order_by("trainee_object_id", "-id")
        decision_numbers = {}
        for obj in decisions:
            if obj.decision_number and obj.trainee_object_id not in decision_numbers:
                decision_numbers[obj.trainee_object_id] = obj.decision_number
        for row in rows:
            decision_number = decision_numbers.get(row.pk)
            if decision_number:
                setattr(row, "رقم_مقرر_الفصل", decision_number)
    except Exception:
        # لا نوقف صفحة المتكونين إذا تعذر الوصول إلى جدول مقررات الفصل.
        pass
    return rows


def _normalized_str(value):
    return normalize_text(value)


def _unique_clean_values(values):
    return unique_clean_values(values)


def _apply_advanced_filters(qs, request):
    q = _normalized_str(request.GET.get("q"))
    semester = _normalized_str(request.GET.get("semester"))
    year = _normalized_str(request.GET.get("year"))
    promotion_id = _normalized_str(request.GET.get("promotion"))
    status = _normalized_str(request.GET.get("status"))
    specialty = _normalized_str(request.GET.get("specialty"))

    if q:
        qs = qs.filter(
            Q(الرقم_التعريفي__icontains=q) |
            Q(اللقب__icontains=q) |
            Q(الاسم__icontains=q) |
            Q(التخصص__icontains=q) |
            Q(رقم_التسجيل__icontains=q) |
            Q(رقم_الهاتف__icontains=q)
        )
    if semester:
        qs = qs.filter(السداسي=semester)
    if specialty:
        qs = qs.filter(التخصص=specialty)
    if year.isdigit():
        qs = qs.filter(الدفعة__السنة=int(year))
    if promotion_id.isdigit():
        qs = qs.filter(الدفعة_id=int(promotion_id))
    if status == "active":
        qs = qs.exclude(الحالة__icontains="مشطوب").exclude(الحالة__icontains="شطب")
    elif status == "removed":
        qs = qs.filter(Q(الحالة__icontains="مشطوب") | Q(الحالة__icontains="شطب") | Q(الحالة__icontains="مفصول") | Q(الحالة__icontains="منقطع"))
    return qs


def _list_context(request, model_cls, program: str, graduates: bool = False):
    if not has_program_permission(request.user, program, "view"):
        deny_with_log(request, program, "view")
    title = build_program_title(program, PROGRAM_TITLES, graduates=graduates)

    cols = _base_cols_for_program(program)
    custom_fields = list(_active_custom_fields(program))
    cols += [(f"cf_{f.id}", f.label) for f in custom_fields]

    base_qs = _get_ordered_rows(model_cls, program, graduates=graduates)
    qs = _apply_advanced_filters(base_qs, request)

    paginator = Paginator(qs, 200)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)
    rows = list(page_obj.object_list)
    rows = _refresh_rows_live_semesters(rows, model_cls)
    rows = _attach_program_list_extra_columns(rows, model_cls, program, graduates=graduates)
    _attach_row_flags(rows)
    rows = _attach_custom_fields(rows, model_cls, program)
    page_obj.object_list = rows
    log_activity(request, "view", program=program, details=("عرض المتخرجين" if graduates else "عرض الحاليين"))

    filters = extract_list_filters(request.GET)

    def _filtered_options_qs(exclude_key: str = ""):
        """Build dependent filter choices from the current active list.

        When the user chooses a promotion, semester, year, status, or specialty,
        the other choice lists are rebuilt from the matching trainees only.
        The excluded key keeps the currently edited list from hiding its own
        alternatives completely.
        """
        option_filters = filters.copy()
        if exclude_key:
            option_filters[exclude_key] = ""
        # Keep search text because it is an intentional global filter.
        fake_request = type("FilterRequest", (), {"GET": option_filters})()
        return _apply_advanced_filters(base_qs.order_by(), fake_request)

    promotion_qs = _filtered_options_qs("promotion")
    promotion_options = list(
        دفعة.objects.filter(
            id__in=promotion_qs.exclude(الدفعة_id__isnull=True).values_list("الدفعة_id", flat=True).distinct(),
            مفعلة=True,
        )
        .order_by("-السنة", "-رقم_الدورة")
        .only("id", "اسم_الدفعة", "السنة")
    )
    year_options = list(
        _filtered_options_qs("year")
        .exclude(الدفعة__السنة__isnull=True)
        .values_list("الدفعة__السنة", flat=True)
        .distinct()
        .order_by("-الدفعة__السنة")
    )
    semester_options = build_semester_options(
        _filtered_options_qs("semester")
        .exclude(السداسي__isnull=True)
        .exclude(السداسي="")
        .values_list("السداسي", flat=True)
    )
    specialty_options = build_specialty_options(
        _filtered_options_qs("specialty")
        .exclude(التخصص__isnull=True)
        .exclude(التخصص="")
        .values_list("التخصص", flat=True)
    )

    return {
        "title": title,
        "rows": rows,
        "cols": cols,
        "program": program,
        "is_staff": request.user.is_staff,
        "graduates": graduates,
        "page_obj": page_obj,
        "paginator": paginator,
        "filters": filters,
        "promotion_options": promotion_options,
        "year_options": year_options,
        "semester_options": semester_options,
        "specialty_options": specialty_options,
        "query_string": build_query_string_without_page(request.GET),
        "list_url": reverse(_program_list_url(program)),
        "graduates_url": reverse("program_graduates_list", args=[program]),
        "can_add": has_program_permission(request.user, program, "add"),
        "can_change": has_program_permission(request.user, program, "change"),
        "can_delete": has_program_permission(request.user, program, "delete"),
        "can_export": can_export_for_user(request.user),
    }


@login_required
def initial_list(request):
    return render(request, "trainees/list.html", _list_context(request, حضوري_أولي, "initial", graduates=False))


@login_required
def apprentice_list(request):
    return render(request, "trainees/list.html", _list_context(request, تمهين, "apprentice", graduates=False))


@login_required
def evening_list(request):
    return render(request, "trainees/list.html", _list_context(request, مسائي_ومعابر, "evening", graduates=False))


@login_required
def crossing_list(request):
    return render(request, "trainees/list.html", _list_context(request, مسائي_ومعابر, "crossing", graduates=False))


@login_required
def program_graduates_list(request, program):
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return HttpResponseBadRequest("برنامج غير صالح")
    return render(request, "trainees/list.html", _list_context(request, ModelCls, program, graduates=True))


def _visible_value(obj, field_name: str):
    if field_name == "__actions__":
        return ""
    if field_name.startswith("cf_"):
        return getattr(obj, "cf", {}).get(field_name, "")
    value = getattr(obj, field_name, "")
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    if field_name == "الدفعة" and value:
        return str(value)
    if isinstance(value, bool):
        return "نعم" if value else "لا"
    return value or ""


def _build_export_rows(model_cls, program: str, request, graduates: bool = False):
    context = _list_context(request, model_cls, program, graduates=graduates)
    qs = _apply_advanced_filters(_get_ordered_rows(model_cls, program, graduates=graduates), request)
    rows = list(qs.select_related("الدفعة")[:5000])
    rows = _refresh_rows_live_semesters(rows, model_cls)
    rows = _attach_program_list_extra_columns(rows, model_cls, program, graduates=graduates)
    rows = _attach_custom_fields(rows, model_cls, program)
    return context["cols"], rows, context["title"]


def _export_excel(title: str, cols, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "البيانات"
    headers = [label for field, label in cols if field != "__actions__"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for obj in rows:
        ws.append([_visible_value(obj, field) for field, _ in cols if field != "__actions__"])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 30)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{title}.xlsx"'
    return response


def _register_pdf_font():
    candidates = [
        Path(settings.BASE_DIR) / "fonts" / "Amiri-Regular.ttf",
        Path(settings.BASE_DIR) / "fonts" / "NotoNaskhArabic-Regular.ttf",
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/tahoma.ttf"),
        Path("C:/Windows/Fonts/tradbdo.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
    ]
    for candidate in candidates:
        try:
            path = str(candidate)
            if not Path(path).exists():
                continue
            if "ArabicUI" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("ArabicUI", path))
            return "ArabicUI"
        except Exception:
            continue
    return "Helvetica"


def _pdf_text(value):
    if value is None:
        return ""
    text = str(value)
    if not text:
        return ""
    try:
        if arabic_reshaper is not None and get_display is not None and any("؀" <= ch <= "ۿ" for ch in text):
            return get_display(arabic_reshaper.reshape(text))
    except Exception:
        pass
    return text


def _pdf_row(values):
    result = []
    for value in values:
        if isinstance(value, str):
            result.append(_pdf_text(value))
        else:
            result.append(value)
    return result


def _export_pdf(title: str, cols, rows):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=20, bottomMargin=20)
    font_name = _register_pdf_font()
    styles = getSampleStyleSheet()
    styles["Normal"].fontName = font_name
    styles["Title"].fontName = font_name
    pdf_cols = [(field, label) for field, label in cols if field in {"الرقم_التعريفي", "اللقب", "الاسم", "التخصص", "رقم_التسجيل", "السداسي", "الحالة", "تاريخ_بداية_التكوين", "تاريخ_نهاية_التكوين"}]
    data = [[label for _, label in pdf_cols]]
    for obj in rows[:1500]:
        data.append([str(_visible_value(obj, field)) for field, _ in pdf_cols])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#EAF1F8")]),
    ]))
    story = [Paragraph(title, styles["Title"]), Spacer(1, 10), table]
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title}.pdf"'
    return response


@login_required
def export_program_data(request, program, fmt):
    if not has_program_permission(request.user, program, "view"):
        deny_with_log(request, program, "export")
    profile = getattr(request.user, "access_profile", None)
    if not (request.user.is_superuser or (profile and (profile.can_export_data or profile.can_manage_all_programs))):
        deny_with_log(request, program, "export")
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return HttpResponseBadRequest("برنامج غير صالح")
    graduates = request.GET.get("graduates") == "1"
    cols, rows, title = _build_export_rows(ModelCls, program, request, graduates=graduates)
    log_activity(request, "export", program=program, details=f"تصدير {fmt}")
    if fmt == "excel":
        return _export_excel(title, cols, rows)
    if fmt == "pdf":
        return _export_pdf(title, cols, rows)
    return HttpResponseBadRequest("صيغة غير صالحة")


@login_required  # سطر كود لتنفيذ منطق/إعداد
def trainee_add(request, program):
    if not has_program_permission(request.user, program, "add"):
        deny_with_log(request, program, "add")
    FormCls = FORM_BY_PROGRAM.get(program)
    if not FormCls:  # شرط (If)
        return HttpResponseBadRequest("برنامج غير صالح")  # إرجاع قيمة من الدالة

    form = FormCls(request.POST or None)  # تعيين قيمة لمتغير/إعداد

    if request.method == "POST" and form.is_valid():  # شرط (If)
        obj = form.save(commit=False)
        _force_evening_type_for_program(obj, program)
        obj.save()
        form.save_m2m()
        _save_custom_fields_from_post(request, obj, program)
        dismissal_sync = _sync_auto_dismissal_decision_for_trainee(program, obj, request.user)
        log_activity(request, "add", program=program, obj=obj, details="إضافة سجل")
        messages.success(request, "تم حفظ البيانات بنجاح ✅")  # إظهار رسالة للمستخدم (نجاح/خطأ) في Django
        if dismissal_sync.get("created"):
            messages.success(request, f"تم إنشاء مقرر الفصل رقم {dismissal_sync['decision'].decision_number} تلقائيًا.")
        elif dismissal_sync.get("restored"):
            messages.success(request, f"تم استرجاع مقرر الفصل رقم {dismissal_sync['decision'].decision_number} من الأرشيف تلقائيًا.")
        elif dismissal_sync.get("archived"):
            messages.info(request, f"تم أرشفة {dismissal_sync['archived']} مقرر فصل سابق لهذا المتكون.")

        action = request.POST.get("action") or "list"  # تعيين قيمة لمتغير/إعداد
        if action == "stay":  # شرط (If)
            return redirect("trainee_edit", program=program, pk=obj.pk) if _has_named_url("trainee_edit") else redirect(_program_list_url(program))  # إرجاع قيمة من الدالة
        if action == "back":  # شرط (If)
            return redirect(request.META.get("HTTP_REFERER") or _program_list_url(program))  # إرجاع قيمة من الدالة

        return redirect(_program_list_url(program))  # إرجاع قيمة من الدالة

    # عرض صفحة الإضافة
    custom_fields = _active_custom_fields(program)  # تعيين قيمة لمتغير/إعداد
    cf_pairs = [(f, "") for f in custom_fields]  # تعيين قيمة لمتغير/إعداد

    return render(  # إرجاع قيمة من الدالة
        request,  # سطر كود لتنفيذ منطق/إعداد
        "trainees/form.html",  # سطر كود لتنفيذ منطق/إعداد
        {  # سطر كود لتنفيذ منطق/إعداد
            "form": form,  # سطر كود لتنفيذ منطق/إعداد
            "program": program,  # سطر كود لتنفيذ منطق/إعداد
            "mode": "add",  # سطر كود لتنفيذ منطق/إعداد
            "title": "إضافة",  # سطر كود لتنفيذ منطق/إعداد
            "custom_fields": custom_fields,  # سطر كود لتنفيذ منطق/إعداد
            "cf_pairs": cf_pairs,  # سطر كود لتنفيذ منطق/إعداد
        },  # سطر كود لتنفيذ منطق/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد


@login_required  # سطر كود لتنفيذ منطق/إعداد
def trainee_edit(request, program, pk):
    if not has_program_permission(request.user, program, "change"):
        deny_with_log(request, program, "change")
    ModelCls = MODEL_BY_PROGRAM.get(program)
    FormCls = FORM_BY_PROGRAM.get(program)  # تعيين قيمة لمتغير/إعداد
    if not ModelCls or not FormCls:  # شرط (If)
        return HttpResponseBadRequest("برنامج غير صالح")  # إرجاع قيمة من الدالة

    obj = get_object_or_404(_filter_queryset_by_program(ModelCls.objects.all(), program), pk=pk)  # تعيين قيمة لمتغير/إعداد
    form = FormCls(request.POST or None, instance=obj)  # تعيين قيمة لمتغير/إعداد

    if request.method == "POST" and form.is_valid():  # شرط (If)
        obj = form.save(commit=False)
        _force_evening_type_for_program(obj, program)
        obj.save()
        form.save_m2m()
        _save_custom_fields_from_post(request, obj, program)
        dismissal_sync = _sync_auto_dismissal_decision_for_trainee(program, obj, request.user)
        log_activity(request, "change", program=program, obj=obj, details="تعديل سجل")
        messages.success(request, "تم تحديث البيانات بنجاح ✅")  # إظهار رسالة للمستخدم (نجاح/خطأ) في Django
        if dismissal_sync.get("created"):
            messages.success(request, f"تم إنشاء مقرر الفصل رقم {dismissal_sync['decision'].decision_number} تلقائيًا.")
        elif dismissal_sync.get("restored"):
            messages.success(request, f"تم استرجاع مقرر الفصل رقم {dismissal_sync['decision'].decision_number} من الأرشيف تلقائيًا.")
        elif dismissal_sync.get("archived"):
            messages.info(request, f"تم أرشفة {dismissal_sync['archived']} مقرر فصل سابق لهذا المتكون.")
        return _redirect_after_save(request, program, obj.pk, default="stay")  # إرجاع قيمة من الدالة

    # عرض صفحة التعديل
    custom_fields = _active_custom_fields(program)  # تعيين قيمة لمتغير/إعداد
    cf_initial = _load_custom_field_initial(obj, program)  # تعيين قيمة لمتغير/إعداد
    cf_pairs = [(f, cf_initial.get(f"id{f.id}", "")) for f in custom_fields]  # تعيين قيمة لمتغير/إعداد

    return render(  # إرجاع قيمة من الدالة
        request,  # سطر كود لتنفيذ منطق/إعداد
        "trainees/form.html",  # سطر كود لتنفيذ منطق/إعداد
        {  # سطر كود لتنفيذ منطق/إعداد
            "form": form,  # سطر كود لتنفيذ منطق/إعداد
            "program": program,  # سطر كود لتنفيذ منطق/إعداد
            "mode": "edit",  # سطر كود لتنفيذ منطق/إعداد
            "obj": obj,  # سطر كود لتنفيذ منطق/إعداد
            "title": "تعديل",  # سطر كود لتنفيذ منطق/إعداد
            "custom_fields": custom_fields,  # سطر كود لتنفيذ منطق/إعداد
            "cf_pairs": cf_pairs,  # سطر كود لتنفيذ منطق/إعداد
        },  # سطر كود لتنفيذ منطق/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد



@login_required
def trainee_media_upload(request, program, pk):
    if request.method != "POST":
        return HttpResponseBadRequest("طريقة غير مدعومة")
    if not has_program_permission(request.user, program, "change"):
        deny_with_log(request, program, "change")
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return HttpResponseBadRequest("برنامج غير صالح")
    obj = get_object_or_404(_filter_queryset_by_program(ModelCls.objects.all(), program), pk=pk)

    photo_file = request.FILES.get("photo_file")
    qr_file = request.FILES.get("qr_file")

    if not photo_file and not qr_file:
        messages.warning(request, "اختر صورة المتكوّن أو صورة QR_Code على الأقل.")
        return redirect("trainee_edit", program=program, pk=pk)

    if photo_file:
        save_uploaded_media(obj, program, photo_file, "صور")
    if qr_file:
        save_uploaded_media(obj, program, qr_file, "QR_Code")

    details = []
    if photo_file:
        details.append("رفع/تغيير الصورة الشخصية")
    if qr_file:
        details.append("رفع/تغيير QR_Code")
    log_activity(request, "change", program=program, obj=obj, details=" + ".join(details) or "رفع وسائط")
    messages.success(request, "تم حفظ الوسائط بنجاح ✅")
    return redirect("trainee_edit", program=program, pk=pk)


@login_required  # سطر كود لتنفيذ منطق/إعداد
def trainee_delete(request, program, pk):
    if not has_program_permission(request.user, program, "delete"):
        deny_with_log(request, program, "delete")
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:  # شرط (If)
        return HttpResponseBadRequest("برنامج غير صالح")  # إرجاع قيمة من الدالة
    obj = get_object_or_404(_filter_queryset_by_program(ModelCls.objects.all(), program), pk=pk)  # تعيين قيمة لمتغير/إعداد
    if request.method == "POST":
        log_activity(request, "delete", program=program, obj=obj, details="حذف سجل")
        obj.delete()
        messages.success(request, "تم الحذف بنجاح ✅")  # إظهار رسالة للمستخدم (نجاح/خطأ) في Django
        return redirect(_program_list_url(program))  # إرجاع قيمة من الدالة
    return render(request, "trainees/confirm_delete.html", {"obj": obj, "program": program, "title": "حذف"})  # إرجاع قيمة من الدالة

def _program_list_url(program: str):  # تعريف دالة (Function)
    return {  # إرجاع قيمة من الدالة
        "initial": "initial_list",  # سطر كود لتنفيذ منطق/إعداد
        "apprentice": "apprentice_list",  # سطر كود لتنفيذ منطق/إعداد
        "evening": "evening_list",  # سطر كود لتنفيذ منطق/إعداد
        "crossing": "crossing_list",  # سطر كود لتنفيذ منطق/إعداد
    }.get(program, "dashboard")  # سطر كود لتنفيذ منطق/إعداد


# -----------------------------
# UI emergency button: recompute all semesters
# -----------------------------


@login_required  # سطر كود لتنفيذ منطق/إعداد
@require_POST  # سطر كود لتنفيذ منطق/إعداد
def recompute_semesters_ui(request):  # تعريف دالة (Function)
    # سطر كود لتنفيذ منطق/إعداد
    """زر طوارئ في الواجهة لإعادة حساب السداسيات.

    - متاح فقط للمستخدمين staff  # سطر كود لتنفيذ منطق/إعداد
    - يعيد حساب السداسي لكل البرامج (حضوري/تمهين/مسائي)  # سطر كود لتنفيذ منطق/إعداد
    # سطر كود لتنفيذ منطق/إعداد
    """
    if not can_access_admin_panel(request.user):
        deny_with_log(request, "", "recompute")

    _recompute_semesters(حضوري_أولي)  # سطر كود لتنفيذ منطق/إعداد
    _recompute_semesters(تمهين)  # سطر كود لتنفيذ منطق/إعداد
    _recompute_semesters(مسائي_ومعابر)  # سطر كود لتنفيذ منطق/إعداد

    log_activity(request, "change", details="إعادة حساب السداسيات")
    messages.success(request, "تمت إعادة حساب السداسيات بنجاح.")
    # الرجوع للصفحة السابقة إن أمكن، وإلا إلى الرئيسية
    return redirect(request.META.get("HTTP_REFERER") or "dashboard")  # إرجاع قيمة من الدالة


def _active_custom_fields(program: str):  # تعريف دالة (Function)
    program = (program or "").strip().lower()  # تعيين قيمة لمتغير/إعداد
    qs = CustomField.objects.filter(active=True)  # تعيين قيمة لمتغير/إعداد
    program_filter = Q(program="all") | Q(program=program)
    if program == "crossing":
        program_filter |= Q(program="evening")
    return qs.filter(program_filter).order_by("order", "id")  # إرجاع قيمة من الدالة




def _load_custom_field_initial(obj, program: str):  # تعريف دالة (Function)
    custom_fields = list(_active_custom_fields(program))  # تعيين قيمة لمتغير/إعداد
    ct = ContentType.objects.get_for_model(obj.__class__)  # تعيين قيمة لمتغير/إعداد
    vals = CustomFieldValue.objects.filter(content_type=ct, object_id=obj.pk, field__in=custom_fields).select_related("field")  # تعيين قيمة لمتغير/إعداد

    initial = {}  # تعيين قيمة لمتغير/إعداد
    for v in vals:  # حلقة تكرار (For)
        # نخزن دائمًا كنص (ISO للتاريخ، 1/0 للمنطقي)
        initial[f"id{v.field_id}"] = v.value_text or ""  # تعيين قيمة لمتغير/إعداد
    return initial  # إرجاع قيمة من الدالة




def _save_custom_fields_from_post(request, obj, program: str):  # تعريف دالة (Function)
    custom_fields = list(_active_custom_fields(program))  # تعيين قيمة لمتغير/إعداد
    ct = ContentType.objects.get_for_model(obj.__class__)  # تعيين قيمة لمتغير/إعداد

    for f in custom_fields:  # حلقة تكرار (For)
        post_key = f"cf_{f.id}"  # تعيين قيمة لمتغير/إعداد

        if f.field_type == "boolean":  # شرط (If)
            raw = "1" if request.POST.get(post_key) in ("1", "on", "true", "True") else "0"  # تعيين قيمة لمتغير/إعداد
        else:  # فرع بديل (Else)
            raw = (request.POST.get(post_key, "") or "").strip()  # تعيين قيمة لمتغير/إعداد

        CustomFieldValue.objects.update_or_create(  # سطر كود لتنفيذ منطق/إعداد
            field=f,  # تعيين قيمة لمتغير/إعداد
            content_type=ct,  # تعيين قيمة لمتغير/إعداد
            object_id=obj.pk,  # تعيين قيمة لمتغير/إعداد
            defaults={"value_text": raw},  # تعيين قيمة لمتغير/إعداد
        )  # سطر كود لتنفيذ منطق/إعداد




def _attach_custom_fields(rows, model_cls, program: str):  # تعريف دالة (Function)
    custom_fields = list(_active_custom_fields(program))  # تعيين قيمة لمتغير/إعداد
    if not rows:  # شرط (If)
        return rows  # إرجاع قيمة من الدالة

    ct = ContentType.objects.get_for_model(model_cls)  # تعيين قيمة لمتغير/إعداد
    obj_ids = [o.pk for o in rows]  # تعيين قيمة لمتغير/إعداد

    vals = (  # تعيين قيمة لمتغير/إعداد
        CustomFieldValue.objects  # سطر كود لتنفيذ منطق/إعداد
        .filter(content_type=ct, object_id__in=obj_ids, field__in=custom_fields)  # تعيين قيمة لمتغير/إعداد
        .values("object_id", "field_id", "value_text")  # سطر كود لتنفيذ منطق/إعداد
    )  # سطر كود لتنفيذ منطق/إعداد
    mp = {(v["object_id"], v["field_id"]): (v["value_text"] or "") for v in vals}  # تعيين قيمة لمتغير/إعداد

    for o in rows:  # حلقة تكرار (For)
        o.cf = {}  # تعيين قيمة لمتغير/إعداد
        for f in custom_fields:  # حلقة تكرار (For)
            v = mp.get((o.pk, f.id), "")  # تعيين قيمة لمتغير/إعداد
            if f.field_type == "boolean":  # شرط (If)
                o.cf[f"cf_{f.id}"] = "نعم" if v == "1" else ""  # تعيين قيمة لمتغير/إعداد
            else:  # فرع بديل (Else)
                o.cf[f"cf_{f.id}"] = v  # تعيين قيمة لمتغير/إعداد
    return rows  # إرجاع قيمة من الدالة



# -------------------------
# Admin quick links (used by training_center/urls.py)
# -------------------------
@login_required
def admin_quick_initial(request):
    if not can_access_admin_panel(request.user):
        deny_with_log(request, "initial", "admin")
    return redirect(_admin_changelist_url(حضوري_أولي))

@login_required
def admin_quick_apprentice(request):
    if not can_access_admin_panel(request.user):
        deny_with_log(request, "apprentice", "admin")
    return redirect(_admin_changelist_url(تمهين))

@login_required
def admin_quick_evening(request):
    if not can_access_admin_panel(request.user):
        deny_with_log(request, "evening", "admin")
    return redirect(_admin_changelist_filtered(مسائي_ومعابر, **{"نوع_التكوين__exact": EVENING_TRAINING_TYPE_EVENING}))

@login_required
def admin_quick_crossing(request):
    if not can_access_admin_panel(request.user):
        deny_with_log(request, "crossing", "admin")
    return redirect(_admin_changelist_filtered(مسائي_ومعابر, **{"نوع_التكوين__exact": EVENING_TRAINING_TYPE_CROSSING}))

@login_required
def admin_quick_filtered(request, program, status):
    if not can_access_admin_panel(request.user):
        deny_with_log(request, program, "admin")
    # سطر كود لتنفيذ منطق/إعداد
    """Quick admin view: open Django admin changelist for a program with unified status filters."""
    model_cls = MODEL_BY_PROGRAM.get(program)  # تعيين قيمة لمتغير/إعداد
    if not model_cls:  # شرط (If)
        return HttpResponseBadRequest("برنامج غير صالح")  # إرجاع قيمة من الدالة

    status = (status or "").strip().lower()  # تعيين قيمة لمتغير/إعداد
    base_filters = {}
    training_type = _evening_type_for_program(program)
    if training_type:
        base_filters["نوع_التكوين__exact"] = training_type

    # all
    if status in ("all", "*", "كل", "الكل"):  # شرط (If)
        return redirect(_admin_changelist_filtered(model_cls, **base_filters) if base_filters else _admin_changelist_url(model_cls))  # إرجاع قيمة من الدالة

    # unified admin filter parameter_name = status_group
    if status in ("active", "نشط"):  # شرط (If)
        return redirect(_admin_changelist_filtered(model_cls, status_group="active", **base_filters))  # إرجاع قيمة من الدالة
    if status in ("removed", "مشطوب"):  # شرط (If)
        return redirect(_admin_changelist_filtered(model_cls, status_group="removed", **base_filters))  # إرجاع قيمة من الدالة
    if status in ("recent_removed", "مشطوب_حديثًا", "مشطوب_حديثا", "حديث"):
        return redirect(_admin_changelist_filtered(model_cls, status_group="recent_removed", **base_filters))  # إرجاع قيمة من الدالة

    if status in ("repeater", "معيد") and program == "apprentice":  # شرط (If)
        return redirect(_admin_changelist_filtered(model_cls, **base_filters, **{"معيد__exact": "1"}))  # إرجاع قيمة من الدالة

    # fallback: use admin search
    return redirect(_admin_changelist_filtered(model_cls, q=status, **base_filters))  # إرجاع قيمة من الدالة


ATTENDANCE_PROGRAMS = {
    "initial": {
        "label": "الحضوري الأولي",
        "description": "الدراسة من الأحد إلى الخميس.",
        "weekday_numbers": [6, 0, 1, 2, 3],
    },
    "apprentice": {
        "label": "التمهين",
        "description": "اختر يومي الدراسة مع إمكانية إضافة يوم ثالث حسب التخصص أو الفوج.",
        "weekday_numbers": [],
    },
    "evening": {
        "label": "الدروس المسائية",
        "description": "اختر أيام الدراسة الخاصة بالفوج قبل عرض جدول الغيابات.",
        "weekday_numbers": [],
    },
    "crossing": {
        "label": "المعابر",
        "description": "تكوين سنة واحدة؛ اختر أيام الدراسة الخاصة بالفوج.",
        "weekday_numbers": [],
    },
}


ATTENDANCE_MONTH_CHOICES = [
    (1, "جانفي 01"),
    (2, "فيفري 02"),
    (3, "مارس 03"),
    (4, "أفريل 04"),
    (5, "ماي 05"),
    (6, "جوان 06"),
    (7, "جويلية 07"),
    (8, "أوت 08"),
    (9, "سبتمبر 09"),
    (10, "أكتوبر 10"),
    (11, "نوفمبر 11"),
    (12, "ديسمبر 12"),
]

ATTENDANCE_MODEL_MAP = {
    "initial": حضوري_أولي,
    "apprentice": تمهين,
    "evening": مسائي_ومعابر,
    "crossing": مسائي_ومعابر,
}

ATTENDANCE_EXPORT_TITLE_MAP = {
    "table": "جدول الغيابات",
    "stats": "إحصائيات الغيابات",
    "saved_stats": "أرشيف إحصائيات الغيابات",
}


def _attendance_month_label(month: int) -> str:
    return dict(ATTENDANCE_MONTH_CHOICES).get(month, str(month))


def _safe_sheet_title(value: str, fallback: str = "ورقة") -> str:
    value = (value or "").strip()
    if not value:
        value = fallback
    forbidden = "[]:*?/\\"
    for ch in forbidden:
        value = value.replace(ch, ' ')
    value = ' '.join(value.split())
    return (value[:31] or fallback)


def _set_download_filename(response, filename: str):
    from django.utils.http import content_disposition_header

    filename = (filename or "download").strip()
    invalid_chars = '<>:"/\\|?*'
    cleaned = []
    for ch in filename:
        cleaned.append(' - ' if ch in {'/', '\\'} else ('_' if ch in invalid_chars else ch))
    filename = ' '.join(''.join(cleaned).split()) or 'download'

    response['Content-Disposition'] = content_disposition_header(True, filename)
    return response

def _attendance_export_filename(kind: str, program: str, scope, specialty: str = "", ext: str = "xlsx") -> str:
    program_label = ATTENDANCE_PROGRAMS[program]["label"]
    month_label = _attendance_month_label(scope["month"])
    parts = [ATTENDANCE_EXPORT_TITLE_MAP.get(kind, "تصدير"), program_label, month_label, str(scope["year"])]
    specialty = (specialty or "").strip()
    if specialty:
        parts.append(specialty)
    return " - ".join(parts) + f".{ext}"


def _group_items_by_specialty(items, accessor):
    grouped = {}
    order = []
    for item in items:
        specialty = (accessor(item) or "بدون تخصص").strip() or "بدون تخصص"
        if specialty not in grouped:
            grouped[specialty] = []
            order.append(specialty)
        grouped[specialty].append(item)
    return [(specialty, grouped[specialty]) for specialty in order]


def _finalize_workbook_response(wb, filename: str):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return _set_download_filename(response, filename)


ATTENDANCE_EXPORT_HEADER_LINES = [
    "الجمهورية الجزائرية الديمقراطية الشعبية",
    "وزارة التكوين والتعليم المهنيين",
    "ولاية تيسمسيلت",
    "المعهد الوطني المتخصص في التكوين المهني",
    "الشهيد تاج الدين حامد عبد الوهاب تيسمسيلت",
    "المديرية الفرعية للإعلام والتوجيه والرقمنة والمساعدة على الإدماج المهني",
    "مصلحة المراقبة العامة",
]


def _attendance_rows_semester_label(rows):
    labels = []
    seen = set()
    for row in rows or []:
        trainee = row.get("trainee") if isinstance(row, dict) else None
        label = (getattr(trainee, "السداسي", "") or "").strip()
        if label and label not in seen:
            seen.add(label)
            labels.append(label)
    if not labels:
        return ""
    if len(labels) == 1:
        return labels[0]
    return "، ".join(labels)


def _attendance_scope_subtitle(scope, specialty_label="", semester_label=""):
    parts = []
    month_label = _attendance_month_label(scope["month"])
    parts.append(f"الشهر: {month_label} {scope['year']}")
    semester_label = (semester_label or scope.get("semester") or "").strip()
    if semester_label:
        parts.append(f"السداسي: {semester_label}")
    return " - ".join([part for part in parts if part])


def _attendance_template_title(program, specialty_label=""):
    specialty_label = (specialty_label or "").strip() or "كل التخصصات"
    if program == "apprentice":
        return f"وثيقة غيابات المتمهنين فرع {specialty_label}"
    return f"وثيقة غيابات المتربصين فرع {specialty_label}"


def _attendance_stats_template_title(program, specialty_label=""):
    base = "إحصائيات غيابات المتمهنين فرع" if program == "apprentice" else "إحصائيات غيابات المتربصين فرع"
    specialty_label = (specialty_label or "").strip()
    return f"{base} {specialty_label}".strip()


def _attendance_stats_scope_subtitle(scope, specialty_label="", semester_label="", batch_display=""):
    parts = []
    month_label = _attendance_month_label(scope["month"])
    parts.append(f"الشهر: {month_label} {scope['year']}")
    semester_label = (semester_label or scope.get("semester") or "").strip()
    if semester_label:
        parts.append(f"السداسي: {semester_label}")
    batch_display = (batch_display or "").strip()
    if batch_display:
        parts.append(batch_display)
    return " - ".join([part for part in parts if part])


def _apply_official_stats_excel_header(ws, total_columns, program, scope, specialty_label="", semester_label="", batch_display=""):
    end_col = max(1, total_columns)
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9E2F3")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, line in enumerate(ATTENDANCE_EXPORT_HEADER_LINES, start=1):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=end_col)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = line
        cell.font = Font(bold=True, size=13 if row_idx <= 2 else 11, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = white_fill

    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=end_col)
    title_cell = ws.cell(row=8, column=1)
    title_cell.value = _attendance_stats_template_title(program, specialty_label)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = dark_fill

    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=end_col)
    subtitle_cell = ws.cell(row=9, column=1)
    subtitle_cell.value = _attendance_stats_scope_subtitle(scope, specialty_label=specialty_label, semester_label=semester_label, batch_display=batch_display)
    subtitle_cell.font = Font(bold=True, size=11, color="000000")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    subtitle_cell.fill = subtitle_fill

    for row_idx in range(1, 10):
        ws.row_dimensions[row_idx].height = 18 if row_idx < 8 else 21


def _apply_official_header(ws, total_columns, program, scope, specialty_label="", semester_label=""):
    end_col = max(1, total_columns)
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    line_fill = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, line in enumerate(ATTENDANCE_EXPORT_HEADER_LINES, start=1):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=end_col)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = line
        cell.font = Font(bold=True, size=13 if row_idx <= 2 else 11, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = line_fill

    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=end_col)
    title_cell = ws.cell(row=8, column=1)
    title_cell.value = _attendance_template_title(program, specialty_label)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = dark_fill

    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=end_col)
    subtitle_cell = ws.cell(row=9, column=1)
    subtitle_cell.value = _attendance_scope_subtitle(scope, specialty_label=specialty_label, semester_label=semester_label)
    subtitle_cell.font = Font(bold=True, size=11, color="000000")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    subtitle_cell.fill = PatternFill("solid", fgColor="D9E2F3")

    for row_idx in range(1, 10):
        ws.row_dimensions[row_idx].height = 18 if row_idx < 8 else 21


def _build_attendance_export_workbook(program, payload):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    rows = payload["rows"]
    scope = payload["scope"]
    slot_count = payload.get("slot_count", 1)
    columns = payload["columns"]
    label_map = _attendance_status_label_map(program)
    show_all_specialties = payload["show_all_specialties"]
    specialty_value = (scope.get("specialty") or "").strip()

    grouped_rows = [(specialty_value or "كل التخصصات", rows)]
    if show_all_specialties:
        grouped_rows = _group_items_by_specialty(rows, lambda row: getattr(row["trainee"], "التخصص", "") or "بدون تخصص")

    for idx, (specialty_name, specialty_rows) in enumerate(grouped_rows, start=1):
        ws = wb.create_sheet(title=_safe_sheet_title(specialty_name or f"تخصص {idx}"))
        ws.sheet_view.rightToLeft = True
        ws.freeze_panes = "C12"

        show_specialty_column = bool(show_all_specialties)
        base_header = ["الرقم", "الاسم و اللقب"]
        if show_specialty_column:
            base_header.append("التخصص")
        base_col_count = len(base_header)
        total_columns = base_col_count + (len(columns) * slot_count)

        semester_label = _attendance_rows_semester_label(specialty_rows)
        _apply_official_header(ws, total_columns, program, scope, specialty_label=specialty_name, semester_label=semester_label)

        header_row_1 = 10
        header_row_2 = 11
        data_start_row = 12

        dark_fill = PatternFill("solid", fgColor="1F4E78")
        light_fill = PatternFill("solid", fgColor="D9E2F3")
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for idx_col, title in enumerate(base_header, start=1):
            ws.merge_cells(start_row=header_row_1, start_column=idx_col, end_row=header_row_2, end_column=idx_col)
            cell = ws.cell(row=header_row_1, column=idx_col)
            cell.value = title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = dark_fill

        current_col = base_col_count + 1
        for col in columns:
            if slot_count > 1:
                ws.merge_cells(start_row=header_row_1, start_column=current_col, end_row=header_row_1, end_column=current_col + slot_count - 1)
                ws.merge_cells(start_row=header_row_2, start_column=current_col, end_row=header_row_2, end_column=current_col + slot_count - 1)
            ws.cell(row=header_row_1, column=current_col).value = col["weekday_label"]
            ws.cell(row=header_row_2, column=current_col).value = str(col["day_num"]).zfill(2)
            for offset in range(slot_count):
                for row_num in (header_row_1, header_row_2):
                    cell = ws.cell(row=row_num, column=current_col + offset)
                    cell.font = Font(bold=True, color="FFFFFF" if row_num == header_row_1 else "000000")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = dark_fill if row_num == header_row_1 else light_fill
                    cell.border = border
            current_col += slot_count

        for row_num in (header_row_1, header_row_2):
            for col_num in range(1, total_columns + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                if col_num <= base_col_count:
                    cell.fill = dark_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif row_num == header_row_2:
                    cell.fill = light_fill
                    cell.font = Font(bold=True, color="000000")

        for row_idx, row in enumerate(specialty_rows, start=data_start_row):
            values = [str(row["index"]).zfill(2), f"{row['trainee'].اللقب} {row['trainee'].الاسم}"]
            if show_specialty_column:
                values.append(getattr(row["trainee"], "التخصص", "") or "")
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            current_col = base_col_count + 1
            for day_cell in row["cells"]:
                if program == "apprentice":
                    slots = day_cell.get("slots", [])
                    for slot_index in range(slot_count):
                        status = slots[slot_index].get("status", "") if slot_index < len(slots) else ""
                        cell = ws.cell(row=row_idx, column=current_col)
                        cell.value = label_map.get(status, "")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border
                        current_col += 1
                else:
                    cell = ws.cell(row=row_idx, column=current_col)
                    cell.value = label_map.get(day_cell.get("status", ""), "")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                    current_col += 1
            ws.row_dimensions[row_idx].height = 19

        from openpyxl.utils import get_column_letter

        ws.column_dimensions["A"].width = 5.0
        ws.column_dimensions["B"].width = 26.0
        next_col = 3
        if show_specialty_column:
            ws.column_dimensions["C"].width = 24.0
            next_col = 4
        for col_num in range(next_col, total_columns + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 4.5 if slot_count > 1 else 5.6

        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 8 if slot_count > 1 else 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.24
        ws.page_margins.right = 0.24
        ws.page_margins.top = 0.2
        ws.page_margins.bottom = 0.2
        ws.print_options.horizontalCentered = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "$1:$11"
        ws.print_area = f"$A$1:${get_column_letter(total_columns)}${max(data_start_row, ws.max_row)}"

    return wb


def _attendance_workbook_to_pdf_bytes(wb):
    import subprocess
    import tempfile

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "attendance_export.xlsx")
        pdf_path = os.path.join(tmpdir, "attendance_export.pdf")
        wb.save(xlsx_path)
        command = [
            shutil.which("libreoffice") or shutil.which("soffice") or "libreoffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            tmpdir,
            xlsx_path,
        ]
        subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        with open(pdf_path, "rb") as handle:
            return handle.read()


AR_WEEKDAYS = {
    0: "الإثنين",
    1: "الثلاثاء",
    2: "الأربعاء",
    3: "الخميس",
    4: "الجمعة",
    5: "السبت",
    6: "الأحد",
}

# في الجداول ذات الأيام المختارة نستعمل أيام العمل من السبت إلى الخميس فقط.
# ترقيم Python calendar: الإثنين=0 ... الأحد=6، لذلك السبت=5.
CUSTOM_ATTENDANCE_WEEKDAY_VALUES = [5, 6, 0, 1, 2, 3]


def _attendance_allowed_programs(user):
    return visible_programs(user)


def _attendance_month_dates(year: int, month: int, weekday_numbers):
    result = []
    month_matrix = calendar.monthcalendar(year, month)
    for week in month_matrix:
        for weekday_idx, day_num in enumerate(week):
            if not day_num:
                continue
            if weekday_idx in weekday_numbers:
                current = date(year, month, day_num)
                result.append({
                    "date": current,
                    "day_num": day_num,
                    "weekday": weekday_idx,
                    "weekday_label": AR_WEEKDAYS.get(weekday_idx, str(weekday_idx)),
                    "iso": current.isoformat(),
                })
    return result



def _is_bridge_specialty(specialty: str) -> bool:
    return "معابر" in ((specialty or "").strip())


def _attendance_scope(program, request):
    today = timezone.localdate()
    month = int(request.GET.get("month") or request.POST.get("month") or today.month)
    year = int(request.GET.get("year") or request.POST.get("year") or today.year)
    specialty = (request.GET.get("specialty") or request.POST.get("specialty") or "").strip()
    promotion_id = request.GET.get("promotion") or request.POST.get("promotion") or ""
    promotion = None
    if promotion_id:
        try:
            promotion = دفعة.objects.get(pk=promotion_id)
        except دفعة.DoesNotExist:
            promotion = None
    weekday1 = request.GET.get("weekday1") or request.POST.get("weekday1") or ""
    weekday2 = request.GET.get("weekday2") or request.POST.get("weekday2") or ""
    weekday3 = request.GET.get("weekday3") or request.POST.get("weekday3") or ""
    weekday4 = request.GET.get("weekday4") or request.POST.get("weekday4") or ""
    weekday5 = request.GET.get("weekday5") or request.POST.get("weekday5") or ""
    show_table = (request.GET.get("show_table") or request.POST.get("show_table") or "").strip()
    action = (request.GET.get("action") or request.POST.get("action") or "").strip()
    return {
        "month": month,
        "year": year,
        "specialty": specialty,
        "promotion": promotion,
        "promotion_obj": promotion,
        "promotion_id": str(promotion.pk) if promotion else "",
        "weekday1": weekday1,
        "weekday2": weekday2,
        "weekday3": weekday3,
        "weekday4": weekday4,
        "weekday5": weekday5,
        "show_table": show_table == "1",
        "action": action,
    }


def _attendance_specialty_options(model_cls, promotion=None, program=None):
    today = timezone.localdate()
    qs = _filter_queryset_by_program(model_cls.objects, program or "").filter(
        Q(تاريخ_نهاية_التكوين__isnull=True) | Q(تاريخ_نهاية_التكوين__gt=today)
    ).exclude(
        Q(الحالة__icontains="مشطوب") |
        Q(الحالة__icontains="شطب") |
        Q(الحالة__icontains="مفصول") |
        Q(الحالة__icontains="منقطع") |
        Q(الحالة__icontains="متوقف") |
        Q(الحالة__icontains="موقوف") |
        Q(الحالة__icontains="انسحب")
    )
    if promotion:
        qs = qs.filter(الدفعة=promotion)
    return sorted(
        _unique_clean_values(
            qs.order_by()
              .exclude(التخصص__isnull=True)
              .exclude(التخصص__exact="")
              .values_list("التخصص", flat=True)
        )
    )


def _attendance_promotion_options(model_cls, program=None):
    today = timezone.localdate()
    promotion_ids = (
        _filter_queryset_by_program(model_cls.objects, program or "").filter(
            Q(تاريخ_نهاية_التكوين__isnull=True) | Q(تاريخ_نهاية_التكوين__gt=today)
        ).exclude(
            Q(الحالة__icontains="مشطوب") |
            Q(الحالة__icontains="شطب") |
            Q(الحالة__icontains="مفصول") |
            Q(الحالة__icontains="منقطع") |
            Q(الحالة__icontains="متوقف") |
            Q(الحالة__icontains="موقوف") |
            Q(الحالة__icontains="انسحب")
        ).exclude(الدفعة__isnull=True)
         .values_list("الدفعة_id", flat=True)
         .distinct()
    )
    return دفعة.objects.filter(id__in=promotion_ids, مفعلة=True).order_by("-السنة", "-رقم_الدورة")


def _attendance_queryset(program, scope):
    model_cls = ATTENDANCE_MODEL_MAP[program]
    today = timezone.localdate()
    qs = _filter_queryset_by_program(model_cls.objects, program).filter(
        Q(تاريخ_نهاية_التكوين__isnull=True) | Q(تاريخ_نهاية_التكوين__gt=today)
    ).exclude(
        Q(الحالة__icontains="مشطوب") |
        Q(الحالة__icontains="شطب") |
        Q(الحالة__icontains="مفصول") |
        Q(الحالة__icontains="منقطع") |
        Q(الحالة__icontains="متوقف") |
        Q(الحالة__icontains="موقوف") |
        Q(الحالة__icontains="انسحب")
    )
    if scope["promotion"]:
        qs = qs.filter(الدفعة=scope["promotion"])
    if scope["specialty"]:
        qs = qs.filter(التخصص=scope["specialty"])
    qs = qs.only(
        "id", "الاسم", "اللقب", "التخصص", "الدفعة_id",
        "تاريخ_نهاية_التكوين", "الحالة"
    ).order_by("التخصص", "اللقب", "الاسم")
    return model_cls, qs


def _attendance_visual_slot_count(program):
    return 4 if program == "apprentice" else 1


def _attendance_status_choices(program):
    if program == "apprentice":
        return [
            ("present", "ح"),
            ("absent", "غ"),
            ("excused", "ع"),
            ("late", "ت"),
        ]
    return list(خليةغياب.STATUS_CHOICES)


def _attendance_status_label_map(program):
    return dict(_attendance_status_choices(program))


def _attendance_weekdays(program, scope):
    weekdays = []

    if _program_requires_custom_weekdays(program, scope.get("specialty", "")):
        for key in _custom_weekday_keys(program):
            value = scope.get(key, "")
            if value == "":
                continue
            try:
                num = int(value)
            except (TypeError, ValueError):
                continue
            if num not in weekdays:
                weekdays.append(num)
        return weekdays

    return ATTENDANCE_PROGRAMS[program]["weekday_numbers"]


def _attendance_weekday_validation(program, scope, requires_custom_weekdays=False):
    if not _program_requires_custom_weekdays(program) and not requires_custom_weekdays:
        return {"is_valid": True, "message": "", "weekdays": ATTENDANCE_PROGRAMS[program]["weekday_numbers"]}

    raw_values = [scope.get(key, "") for key in _custom_weekday_keys(program)]

    cleaned = []
    provided = []
    invalid = False
    for value in raw_values:
        if value in (None, ""):
            continue
        provided.append(value)
        try:
            num = int(value)
        except (TypeError, ValueError):
            invalid = True
            continue
        if num not in cleaned:
            cleaned.append(num)

    min_required = 2
    max_allowed = _program_max_custom_weekdays(program)

    if invalid:
        return {"is_valid": False, "message": "قيمة يوم الدراسة غير صحيحة.", "weekdays": cleaned}
    if len(provided) != len(cleaned):
        return {"is_valid": False, "message": "لا يمكن اختيار نفس اليوم أكثر من مرة.", "weekdays": cleaned}
    if len(cleaned) < min_required:
        return {"is_valid": False, "message": "اختر أيام الدراسة قبل عرض الجدول.", "weekdays": cleaned}
    if len(cleaned) > max_allowed:
        return {"is_valid": False, "message": "عدد أيام الدراسة المحدد أكبر من المسموح.", "weekdays": cleaned}

    return {"is_valid": True, "message": "", "weekdays": cleaned}


def _attendance_stats_payload(program, request):
    payload = _attendance_table_payload(program, request)
    columns = payload["columns"]
    rows = payload["rows"]
    slot_count = payload.get("slot_count", 1)

    stats_rows = []
    totals_counter = Counter()

    for row in rows:
        counter = Counter()
        total_recorded = 0
        for cell in row.get("cells", []):
            if program == "apprentice":
                for slot in cell.get("slots", []):
                    status = (slot.get("status") or "").strip()
                    if not status:
                        continue
                    counter[status] += 1
                    total_recorded += 1
            else:
                status = (cell.get("status") or "").strip()
                if not status:
                    continue
                counter[status] += 1
                total_recorded += 1

        absence_rate = round((counter["absent"] / total_recorded) * 100, 2) if total_recorded else 0
        stats_rows.append({
            "index": row["index"],
            "trainee": row["trainee"],
            "present_count": counter["present"],
            "absent_count": counter["absent"],
            "excused_count": counter["excused"],
            "late_count": counter["late"],
            "total_recorded": total_recorded,
            "absence_rate": absence_rate,
        })
        totals_counter.update(counter)
        totals_counter["total_recorded"] += total_recorded

    stats_rows.sort(key=lambda item: (-item["absence_rate"], -item["absent_count"], getattr(item["trainee"], "التخصص", "") or "", getattr(item["trainee"], "اللقب", "") or "", getattr(item["trainee"], "الاسم", "") or ""))
    for index, row in enumerate(stats_rows, start=1):
        row["display_index"] = index

    trainee_count = len(stats_rows)
    average_absence_rate = round(sum(item["absence_rate"] for item in stats_rows) / trainee_count, 2) if trainee_count else 0

    payload.update({
        "stats_rows": stats_rows,
        "stats_totals": {
            "present_count": totals_counter["present"],
            "absent_count": totals_counter["absent"],
            "excused_count": totals_counter["excused"],
            "late_count": totals_counter["late"],
            "total_recorded": totals_counter["total_recorded"],
        },
        "average_absence_rate": average_absence_rate,
        "trainee_count": trainee_count,
        "displayed_days_count": len(columns),
    })
    return payload




def _attendance_stats_scope_filters(program, scope):
    return {
        "program": program,
        "year": scope["year"],
        "month": scope["month"],
        "batch": scope.get("promotion_obj"),
        "specialty": (scope.get("specialty") or ""),
    }



def _attendance_saved_stats_program_options(user):
    return [
        {
            "code": code,
            "label": ATTENDANCE_PROGRAMS[code]["label"],
        }
        for code in _attendance_allowed_programs(user)
        if code in ATTENDANCE_PROGRAMS
    ]


def _attendance_saved_stats_summary(program, scope):
    filters = _attendance_stats_scope_filters(program, scope)
    qs = AttendanceStatSnapshot.objects.filter(**filters)
    summary = qs.aggregate(row_count=Count("id"), latest_saved_at=Max("updated_at"), latest_created_at=Max("created_at"))
    return {
        "row_count": summary.get("row_count") or 0,
        "latest_saved_at": summary.get("latest_saved_at") or summary.get("latest_created_at"),
    }


def _save_attendance_stats_snapshot(program, payload, user):
    scope = payload["scope"]
    filters = _attendance_stats_scope_filters(program, scope)
    rows = payload.get("stats_rows") or []
    with transaction.atomic():
        AttendanceStatSnapshot.objects.filter(**filters).delete()
        objects = []
        for row in rows:
            trainee = row["trainee"]
            objects.append(AttendanceStatSnapshot(
                **filters,
                trainee_id=trainee.pk,
                trainee_name=f"{getattr(trainee, 'اللقب', '')} {getattr(trainee, 'الاسم', '')}".strip(),
                trainee_specialty=getattr(trainee, "التخصص", "") or "",
                present_count=row["present_count"],
                absent_count=row["absent_count"],
                excused_count=row["excused_count"],
                late_count=row["late_count"],
                total_recorded=row["total_recorded"],
                absence_rate=row["absence_rate"],
                saved_by=user if getattr(user, "is_authenticated", False) else None,
            ))
        if objects:
            AttendanceStatSnapshot.objects.bulk_create(objects, batch_size=500)
    return len(rows)


def _delete_attendance_stats_snapshot(program, scope):
    filters = _attendance_stats_scope_filters(program, scope)
    deleted_count, _ = AttendanceStatSnapshot.objects.filter(**filters).delete()
    return deleted_count


def _delete_old_attendance_stats(program, cutoff_year, cutoff_month):
    deleted_count, _ = AttendanceStatSnapshot.objects.filter(program=program).filter(
        Q(year__lt=cutoff_year) | Q(year=cutoff_year, month__lt=cutoff_month)
    ).delete()
    return deleted_count

def _attendance_table_payload(program, request):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()

    scope = _attendance_scope(program, request)
    action = scope.get("action", "")
    show_all_specialties = (action == "all")

    if show_all_specialties:
        scope["specialty"] = ""

    requires_custom_weekdays = _program_requires_custom_weekdays(program, scope.get("specialty", ""))
    allow_third_weekday = _program_allows_third_weekday(program)

    model_cls, trainee_qs = _attendance_queryset(program, scope)
    specialty_options = _attendance_specialty_options(model_cls, scope["promotion"], program=program)
    promotion_options = _attendance_promotion_options(model_cls, program=program)
    weekday_validation = _attendance_weekday_validation(program, scope, requires_custom_weekdays=requires_custom_weekdays)
    weekdays = weekday_validation["weekdays"] if requires_custom_weekdays else _attendance_weekdays(program, scope)

    can_prepare_table = False
    if show_all_specialties:
        can_prepare_table = True
    elif scope["promotion"] and scope["specialty"]:
        can_prepare_table = True

    if requires_custom_weekdays:
        can_prepare_table = can_prepare_table and weekday_validation["is_valid"]

    show_table = bool(scope.get("show_table")) and can_prepare_table

    columns = []
    if show_table:
        columns = _attendance_month_dates(scope["year"], scope["month"], weekdays)

    sheet = None
    if show_table:
        defaults = {
            "يوم_الدراسة_1": weekdays[0] if len(weekdays) > 0 else None,
            "يوم_الدراسة_2": weekdays[1] if len(weekdays) > 1 else None,
        }
        for idx in range(3, 6):
            field_name = f"يوم_الدراسة_{idx}"
            if hasattr(كشفغياب, field_name):
                defaults[field_name] = weekdays[idx - 1] if len(weekdays) >= idx else None
        sheet, _ = كشفغياب.objects.get_or_create(
            البرنامج=program,
            الدفعة=scope["promotion"],
            التخصص=scope["specialty"],
            الشهر=scope["month"],
            السنة=scope["year"],
            defaults={**defaults, "created_by": request.user},
        )
        dirty = False
        update_fields = []
        if requires_custom_weekdays:
            for field_name in ("يوم_الدراسة_1", "يوم_الدراسة_2", "يوم_الدراسة_3", "يوم_الدراسة_4", "يوم_الدراسة_5"):
                if field_name not in defaults or not hasattr(sheet, field_name):
                    continue
                if getattr(sheet, field_name) != defaults[field_name]:
                    setattr(sheet, field_name, defaults[field_name])
                    update_fields.append(field_name)
                    dirty = True
        if dirty:
            update_fields.append("updated_at")
            sheet.save(update_fields=update_fields)

    trainees = list(trainee_qs) if columns else []
    slot_count = _attendance_visual_slot_count(program)
    entry_map = {}
    if sheet and trainees and columns:
        trainee_ids = [t.pk for t in trainees]
        column_dates = [col["date"] for col in columns]
        for entry in خليةغياب.objects.filter(
            الكشف=sheet,
            trainee_id__in=trainee_ids,
            التاريخ__in=column_dates
        ).only("trainee_id", "التاريخ", "الحالة", "ملاحظة", "رقم_الخانة"):
            if program == "apprentice":
                entry_map[(entry.trainee_id, entry.التاريخ.isoformat(), entry.رقم_الخانة)] = entry
            else:
                entry_map[(entry.trainee_id, entry.التاريخ.isoformat())] = entry

    rows = []
    for idx, trainee in enumerate(trainees, start=1):
        cells = []
        status_label_map = _attendance_status_label_map(program)
        for col in columns:
            if program == "apprentice":
                slots = []
                for slot_no in range(1, slot_count + 1):
                    entry = entry_map.get((trainee.pk, col["iso"], slot_no))
                    status_value = entry.الحالة if entry else ""
                    slots.append({
                        "slot": slot_no,
                        "status": status_value,
                        "display_label": status_label_map.get(status_value, ""),
                        "note": entry.ملاحظة if entry else "",
                    })
                cells.append({
                    "date": col["date"],
                    "iso": col["iso"],
                    "slots": slots,
                })
            else:
                entry = entry_map.get((trainee.pk, col["iso"]))
                status_value = entry.الحالة if entry else ""
                cells.append({
                    "date": col["date"],
                    "iso": col["iso"],
                    "status": status_value,
                    "display_label": status_label_map.get(status_value, ""),
                    "note": entry.ملاحظة if entry else "",
                })
        rows.append({
            "index": idx,
            "trainee": trainee,
            "cells": cells,
        })

    return {
        "scope": scope,
        "action": action,
        "show_all_specialties": show_all_specialties,
        "is_bridge_specialty": requires_custom_weekdays,
        "requires_custom_weekdays": requires_custom_weekdays,
        "allow_third_weekday": allow_third_weekday,
        "model_cls": model_cls,
        "specialty_options": specialty_options,
        "promotion_options": promotion_options,
        "weekdays": weekdays,
        "weekday_validation": weekday_validation,
        "show_table": show_table,
        "columns": columns,
        "sheet": sheet,
        "rows": rows,
        "slot_count": slot_count,
        "display_columns_count": len(columns) * slot_count,
    }


def _write_attendance_excel_sheet(ws, program, payload, rows, sheet_title=None, specialty_label=""):
    program_label = ATTENDANCE_PROGRAMS[program]["label"]
    sheet = payload.get("sheet")
    columns = payload["columns"]
    slot_count = payload.get("slot_count", 1)
    show_specialty_column = bool(payload["show_all_specialties"] and not specialty_label)
    label_map = _attendance_status_label_map(program)

    ws.title = _safe_sheet_title(sheet_title or (specialty_label or "الغيابات"))
    ws.append([f"جدول الغيابات - {program_label}"])
    subtitle = str(sheet) if sheet else ""
    if specialty_label:
        subtitle = f"{subtitle} - {specialty_label}" if subtitle else specialty_label
    ws.append([subtitle])

    base_header = ["الرقم", "الاسم واللقب"]
    if show_specialty_column:
        base_header.append("التخصص")

    ws.append(base_header)
    ws.append([""] * len(base_header))

    base_col_count = len(base_header)
    current_col = base_col_count + 1
    for col in columns:
        if slot_count > 1:
            ws.merge_cells(start_row=3, start_column=current_col, end_row=3, end_column=current_col + slot_count - 1)
            ws.merge_cells(start_row=4, start_column=current_col, end_row=4, end_column=current_col + slot_count - 1)
        ws.cell(row=3, column=current_col).value = col["weekday_label"]
        ws.cell(row=4, column=current_col).value = str(col["day_num"]).zfill(2)
        current_col += slot_count

    for idx, title in enumerate(base_header, start=1):
        ws.merge_cells(start_row=3, start_column=idx, end_row=4, end_column=idx)
        ws.cell(row=3, column=idx).value = title

    for row_idx, row in enumerate(rows, start=5):
        values = [str(row["index"]).zfill(2), f"{row['trainee'].اللقب} {row['trainee'].الاسم}"]
        if show_specialty_column:
            values.append(getattr(row["trainee"], "التخصص", "") or "")
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

        current_col = base_col_count + 1
        for cell in row["cells"]:
            if program == "apprentice":
                for slot in cell.get("slots", []):
                    ws.cell(row=row_idx, column=current_col).value = label_map.get(slot.get("status", ""), "")
                    current_col += 1
            else:
                ws.cell(row=row_idx, column=current_col).value = label_map.get(cell.get("status", ""), "")
                current_col += 1

    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9E2F3")
    border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))

    for row_num in (1, 2):
        for cell in ws[row_num]:
            cell.font = Font(bold=True, size=14 if row_num == 1 else 11, color="FFFFFF" if row_num == 1 else "000000")
            if row_num == 1:
                cell.fill = title_fill
            else:
                cell.fill = subtitle_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    if ws.max_column >= 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)

    for row_num in (3, 4):
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color="FFFFFF" if row_num == 3 else "000000")
            cell.fill = title_fill if row_num == 3 else subtitle_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    from openpyxl.utils import get_column_letter
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    if show_specialty_column:
        ws.column_dimensions['C'].width = 24
        start_idx = 4
    else:
        start_idx = 3
    for i in range(start_idx, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 6.5 if slot_count > 1 else 11


def _attendance_export_excel(program, payload):
    wb = _build_attendance_export_workbook(program, payload)
    scope = payload["scope"]
    specialty = (scope.get("specialty") or "").strip()
    filename = _attendance_export_filename("table", program, scope, specialty="" if payload["show_all_specialties"] else specialty, ext="xlsx")
    return _finalize_workbook_response(wb, filename)


def _attendance_export_pdf_bytes(program, payload):
    font_name = _register_pdf_font()
    styles = getSampleStyleSheet()
    for style_name in ("Normal", "Title", "Heading2"):
        styles[style_name].fontName = font_name
        styles[style_name].alignment = 1

    rows = payload["rows"]
    columns = payload["columns"]
    scope = payload["scope"]
    slot_count = payload.get("slot_count", 1)
    label_map = _attendance_status_label_map(program)
    show_all_specialties = payload["show_all_specialties"]
    specialty_value = (scope.get("specialty") or "").strip()

    grouped_rows = [(specialty_value or "كل التخصصات", rows)]
    if show_all_specialties:
        grouped_rows = _group_items_by_specialty(rows, lambda row: getattr(row["trainee"], "التخصص", "") or "بدون تخصص")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        rightMargin=18,
        leftMargin=18,
        topMargin=16,
        bottomMargin=16,
    )

    story = []
    dark = colors.HexColor("#1F4E78")
    light = colors.HexColor("#D9E2F3")
    white = colors.white

    for index, (specialty_name, specialty_rows) in enumerate(grouped_rows, start=1):
        semester_label = _attendance_rows_semester_label(specialty_rows)

        header_table = Table([[ _pdf_text(line) ] for line in ATTENDANCE_EXPORT_HEADER_LINES], colWidths=[doc.width], hAlign="CENTER")
        header_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 1), 12),
            ("FONTSIZE", (0, 2), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 6))

        title_table = Table([[ _pdf_text(_attendance_template_title(program, specialty_name)) ]], colWidths=[doc.width], hAlign="CENTER")
        title_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), dark),
            ("TEXTCOLOR", (0, 0), (-1, -1), white),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 15),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(title_table)

        subtitle_table = Table([[ _pdf_text(_attendance_scope_subtitle(scope, specialty_label=specialty_name, semester_label=semester_label)) ]], colWidths=[doc.width], hAlign="CENTER")
        subtitle_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), light),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(subtitle_table)
        story.append(Spacer(1, 10))

        header_row_1 = ["الرقم", "الاسم و اللقب"]
        header_row_2 = ["", ""]
        col_widths = [34, 140]
        show_specialty_column = bool(show_all_specialties)
        if show_specialty_column:
            header_row_1.append("التخصص")
            header_row_2.append("")
            col_widths.append(115)

        for col in columns:
            if slot_count > 1:
                for _slot_index in range(slot_count):
                    header_row_1.append(_pdf_text(col["weekday_label"]))
                    header_row_2.append(str(col["day_num"]).zfill(2))
                    col_widths.append(20)
            else:
                header_row_1.append(_pdf_text(col["weekday_label"]))
                header_row_2.append(str(col["day_num"]).zfill(2))
                col_widths.append(30)

        table_data = [_pdf_row(header_row_1), _pdf_row(header_row_2)]
        for row in specialty_rows:
            values = [str(row["index"]).zfill(2), f"{row['trainee'].اللقب} {row['trainee'].الاسم}"]
            if show_specialty_column:
                values.append(getattr(row["trainee"], "التخصص", "") or "")
            for day_cell in row["cells"]:
                if program == "apprentice":
                    slots = day_cell.get("slots", [])
                    for slot in slots[:slot_count]:
                        values.append(label_map.get(slot.get("status", ""), ""))
                    if len(slots) < slot_count:
                        values.extend([""] * (slot_count - len(slots)))
                else:
                    values.append(label_map.get(day_cell.get("status", ""), ""))
            table_data.append(_pdf_row(values))

        table = Table(table_data, colWidths=col_widths, repeatRows=2, hAlign="CENTER")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), dark),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("BACKGROUND", (0, 1), (-1, 1), light),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 1), 8),
            ("FONTSIZE", (0, 2), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.whitesmoke, colors.HexColor("#F4F7FB")]),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
        if index < len(grouped_rows):
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _attendance_export_pdf(program, payload):
    pdf_bytes = _attendance_export_pdf_bytes(program, payload)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    specialty = "" if payload["show_all_specialties"] else (payload["scope"].get("specialty") or "")
    return _set_download_filename(response, _attendance_export_filename("table", program, payload["scope"], specialty=specialty, ext="pdf"))


def _write_attendance_stats_excel_sheet(ws, program, payload, stats_rows, sheet_title=None, specialty_label=""):
    show_specialty_column = bool(payload["show_all_specialties"] and not specialty_label)
    scope = payload["scope"]
    semester_label = _attendance_rows_semester_label(stats_rows)
    batch = scope.get("batch")
    batch_display = f"الدفعة: {batch.السنة} / {batch.رقم_الدورة}" if batch else ""

    ws.title = _safe_sheet_title(sheet_title or specialty_label or "الإحصائيات")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A16"

    headers = ["الترتيب", "الاسم واللقب"]
    if show_specialty_column:
        headers.append("التخصص")
    headers.extend(["عدد الحضور", "عدد الغيابات", "غياب بعذر", "عدد التأخرات", "المجموع المسجل", "نسبة الغياب"])

    average_absence_rate = round(sum(item["absence_rate"] for item in stats_rows) / len(stats_rows), 2) if stats_rows else 0
    stats_totals = {
        "present_count": sum(item["present_count"] for item in stats_rows),
        "absent_count": sum(item["absent_count"] for item in stats_rows),
        "excused_count": sum(item["excused_count"] for item in stats_rows),
        "late_count": sum(item["late_count"] for item in stats_rows),
        "total_recorded": sum(item["total_recorded"] for item in stats_rows),
    }

    summary_rows = [
        ["عدد المتكونين", len(stats_rows)],
        ["عدد الأيام المعروضة", payload["displayed_days_count"]],
        ["متوسط نسبة الغياب", f"{average_absence_rate}%"],
        ["طريقة الحساب", "بالتفصيل حسب الخانات" if payload.get("slot_count", 1) > 1 else "حسب الأيام المسجلة"],
    ]

    total_columns = len(headers)
    _apply_official_stats_excel_header(ws, total_columns, program, scope, specialty_label=specialty_label, semester_label=semester_label, batch_display=batch_display)

    summary_start_row = 11
    for offset, item in enumerate(summary_rows):
        row_num = summary_start_row + offset
        ws.cell(row=row_num, column=1, value=item[0])
        ws.cell(row=row_num, column=2, value=item[1])

    header_row = summary_start_row + len(summary_rows) + 1
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)

    data_start_row = header_row + 1
    for n, row in enumerate(stats_rows, start=1):
        values = [n, f"{row['trainee'].اللقب} {row['trainee'].الاسم}"]
        if show_specialty_column:
            values.append(getattr(row["trainee"], "التخصص", "") or "")
        values.extend([row["present_count"], row["absent_count"], row["excused_count"], row["late_count"], row["total_recorded"], row["absence_rate"]])
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=data_start_row + n - 1, column=col_idx, value=value)

    final_row = data_start_row + len(stats_rows)
    if stats_rows:
        total_row = ["", "الإجمالي"]
        if show_specialty_column:
            total_row.append("")
        total_row.extend([stats_totals["present_count"], stats_totals["absent_count"], stats_totals["excused_count"], stats_totals["late_count"], stats_totals["total_recorded"], average_absence_rate])
        for col_idx, value in enumerate(total_row, start=1):
            ws.cell(row=final_row, column=col_idx, value=value)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9E2F3")
    thin_border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))

    for row_num in range(summary_start_row, summary_start_row + len(summary_rows)):
        ws.row_dimensions[row_num].height = 23
        for col_idx in range(1, 3):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=1).fill = subtitle_fill
        ws.cell(row=row_num, column=2).font = Font(bold=True)

    for col_idx in range(1, total_columns + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    end_data_row = final_row if stats_rows else data_start_row - 1
    for row_cells in ws.iter_rows(min_row=data_start_row, max_row=end_data_row, min_col=1, max_col=total_columns):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    if stats_rows:
        for cell in ws[final_row]:
            cell.font = Font(bold=True)
            cell.fill = subtitle_fill

    from openpyxl.utils import get_column_letter
    widths = {1: 18, 2: 26}
    if show_specialty_column:
        widths[3] = 24
    data_widths = {
        1: 10,
        2: 28,
        3: 24 if show_specialty_column else 16,
    }
    for col_idx in range(1, total_columns + 1):
        default_width = 16
        if col_idx == 1:
            default_width = 18
        elif col_idx == 2:
            default_width = 26
        elif show_specialty_column and col_idx == 3:
            default_width = 24
        ws.column_dimensions[get_column_letter(col_idx)].width = default_width

def _attendance_stats_export_excel(program, payload):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    stats_rows = payload["stats_rows"]
    scope = payload["scope"]
    specialty = (scope.get("specialty") or "").strip()

    if payload["show_all_specialties"]:
        groups = _group_items_by_specialty(stats_rows, lambda row: getattr(row["trainee"], "التخصص", "") or "بدون تخصص")
        for idx, (specialty_name, specialty_rows) in enumerate(groups):
            ws = wb.create_sheet(title=_safe_sheet_title(specialty_name or f"تخصص {idx + 1}"))
            _write_attendance_stats_excel_sheet(ws, program, payload, specialty_rows, sheet_title=specialty_name, specialty_label=specialty_name)
    else:
        ws = wb.create_sheet(title=_safe_sheet_title(specialty or "الإحصائيات"))
        _write_attendance_stats_excel_sheet(ws, program, payload, stats_rows, sheet_title=specialty or "الإحصائيات", specialty_label=specialty)

    filename = _attendance_export_filename("stats", program, scope, specialty="" if payload["show_all_specialties"] else specialty, ext="xlsx")
    return _finalize_workbook_response(wb, filename)


def _attendance_stats_export_pdf(program, payload):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=16, bottomMargin=16)
    font_name = _register_pdf_font()
    styles = getSampleStyleSheet()
    for style_name in ("Normal", "Title", "Heading2"):
        styles[style_name].fontName = font_name
        styles[style_name].alignment = 1

    scope = payload["scope"]
    batch = scope.get("batch")
    batch_display = f"الدفعة: {batch.السنة} / {batch.رقم_الدورة}" if batch else ""
    show_all_specialties = payload["show_all_specialties"]

    grouped_rows = [(scope.get("specialty") or "كل التخصصات", payload["stats_rows"])]
    if show_all_specialties:
        grouped_rows = _group_items_by_specialty(payload["stats_rows"], lambda row: getattr(row["trainee"], "التخصص", "") or "بدون تخصص")

    dark = colors.HexColor("#1F4E78")
    light = colors.HexColor("#D9E2F3")
    white = colors.white
    story = []

    for idx2, (specialty_name, stats_rows) in enumerate(grouped_rows, start=1):
        semester_label = _attendance_rows_semester_label(stats_rows)
        average_absence_rate = round(sum(item["absence_rate"] for item in stats_rows) / len(stats_rows), 2) if stats_rows else 0
        stats_totals = {
            "present_count": sum(item["present_count"] for item in stats_rows),
            "absent_count": sum(item["absent_count"] for item in stats_rows),
            "excused_count": sum(item["excused_count"] for item in stats_rows),
            "late_count": sum(item["late_count"] for item in stats_rows),
            "total_recorded": sum(item["total_recorded"] for item in stats_rows),
        }

        header_table = Table([[_pdf_text(line)] for line in ATTENDANCE_EXPORT_HEADER_LINES], colWidths=[doc.width], hAlign="CENTER")
        header_table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, 1), 12), ("FONTSIZE", (0, 2), (-1, -1), 10), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("BOTTOMPADDING", (0, 0), (-1, -1), 1), ("TOPPADDING", (0, 0), (-1, -1), 1)]))
        story.append(header_table)
        story.append(Spacer(1, 6))

        title_text = _attendance_stats_template_title(program, specialty_name if specialty_name != "كل التخصصات" else "")
        title_table = Table([[_pdf_text(title_text)]], colWidths=[doc.width], hAlign="CENTER")
        title_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), dark), ("TEXTCOLOR", (0, 0), (-1, -1), white), ("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, -1), 15), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
        story.append(title_table)

        subtitle_table = Table([[_pdf_text(_attendance_stats_scope_subtitle(scope, specialty_label=specialty_name if specialty_name != "كل التخصصات" else "", semester_label=semester_label, batch_display=batch_display))]], colWidths=[doc.width], hAlign="CENTER")
        subtitle_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), light), ("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, -1), 11), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        story.append(subtitle_table)
        story.append(Spacer(1, 8))

        summary_data = [_pdf_row(["عدد المتكونين", str(len(stats_rows))]), _pdf_row(["عدد الأيام المعروضة", str(payload["displayed_days_count"])]), _pdf_row(["متوسط نسبة الغياب", f"{average_absence_rate}%"]), _pdf_row(["طريقة الحساب", "بالتفصيل حسب الخانات" if payload.get("slot_count", 1) > 1 else "حسب الأيام المسجلة"])]
        summary_table = Table(summary_data, colWidths=[150, 260], hAlign="RIGHT")
        summary_table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), font_name), ("GRID", (0, 0), (-1, -1), 0.5, colors.black), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("BACKGROUND", (0, 0), (0, -1), light), ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        story.append(summary_table)
        story.append(Spacer(1, 10))

        headers = ["الترتيب", "الاسم واللقب"]
        if show_all_specialties:
            headers.append("التخصص")
        headers.extend(["عدد الحضور", "عدد الغيابات", "غياب بعذر", "عدد التأخرات", "المجموع المسجل", "نسبة الغياب"])
        data = [_pdf_row(headers)]
        for row in stats_rows:
            line = [row["display_index"], f"{row['trainee'].اللقب} {row['trainee'].الاسم}"]
            if show_all_specialties:
                line.append(getattr(row["trainee"], "التخصص", "") or "")
            line.extend([row["present_count"], row["absent_count"], row["excused_count"], row["late_count"], row["total_recorded"], f"{row['absence_rate']}%"])
            data.append(_pdf_row(line))

        if stats_rows:
            total_line = ["", "الإجمالي"]
            if show_all_specialties:
                total_line.append("")
            total_line.extend([stats_totals["present_count"], stats_totals["absent_count"], stats_totals["excused_count"], stats_totals["late_count"], stats_totals["total_recorded"], f"{average_absence_rate}%"])
            data.append(_pdf_row(total_line))

        col_widths = [42, 150]
        if show_all_specialties:
            col_widths.append(95)
        col_widths.extend([58, 58, 58, 58, 62, 62])
        table = Table(data, repeatRows=1, colWidths=col_widths, hAlign="CENTER")
        style_cmds = [("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, 0), 9), ("FONTSIZE", (0, 1), (-1, -1), 8), ("BACKGROUND", (0, 0), (-1, 0), dark), ("TEXTCOLOR", (0, 0), (-1, 0), white), ("GRID", (0, 0), (-1, -1), 0.45, colors.black), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#F4F7FB")]), ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]
        if stats_rows:
            style_cmds.append(("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), light))
        table.setStyle(TableStyle(style_cmds))
        story.append(table)

        if idx2 < len(grouped_rows):
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    specialty = "" if payload["show_all_specialties"] else (scope.get("specialty") or "")
    return _set_download_filename(response, _attendance_export_filename("stats", program, scope, specialty=specialty, ext="pdf"))

def _write_saved_attendance_stats_excel_sheet(ws, context, detail_rows, sheet_title=None, specialty_label=""):
    detail_scope = context["detail_scope"]
    show_specialty_column = not bool(detail_scope.get("specialty")) and not specialty_label
    semester_label = detail_scope.get("semester") or ""
    batch = detail_scope.get("batch")
    batch_display = f"الدفعة: {batch.السنة} / {batch.رقم_الدورة}" if batch else ""

    ws.title = _safe_sheet_title(sheet_title or specialty_label or "الإحصائيات المحفوظة")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A15"

    latest_saved_at = max((item.updated_at for item in detail_rows), default=None)
    average_absence_rate = round(sum(item.absence_rate for item in detail_rows) / len(detail_rows), 2) if detail_rows else 0
    summary_rows = [["عدد المتكونين", len(detail_rows)], ["متوسط نسبة الغياب", f"{average_absence_rate}%"], ["آخر حفظ", latest_saved_at.strftime("%Y-%m-%d %H:%M") if latest_saved_at else ""]]

    headers = ["الاسم واللقب"]
    if show_specialty_column:
        headers.append("التخصص")
    headers.extend(["عدد الحضور", "عدد الغيابات", "غياب بعذر", "عدد التأخرات", "المجموع المسجل", "نسبة الغياب", "تاريخ الحفظ"])

    total_columns = len(headers)
    official_specialty = specialty_label or (detail_scope.get("specialty_display") if detail_scope.get("specialty") else "") or ""
    _apply_official_stats_excel_header(ws, total_columns, detail_scope["program"], detail_scope, specialty_label=official_specialty, semester_label=semester_label, batch_display=batch_display)

    summary_start_row = 11
    for offset, item in enumerate(summary_rows):
        row_num = summary_start_row + offset
        ws.cell(row=row_num, column=1, value=item[0])
        ws.cell(row=row_num, column=2, value=item[1])

    header_row = summary_start_row + len(summary_rows) + 1
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)

    data_start_row = header_row + 1
    for r, row in enumerate(detail_rows, start=0):
        values = [row.trainee_name]
        if show_specialty_column:
            values.append(row.trainee_specialty or "")
        values.extend([row.present_count, row.absent_count, row.excused_count, row.late_count, row.total_recorded, row.absence_rate, row.updated_at.strftime("%Y-%m-%d %H:%M")])
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=data_start_row + r, column=col_idx, value=value)

    final_row = data_start_row + len(detail_rows)
    if detail_rows:
        total_row = ["الإجمالي"]
        if show_specialty_column:
            total_row.append("")
        total_row.extend([sum(item.present_count for item in detail_rows), sum(item.absent_count for item in detail_rows), sum(item.excused_count for item in detail_rows), sum(item.late_count for item in detail_rows), sum(item.total_recorded for item in detail_rows), average_absence_rate, ""])
        for col_idx, value in enumerate(total_row, start=1):
            ws.cell(row=final_row, column=col_idx, value=value)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9E2F3")
    thin_border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))

    for row_num in range(summary_start_row, summary_start_row + len(summary_rows)):
        ws.row_dimensions[row_num].height = 23
        for col_idx in range(1, 3):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=1).fill = subtitle_fill
        ws.cell(row=row_num, column=2).font = Font(bold=True)

    for col_idx in range(1, total_columns + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    end_data_row = final_row if detail_rows else data_start_row - 1
    for row_cells in ws.iter_rows(min_row=data_start_row, max_row=end_data_row, min_col=1, max_col=total_columns):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    if detail_rows:
        for cell in ws[final_row]:
            cell.font = Font(bold=True)
            cell.fill = subtitle_fill

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, total_columns + 1):
        if col_idx == 1:
            width = 28
        elif show_specialty_column and col_idx == 2:
            width = 24
        else:
            width = 16
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 18)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 26)



def _saved_attendance_stats_archive_context(request):
    return build_saved_attendance_stats_archive_context(
        request=request,
        attendance_programs=ATTENDANCE_PROGRAMS,
        has_program_permission_func=has_program_permission,
        snapshots_model=AttendanceStatSnapshot,
    )

def _saved_attendance_stats_export_excel(context):
    return build_saved_attendance_stats_excel_response(
        context,
        safe_sheet_title=_safe_sheet_title,
        apply_official_stats_excel_header=_apply_official_stats_excel_header,
        group_items_by_specialty=_group_items_by_specialty,
        attendance_export_filename=_attendance_export_filename,
        finalize_workbook_response=_finalize_workbook_response,
    )


def _saved_attendance_stats_export_pdf(context):
    return build_saved_attendance_stats_pdf_response(
        context,
        register_pdf_font=_register_pdf_font,
        pdf_text=_pdf_text,
        header_lines=ATTENDANCE_EXPORT_HEADER_LINES,
        attendance_stats_template_title=_attendance_stats_template_title,
        attendance_stats_scope_subtitle=_attendance_stats_scope_subtitle,
        pdf_row=_pdf_row,
        attendance_export_filename=_attendance_export_filename,
        set_download_filename=_set_download_filename,
        group_items_by_specialty=_group_items_by_specialty,
    )


@login_required
def saved_attendance_stats_archive(request):
    context = _saved_attendance_stats_archive_context(request)
    log_activity(request, "view", program=context["selected_program"] or "attendance_archive", details="عرض أرشيف إحصائيات الغياب المحفوظة")
    return render(request, "trainees/attendance_saved_stats.html", context)


@login_required
def saved_attendance_stats_export(request, fmt):
    context = _saved_attendance_stats_archive_context(request)
    detail_scope = context.get("detail_scope")
    detail_rows = context.get("detail_rows") or []
    if not detail_scope or not detail_rows:
        messages.error(request, "اختر النمط والشهر والسنة أولاً ثم افتح نطاقًا محفوظًا قبل التصدير.")
        return redirect(reverse("attendance_saved_stats") + ("?" + request.GET.urlencode() if request.GET else ""))

    log_activity(request, "export", program=detail_scope["program"], details=f"تصدير أرشيف الإحصائيات المحفوظة {fmt}")
    if fmt == "excel":
        return _saved_attendance_stats_export_excel(context)
    if fmt == "pdf":
        return _saved_attendance_stats_export_pdf(context)
    return HttpResponseBadRequest("صيغة غير صالحة")



def _attendance_actions_list_url(program: str, source: str) -> str:
    if source == "slots":
        return reverse("attendance_slot_actions", args=[program])
    return reverse("attendance_actions", args=[program])


def _attendance_actions_query_with_source(request, source: str) -> str:
    params = request.GET.copy()
    if source == "slots":
        params["source"] = "slots"
    else:
        params.pop("source", None)
    return params.urlencode()


def _attendance_actions_query_with_overrides(request, source: str, **overrides) -> str:
    params = request.GET.copy()
    if source == "slots":
        params["source"] = "slots"
    else:
        params.pop("source", None)
    for key, value in overrides.items():
        if value is None:
            params.pop(key, None)
        else:
            params[key] = value
    return params.urlencode()


def _attendance_actions_page(request, program, source="daily"):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()

    source = "slots" if source == "slots" else "daily"
    require_program_permission(request, program, "view")
    scope = _attendance_scope(program, request)
    action_type = (request.GET.get("action_type") or "").strip()
    status = (request.GET.get("status") or "").strip()
    raw_archive_state = request.GET.get("archive_state")
    archive_state = "active" if raw_archive_state is None else raw_archive_state.strip()

    try:
        if source == "slots":
            slot_payload = build_attendance_slot_payload(program, request)
            if slot_payload.get("sheet"):
                sync_slot_attendance_actions(program, slot_payload, request.user)
        else:
            if _program_requires_custom_weekdays(program, scope.get("specialty", "")):
                validation = _attendance_weekday_validation(program, scope, requires_custom_weekdays=True)
                if validation.get("is_valid"):
                    scope = dict(scope)
                    scope["visible_dates"] = [col["date"] for col in _attendance_month_dates(scope["year"], scope["month"], validation.get("weekdays") or [])]
            payload = build_monthly_action_payload(program, scope, _attendance_queryset)
            sync_attendance_actions(program, payload, request.user)
    except Exception:
        pass

    qs = attendance_actions_qs(program, scope, action_type=action_type, status=status, archive_state=archive_state, source=source)
    actions = list(qs)
    summary = summarize_attendance_actions(actions)
    current_query = _attendance_actions_query_with_source(request, source)
    log_label = "إعذارات الحصص" if source == "slots" else "الإعذارات اليومية"
    log_activity(request, "view", program=program, details=f"عرض {log_label} {ATTENDANCE_PROGRAMS[program]['label']}")

    if source == "slots" and program in ATTENDANCE_SLOT_PROGRAMS:
        table_back_url = reverse(ATTENDANCE_SLOT_PROGRAMS[program]["table_url_name"]) + ("?" + current_query if current_query else "")
        stats_back_url = reverse(ATTENDANCE_SLOT_PROGRAMS[program]["stats_url_name"]) + ("?" + current_query if current_query else "")
        page_title = f"إعذارات الحصص - {ATTENDANCE_PROGRAMS[program]['label']}"
        page_heading = f"{ATTENDANCE_PROGRAMS[program]['label']} - إعذارات الغياب بالحصة"
        page_note = "هذه صفحة مستقلة عن الإعذارات اليومية؛ تعرض فقط الإعذارات الناتجة عن جدول الغياب بالحصة الجديد."
    else:
        table_back_url = reverse("attendance_program", args=[program]) + ("?" + current_query if current_query else "")
        stats_back_url = reverse("attendance_stats", args=[program]) + ("?" + current_query if current_query else "")
        page_title = f"إدارة الأعذار - {ATTENDANCE_PROGRAMS[program]['label']}"
        page_heading = f"{ATTENDANCE_PROGRAMS[program]['label']} - إدارة الأعذار والاستدعاءات"
        page_note = "هذه صفحة الإعذارات اليومية القديمة، وتتم مزامنتها مع جدول الغيابات القديم عند فتح الصفحة أو تحديث العرض."

    return render(request, "trainees/attendance_actions.html", {
        "title": page_title,
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "scope": scope,
        "actions": actions,
        "summary": summary,
        "promotion_options": _attendance_promotion_options(MODEL_BY_PROGRAM[program], program=program),
        "specialty_options": _attendance_specialty_options(MODEL_BY_PROGRAM[program], scope.get("promotion_obj"), program=program),
        "month_choices": [{"value": value, "label": label} for value, label in ATTENDANCE_MONTH_CHOICES],
        "selected_action_type": action_type,
        "selected_status": status,
        "selected_archive_state": archive_state,
        "action_type_choices": AttendanceAction.ACTION_TYPE_CHOICES,
        "status_choices": AttendanceAction.STATUS_CHOICES,
        "archive_state_choices": AttendanceAction.ARCHIVE_STATE_CHOICES,
        "current_query": current_query,
        "actions_list_url": _attendance_actions_list_url(program, source),
        "active_archive_query": _attendance_actions_query_with_overrides(request, source, archive_state="active"),
        "archived_archive_query": _attendance_actions_query_with_overrides(request, source, archive_state="archived"),
        "all_archive_query": _attendance_actions_query_with_overrides(request, source, archive_state=""),
        "is_archive_view": archive_state == "archived",
        "actions_source": source,
        "actions_page_heading": page_heading,
        "actions_page_note": page_note,
        "table_back_url": table_back_url,
        "stats_back_url": stats_back_url,
    })


@login_required
def attendance_actions(request, program):
    return _attendance_actions_page(request, program, source="daily")


@login_required
def attendance_slot_actions(request, program):
    return _attendance_actions_page(request, program, source="slots")


@login_required
def attendance_actions_preview(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    scope = _attendance_scope(program, request)
    source = "slots" if (request.GET.get("source") or "").strip() == "slots" else "daily"
    action_type = (request.GET.get("action_type") or "").strip()
    status = (request.GET.get("status") or "").strip()
    raw_archive_state = request.GET.get("archive_state")
    archive_state = "active" if raw_archive_state is None else raw_archive_state.strip()
    selected_ids = selected_action_ids_from_request(request)
    qs = attendance_actions_qs(program, scope, action_type=action_type, status=status, archive_state=archive_state, source=source)
    if selected_ids:
        qs = qs.filter(pk__in=selected_ids)
    actions = list(qs)
    if not actions:
        messages.error(request, "لا توجد وثائق للمعاينة حسب الفلاتر الحالية.")
        base_url = _attendance_actions_list_url(program, source)
        return redirect(base_url + ("?" + request.GET.urlencode() if request.GET else ""))
    log_activity(request, "view", program=program, details=f"معاينة مجموعة وثائق الغياب ({len(actions)})")
    return render(request, "trainees/attendance_actions_preview.html", {
        "title": f"معاينة وثائق {'إعذارات الحصص' if source == 'slots' else 'الأعذار'} - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "actions": actions,
        "current_query": request.GET.urlencode(),
        "actions_source": source,
    })


@login_required
def attendance_action_edit(request, pk):
    action = get_object_or_404(AttendanceAction, pk=pk)
    require_program_permission(request, action.program, "change")

    if request.method == "POST":
        form = AttendanceActionForm(request.POST, instance=action)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            if obj.absence_start_date and obj.send_date and obj.status == "pending":
                obj.status = "ready"
            obj.save()
            log_activity(request, "change", program=action.program, details=f"تعديل إجراء غياب: {action.trainee_name}")
            messages.success(request, "تم تحديث بيانات الإعذار بنجاح.")
            return redirect(_attendance_actions_list_url(action.program, attendance_action_source(action)) + "?" + attendance_action_base_query(action))
    else:
        form = AttendanceActionForm(instance=action)
        if not action.document_number:
            form.fields["document_number"].initial = next_attendance_document_number(action.year)

    return render(request, "trainees/attendance_action_form.html", {
        "title": f"تعديل {action.get_action_type_display()}",
        "action_obj": action,
        "form": form,
        "base_query": attendance_action_base_query(action),
        "back_url": _attendance_actions_list_url(action.program, attendance_action_source(action)) + "?" + attendance_action_base_query(action),
    })



def _attendance_action_source_from_query(query: str) -> str:
    return "slots" if "source=slots" in (query or "") else "daily"


def _attendance_bulk_edit_redirect_url(program, return_query=""):
    return_query = (return_query or "").strip()
    source = _attendance_action_source_from_query(return_query)
    base_url = _attendance_actions_list_url(program, source)
    if return_query:
        return f"{base_url}?{return_query}"
    return base_url


@login_required
def attendance_actions_bulk_edit(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        return HttpResponseBadRequest("نمط غير صالح")
    require_program_permission(request, program, "change")

    if request.method == "POST":
        ids = selected_action_ids_from_request(request)
    else:
        ids = [int(v) for v in request.GET.getlist("ids") if str(v).isdigit()]
    ids = list(dict.fromkeys(ids))

    return_query = (request.POST.get("return_query") if request.method == "POST" else request.GET.get("return_query") or request.META.get("QUERY_STRING") or "").strip()
    if not ids:
        messages.error(request, "اختر متكونًا واحدًا على الأقل قبل فتح التعديل الجماعي.")
        return redirect(_attendance_bulk_edit_redirect_url(program, return_query))

    actions_qs = filter_records_by_split_program_for_active_trainees(
        AttendanceAction.objects.filter(pk__in=ids),
        program,
        MODEL_BY_PROGRAM.get(program),
    )
    actions = list(actions_qs.order_by("trainee_name", "pk"))
    if not actions:
        messages.error(request, "لم يتم العثور على السجلات المحددة.")
        return redirect(_attendance_bulk_edit_redirect_url(program, return_query))

    action_id_set = {obj.pk for obj in actions}
    if request.method == "POST":
        with transaction.atomic():
            updated_count = 0
            for obj in actions:
                prefix = f"row_{obj.pk}_"
                document_number = (request.POST.get(prefix + "document_number") or "").strip()
                absence_start_date_raw = request.POST.get(prefix + "absence_start_date")
                send_date_raw = request.POST.get(prefix + "send_date")
                status = (request.POST.get(prefix + "status") or "pending").strip()
                notes = (request.POST.get(prefix + "notes") or "").strip()

                absence_start_date = parse_bulk_action_date(absence_start_date_raw, DATE_INPUT_FORMATS)
                send_date = parse_bulk_action_date(send_date_raw, DATE_INPUT_FORMATS)

                if absence_start_date_raw and absence_start_date is None:
                    messages.error(request, f"تاريخ بداية الغياب غير صالح للسجل: {obj.trainee_name}")
                    return redirect(request.path + "?" + urlencode([("ids", pk) for pk in ids] + ([('return_query', return_query)] if return_query else []), doseq=True))
                if send_date_raw and send_date is None:
                    messages.error(request, f"تاريخ التحرير/الإرسال غير صالح للسجل: {obj.trainee_name}")
                    return redirect(request.path + "?" + urlencode([("ids", pk) for pk in ids] + ([('return_query', return_query)] if return_query else []), doseq=True))

                obj.document_number = document_number
                obj.absence_start_date = absence_start_date
                obj.send_date = send_date
                if status in dict(AttendanceAction.STATUS_CHOICES):
                    obj.status = status
                obj.notes = notes
                obj.updated_by = request.user
                if obj.absence_start_date and obj.send_date and obj.status == "pending":
                    obj.status = "ready"
                obj.full_clean()
                obj.save()
                updated_count += 1

        log_activity(request, "change", program=program, details=f"تعديل جماعي لإدارة الأعذار ({updated_count} سجل/سجلات)")
        messages.success(request, f"تم حفظ التعديلات الجماعية لعدد {updated_count} سجل/سجلات.")
        return redirect(_attendance_bulk_edit_redirect_url(program, return_query))

    query_pairs = [("ids", obj.pk) for obj in actions]
    if return_query:
        query_pairs.append(("return_query", return_query))
    query_string = urlencode(query_pairs, doseq=True)
    return render(request, "trainees/attendance_actions_bulk_form.html", {
        "title": f"تعديل جماعي - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "actions": actions,
        "return_query": return_query,
        "status_choices": AttendanceAction.STATUS_CHOICES,
        "back_url": _attendance_bulk_edit_redirect_url(program, return_query),
        "query_string": query_string,
    })


@login_required
def attendance_action_archive_toggle(request, pk):
    action = get_object_or_404(AttendanceAction, pk=pk)
    require_program_permission(request, action.program, "change")
    target = (request.POST.get("target") or "archive").strip()
    note_changed = False
    if target == "restore":
        action.is_archived = False
        action.archived_at = None
        clear_attendance_action_deletion(action)
        note_changed = _clear_attendance_action_archive_markers(action)
        msg = "تم استرجاع السجل من الأرشيف."
        log_label = "استرجاع"
    else:
        action.is_archived = True
        action.archived_at = timezone.now()
        # هذه أرشفة يدوية: تبقى مخفية من الجدول النشط حتى يضغط المستخدم زر استرجاع،
        # حتى لو بقيت الغيابات مستحقة عند المزامنة.
        note_changed = _mark_attendance_action_manual_archive(action)
        msg = "تم حفظ السجل في الأرشيف."
        log_label = "أرشفة"
    action.updated_by = request.user
    update_fields = ["is_archived", "archived_at", "updated_by", "updated_at"]
    if note_changed:
        update_fields.append("notes")
    action.save(update_fields=update_fields)
    log_activity(request, "change", program=action.program, details=f"{log_label} إجراء غياب: {action.trainee_name}")
    messages.success(request, msg)
    return redirect(_attendance_actions_list_url(action.program, attendance_action_source(action)) + "?" + attendance_action_base_query(action))


@login_required
def attendance_action_delete(request, pk):
    action = get_object_or_404(AttendanceAction, pk=pk)
    require_program_permission(request, action.program, "delete")
    if not action.is_archived:
        messages.error(request, "لا يمكن حذف سجل غير مؤرشف من هذا الزر. استخدم الحذف المباشر.")
        return redirect(_attendance_actions_list_url(action.program, attendance_action_source(action)) + "?" + attendance_action_base_query(action))
    program = action.program
    trainee_name = action.trainee_name
    query = attendance_action_base_query(action)
    register_attendance_action_deletion(action, request.user)
    action.delete()
    log_activity(request, "delete", program=program, details=f"حذف إجراء غياب مؤرشف: {trainee_name}")
    messages.success(request, "تم حذف السجل المؤرشف نهائيًا.")
    return redirect(_attendance_actions_list_url(program, _attendance_action_source_from_query(query)) + "?" + query)


@login_required
def attendance_action_delete_direct(request, pk):
    action = get_object_or_404(AttendanceAction, pk=pk)
    require_program_permission(request, action.program, "delete")
    if action.is_archived:
        messages.error(request, "هذا السجل مؤرشف. استخدم زر حذف الأرشيف لحذفه نهائيًا.")
        return redirect(_attendance_actions_list_url(action.program, attendance_action_source(action)) + "?" + attendance_action_base_query(action))
    program = action.program
    trainee_name = action.trainee_name
    query = attendance_action_base_query(action)
    register_attendance_action_deletion(action, request.user)
    action.delete()
    log_activity(request, "delete", program=program, details=f"حذف مباشر لإجراء غياب: {trainee_name}")
    messages.success(request, "تم حذف السجل نهائيًا قبل الأرشفة.")
    return redirect(_attendance_actions_list_url(program, _attendance_action_source_from_query(query)) + "?" + query)


@login_required
def attendance_actions_bulk(request):
    if request.method != "POST":
        return HttpResponseBadRequest("طريقة غير مدعومة")

    ids = selected_action_ids_from_request(request)
    program = (request.POST.get("program") or "").strip()
    if not ids or not program or program not in ATTENDANCE_PROGRAMS:
        messages.error(request, "اختر متكونًا واحدًا على الأقل قبل تنفيذ العملية.")
        return redirect(request.META.get("HTTP_REFERER") or reverse("attendance_home"))

    target = (request.POST.get("bulk_action") or "").strip()
    required_perm = "delete" if target in {"delete", "delete_direct"} else "change"
    require_program_permission(request, program, required_perm)
    qs = filter_records_by_split_program_for_active_trainees(
        AttendanceAction.objects.filter(pk__in=ids),
        program,
        MODEL_BY_PROGRAM.get(program),
    )
    query = (request.POST.get("return_query") or "").strip()
    base_redirect = _attendance_actions_list_url(program, _attendance_action_source_from_query(query))
    if query:
        base_redirect += "?" + query

    if not qs.exists():
        messages.error(request, "لم يتم العثور على السجلات المحددة.")
        return redirect(base_redirect)

    if target == "archive":
        archivable = list(qs.filter(is_archived=False))
        updated = 0
        for obj in archivable:
            obj.is_archived = True
            obj.archived_at = timezone.now()
            note_changed = _mark_attendance_action_manual_archive(obj)
            obj.updated_by = request.user
            update_fields = ["is_archived", "archived_at", "updated_by", "updated_at"]
            if note_changed:
                update_fields.append("notes")
            obj.save(update_fields=update_fields)
            updated += 1
        messages.success(request, f"تم حفظ {updated} سجل/سجلات في الأرشيف.")
        return redirect(base_redirect)
    if target == "restore":
        restorable = list(qs.filter(is_archived=True))
        updated = 0
        for obj in restorable:
            clear_attendance_action_deletion(obj)
            obj.is_archived = False
            obj.archived_at = None
            note_changed = _clear_attendance_action_archive_markers(obj)
            obj.updated_by = request.user
            update_fields = ["is_archived", "archived_at", "updated_by", "updated_at"]
            if note_changed:
                update_fields.append("notes")
            obj.save(update_fields=update_fields)
            updated += 1
        messages.success(request, f"تمت استعادة {updated} سجل/سجلات من الأرشيف.")
        return redirect(base_redirect)
    if target == "delete_direct":
        deletable = list(qs.filter(is_archived=False))
        skipped = qs.exclude(is_archived=False).count()
        deleted = len(deletable)
        if deleted:
            for obj in deletable:
                register_attendance_action_deletion(obj, request.user)
            AttendanceAction.objects.filter(pk__in=[obj.pk for obj in deletable]).delete()
        if deleted:
            messages.success(request, f"تم حذف {deleted} سجل/سجلات مباشرةً نهائيًا.")
        if skipped:
            messages.warning(request, f"تم تجاهل {skipped} سجل/سجلات مؤرشفة. استخدم حذف الأرشيف لها.")
        return redirect(base_redirect)
    if target == "delete":
        deletable = list(qs.filter(is_archived=True))
        skipped = qs.exclude(is_archived=True).count()
        deleted = len(deletable)
        if deleted:
            for obj in deletable:
                register_attendance_action_deletion(obj, request.user)
            AttendanceAction.objects.filter(pk__in=[obj.pk for obj in deletable]).delete()
        if deleted:
            messages.success(request, f"تم حذف {deleted} سجل/سجلات مؤرشفة نهائيًا.")
        if skipped:
            messages.warning(request, f"تم تجاهل {skipped} سجل/سجلات غير مؤرشفة.")
        return redirect(base_redirect)

    messages.error(request, "العملية المطلوبة غير مدعومة.")
    return redirect(base_redirect)


@login_required
def attendance_action_print(request, pk):
    action = get_object_or_404(AttendanceAction, pk=pk)
    require_program_permission(request, action.program, "view")
    download = (request.GET.get("download") or "").strip().lower()
    if download == "pdf":
        return build_attendance_action_pdf_response(action)
    if download == "word":
        return build_attendance_action_word_response(action)
    return render(
        request,
        "trainees/attendance_action_print.html",
        attendance_action_document_context(action, preview_query=attendance_action_base_query(action)),
    )

@login_required
def attendance_stats(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()

    require_program_permission(request, program, "view")
    payload = _attendance_stats_payload(program, request)
    scope = payload["scope"]
    hide_table = request.GET.get("hide_table") == "1"

    if request.method == "POST":
        post_action = (request.POST.get("post_action") or "").strip()

        if post_action == "save_stats":
            if not payload["columns"]:
                messages.error(request, "اعرض الجدول أولاً أو أكمل الفلاتر المطلوبة ثم احسب نسبة الغيابات.")
                return redirect(reverse("attendance_program", args=[program]) + ("?" + request.GET.urlencode() if request.GET else ""))

            require_program_permission(request, program, "change")
            saved_count = _save_attendance_stats_snapshot(program, payload, request.user)
            log_activity(request, "change", program=program, details=f"حفظ إحصائيات الغيابات {ATTENDANCE_PROGRAMS[program]['label']}")
            messages.success(request, f"تم حفظ {saved_count} سجل من الإحصائيات الحالية.")
            return redirect(reverse("attendance_stats", args=[program]) + ("?" + request.GET.urlencode() if request.GET else ""))

        if post_action == "delete_saved_stats":
            require_program_permission(request, program, "delete")
            deleted_count = _delete_attendance_stats_snapshot(program, scope)
            log_activity(
                request,
                "delete",
                program=program,
                details=f"حذف الإحصائيات المحفوظة الحالية {ATTENDANCE_PROGRAMS[program]['label']}"
            )
            messages.success(request, f"تم حذف {deleted_count} سجل من الإحصائيات المحفوظة الحالية.")

            params = request.GET.copy()
            params.pop("show_table", None)
            params["hide_table"] = "1"

            redirect_url = reverse("attendance_stats", args=[program])
            query = params.urlencode()
            if query:
                redirect_url += f"?{query}"
            return redirect(redirect_url)

        if post_action == "delete_old_stats":
            require_program_permission(request, program, "delete")
            cutoff_month, cutoff_year = parse_old_stats_cutoff(request.POST)

            if not valid_old_stats_cutoff(cutoff_month, cutoff_year):
                messages.error(request, "حدد شهرًا وسنة صالحين لحذف الإحصائيات القديمة.")
            else:
                deleted_count = _delete_old_attendance_stats(program, cutoff_year, cutoff_month)
                log_activity(
                    request,
                    "delete",
                    program=program,
                    details=f"حذف الإحصائيات الأقدم من {cutoff_month:02d}/{cutoff_year} - {ATTENDANCE_PROGRAMS[program]['label']}"
                )
                messages.success(request, f"تم حذف {deleted_count} سجل من الإحصائيات الأقدم من {cutoff_month:02d}/{cutoff_year}.")
            return redirect(reverse("attendance_stats", args=[program]) + ("?" + request.GET.urlencode() if request.GET else ""))

    if not payload["columns"] and not hide_table:
        messages.error(request, "اعرض الجدول أولاً أو أكمل الفلاتر المطلوبة ثم احسب نسبة الغيابات.")
        return redirect(reverse("attendance_program", args=[program]) + ("?" + request.GET.urlencode() if request.GET else ""))

    saved_stats_summary = _attendance_saved_stats_summary(program, scope)
    show_stats_table = bool(scope.get("show_table")) and not hide_table
    show_stats_table_url = reverse("attendance_stats", args=[program])
    show_stats_table_query = build_preserved_query(request.GET, remove_hide_table=True, force_show_table=True)
    show_stats_table_url += "?" + show_stats_table_query

    log_activity(request, "view", program=program, details=f"عرض إحصائيات الغيابات {ATTENDANCE_PROGRAMS[program]['label']}")
    return render(request, "trainees/attendance_stats.html", {
        "title": f"إحصائيات الغيابات - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "program_description": ATTENDANCE_PROGRAMS[program]["description"],
        "scope": scope,
        "sheet": payload["sheet"],
        "show_all_specialties": payload["show_all_specialties"],
        "stats_rows": payload["stats_rows"],
        "stats_totals": payload["stats_totals"],
        "trainee_count": payload["trainee_count"],
        "average_absence_rate": payload["average_absence_rate"],
        "displayed_days_count": payload["displayed_days_count"],
        "slot_count": payload.get("slot_count", 1),
        "saved_stats_summary": saved_stats_summary,
        "current_query": request.GET.urlencode(),
        "show_stats_table": show_stats_table,
        "show_stats_table_url": show_stats_table_url,
        "can_change": has_program_permission(request.user, program, "change"),
        "can_delete": has_program_permission(request.user, program, "delete"),
        "month_choices": [{"value": value, "label": label} for value, label in ATTENDANCE_MONTH_CHOICES],
    })

@login_required
def attendance_program(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()

    require_program_permission(request, program, "view")
    payload = _attendance_table_payload(program, request)
    scope = payload["scope"]
    action = payload["action"]
    show_all_specialties = payload["show_all_specialties"]
    is_bridge_specialty = payload["is_bridge_specialty"]
    requires_custom_weekdays = payload.get("requires_custom_weekdays", is_bridge_specialty)
    allow_third_weekday = payload.get("allow_third_weekday", False)
    columns = payload["columns"]
    rows = payload["rows"]
    sheet = payload["sheet"]

    if request.method == "POST":
        post_action = resolve_attendance_post_action(request.POST)
        should_process_save = should_process_attendance_save(request.POST, action=post_action)
        should_process_delete = should_process_attendance_delete(request.POST, action=post_action)

        if should_process_save or should_process_delete:
            require_program_permission(request, program, "change")
            if not sheet:
                messages.error(request, "حدد الخيارات المطلوبة ثم اعرض الجدول أولاً.")
                return redirect(request.path)

        trainee_ids = [row["trainee"].pk for row in rows]
        column_dates = [col["date"] for col in columns]

        if should_process_delete:
            deleted_count = delete_saved_attendance_entries(sheet, trainee_ids, column_dates)
            log_activity(request, "delete", program=program, details=f"حذف بيانات جدول الغيابات {ATTENDANCE_PROGRAMS[program]['label']}")
            messages.success(request, f"تم حذف {deleted_count} سجل محفوظ من جدول الغيابات.")

            query = build_preserved_query(request.POST, remove_status_fields=True, remove_post_action=True)
            return redirect(request.path + (f"?{query}" if query else ""))

        existing_entries = existing_attendance_entries(program, sheet, trainee_ids, column_dates)
        change_set = build_attendance_changes(
            program=program,
            rows=rows,
            post_data=request.POST,
            sheet=sheet,
            user=request.user,
            existing_entries=existing_entries,
        )
        persist_attendance_changes(
            to_delete_ids=change_set["to_delete_ids"],
            to_create=change_set["to_create"],
            to_update=change_set["to_update"],
        )
        saved = change_set["saved"]

        action_scope = dict(scope)
        action_scope["visible_dates"] = [col["date"] for col in columns]
        sync_summary = _sync_attendance_actions(program, build_monthly_action_payload(program, action_scope, _attendance_queryset), request.user)
        log_activity(request, "change", program=program, details=f"حفظ جدول الغيابات {ATTENDANCE_PROGRAMS[program]['label']}")
        if sync_summary["created"] or sync_summary.get("archived"):
            details = []
            if sync_summary["created"]:
                details.append(f"إضافة {sync_summary['created']} إجراء جديد تلقائيًا")
            if sync_summary.get("archived"):
                details.append(f"أرشفة {sync_summary['archived']} إجراء سابق")
            messages.success(request, f"تم حفظ {saved} خلية في جدول الغيابات، و{ '، '.join(details) } في جدول الأعذار.")
        else:
            messages.success(request, f"تم حفظ {saved} خلية في جدول الغيابات.")

        query = build_preserved_query(request.POST, remove_status_fields=True)
        return redirect(request.path + (f"?{query}" if query else ""))

    weekday_choices = _attendance_weekday_choices_for_program(program)
    month_choices = [{"value": value, "label": label} for value, label in ATTENDANCE_MONTH_CHOICES]
    log_activity(request, "view", program=program, details=f"عرض جدول غيابات {ATTENDANCE_PROGRAMS[program]['label']}")
    return render(request, "trainees/attendance_grid.html", {
        "title": f"جدول الغيابات - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "program_description": ATTENDANCE_PROGRAMS[program]["description"],
        "scope": scope,
        "show_table": payload["show_table"],
        "show_all_specialties": show_all_specialties,
        "current_action": action,
        "promotion_options": payload["promotion_options"],
        "specialty_options": payload["specialty_options"],
        "weekday_choices": weekday_choices,
        "month_choices": month_choices,
        "columns": columns,
        "rows": rows,
        "sheet": sheet,
        "is_bridge_specialty": requires_custom_weekdays,
        "requires_custom_weekdays": requires_custom_weekdays,
        "allow_third_weekday": allow_third_weekday,
        "allow_fourth_weekday": program in {"evening", "crossing"},
        "allow_fifth_weekday": program in {"evening", "crossing"},
        "can_change": has_program_permission(request.user, program, "change"),
        "status_choices": _attendance_status_choices(program),
        "status_display_map": _attendance_status_label_map(program),
        "slot_count": payload.get("slot_count", 1),
        "display_columns_count": payload.get("display_columns_count", len(columns)),
        "current_query": request.GET.urlencode(),
        "weekday_validation": payload.get("weekday_validation", {"is_valid": True, "message": ""}),
    })



def _sync_attendance_actions(program, payload, user):
    return sync_attendance_actions(program, payload, user)


@login_required
def management_overview(request):
    if not can_access_admin_panel(request.user) and not getattr(request.user, "is_superuser", False):
        messages.error(request, "غير مصرح لك بفتح لوحة المتابعة الإدارية.")
        return redirect("dashboard")

    # زر مقرر الفصل يستعمل تحويلًا من الخادم أيضًا، وليس JavaScript فقط.
    # هذا يحل حالة بقاء الصفحة نفسها إذا كان المتصفح أو الكاش يتجاهل onclick/href.
    dismissal_program = (request.GET.get("open_dismissal") or request.GET.get("dismissal") or "").strip()
    if dismissal_program in ATTENDANCE_PROGRAMS:
        return redirect(reverse("dismissal_decisions", args=[dismissal_program]) + "?scope=current")

    sanctions_program = (request.GET.get("open_sanctions") or request.GET.get("sanctions") or "").strip()
    if sanctions_program in ATTENDANCE_PROGRAMS:
        return redirect(reverse("sanction_records", args=[sanctions_program]) + "?scope=current")

    summons_program = (request.GET.get("open_summons") or request.GET.get("summons") or "").strip()
    if summons_program in ATTENDANCE_PROGRAMS:
        return redirect(reverse("summons_records", args=[summons_program]) + "?scope=current")

    return render(request, "trainees/management_overview.html", {
        "title": "المتابعة الإدارية",
        "program_cards": [
            {"code": "initial", "label": "الحضوري الأولي", "list_url": reverse("initial_list"), "attendance_url": reverse("attendance_program", args=["initial"]), "actions_url": reverse("attendance_actions", args=["initial"]), "slots_actions_url": reverse("attendance_slot_actions", args=["initial"]), "dismissal_url": reverse("dismissal_decisions", args=["initial"]) + "?scope=current", "sanctions_url": reverse("sanction_records", args=["initial"]) + "?scope=current", "summons_url": reverse("summons_records", args=["initial"]) + "?scope=current", "intermittent_summons_url": reverse("summons_records", args=["initial"]) + "?scope=current&summons_type=intermittent_absence"},
            {"code": "apprentice", "label": "عن طريق التمهين", "list_url": reverse("apprentice_list"), "attendance_url": reverse("attendance_program", args=["apprentice"]), "actions_url": reverse("attendance_actions", args=["apprentice"]), "slots_actions_url": reverse("attendance_slot_actions", args=["apprentice"]), "dismissal_url": reverse("dismissal_decisions", args=["apprentice"]) + "?scope=current", "sanctions_url": reverse("sanction_records", args=["apprentice"]) + "?scope=current", "summons_url": reverse("summons_records", args=["apprentice"]) + "?scope=current", "intermittent_summons_url": reverse("summons_records", args=["apprentice"]) + "?scope=current&summons_type=intermittent_absence"},
            {"code": "evening", "label": "الدروس المسائية", "list_url": reverse("evening_list"), "attendance_url": reverse("attendance_program", args=["evening"]), "actions_url": reverse("attendance_actions", args=["evening"]), "slots_actions_url": reverse("attendance_slot_actions", args=["evening"]), "dismissal_url": reverse("dismissal_decisions", args=["evening"]) + "?scope=current", "sanctions_url": reverse("sanction_records", args=["evening"]) + "?scope=current", "summons_url": reverse("summons_records", args=["evening"]) + "?scope=current", "intermittent_summons_url": reverse("summons_records", args=["evening"]) + "?scope=current&summons_type=intermittent_absence"},
            {"code": "crossing", "label": "المعابر", "list_url": reverse("crossing_list"), "attendance_url": reverse("attendance_program", args=["crossing"]), "actions_url": reverse("attendance_actions", args=["crossing"]), "slots_actions_url": reverse("attendance_slot_actions", args=["crossing"]), "dismissal_url": reverse("dismissal_decisions", args=["crossing"]) + "?scope=current", "sanctions_url": reverse("sanction_records", args=["crossing"]) + "?scope=current", "summons_url": reverse("summons_records", args=["crossing"]) + "?scope=current", "intermittent_summons_url": reverse("summons_records", args=["crossing"]) + "?scope=current&summons_type=intermittent_absence"},
        ],
    })


@login_required
def attendance_home(request):
    allowed_programs = _attendance_allowed_programs(request.user)
    if not allowed_programs:
        messages.error(request, "لا تملك صلاحية الوصول إلى صفحات الغيابات.")
        return redirect("dashboard")
    cards = build_attendance_home_cards(allowed_programs, ATTENDANCE_PROGRAMS)
    for card in cards:
        key = card["code"]
        slot_route_names = {
            "initial": ("attendance_initial_slots", "attendance_initial_slots_stats"),
            "apprentice": ("attendance_apprentice_slots", "attendance_apprentice_slots_stats"),
            "evening": ("attendance_evening_slots", "attendance_evening_slots_stats"),
            "crossing": ("attendance_crossing_slots", "attendance_crossing_slots_stats"),
        }
        slots_table_name, slots_stats_name = slot_route_names.get(key, (None, None))
        card.update({
            "program": key,
            "table_url": reverse("attendance_program", args=[key]),
            "stats_url": reverse("attendance_stats", args=[key]),
            "actions_url": reverse("attendance_actions", args=[key]),
            "slots_actions_url": reverse("attendance_slot_actions", args=[key]),
            "slots_table_url": reverse(slots_table_name) if slots_table_name else "",
            "slots_stats_url": reverse(slots_stats_name) if slots_stats_name else "",
            "dismissal_url": reverse("dismissal_decisions", args=[key]),
            "sanctions_url": reverse("sanction_records", args=[key]),
            "summons_url": reverse("summons_records", args=[key]),
            "intermittent_summons_url": reverse("summons_records", args=[key]) + "?scope=current&summons_type=intermittent_absence",
        })
    return render(request, "trainees/attendance_home.html", {"title": "اختيار نمط الغيابات", "cards": cards})


# -----------------------------
# Dismissal decisions / مقرر الفصل
# -----------------------------

def _dismissal_scope_label(scope):
    return "المتخرجين" if scope == "graduated" else "الحاليين"


def _valid_dismissal_scope(value):
    return "graduated" if str(value or "").strip() == "graduated" else "current"


def _dismissal_scope_for_trainee(trainee):
    end_date = getattr(trainee, "تاريخ_نهاية_التكوين", None)
    today = timezone.localdate()
    if end_date and end_date <= today:
        return "graduated"
    return "current"


def _dismissal_scope_filter_value(value):
    value = str(value or "").strip()
    if value in {"current", "graduated"}:
        return value
    return ""


def _dismissal_event_values_from_trainee(trainee):
    record_number = str(getattr(trainee, "رقم_الشطب", "") or "").strip()
    removal_date = getattr(trainee, "تاريخ_الشطب", None)
    is_removed = unified_status_code(getattr(trainee, "الحالة", "")) == "removed"
    return is_removed, removal_date, record_number


def _next_dismissal_decision_number(event_date=None):
    target_year = (event_date or timezone.localdate()).year
    qs = DismissalDecision.objects.filter(
        Q(decision_date__year=target_year) |
        Q(disciplinary_record_date__year=target_year) |
        Q(dismissal_start_date__year=target_year)
    )
    max_num = 0
    max_width = 0
    for raw in qs.values_list("decision_number", flat=True):
        text = str(raw or "").strip()
        if not text:
            continue
        match = re.fullmatch(r"0*(\d+)", text)
        if not match:
            continue
        number = int(match.group(1))
        if number > max_num:
            max_num = number
            max_width = len(text) if text.startswith("0") else 0
    next_number = max_num + 1
    return str(next_number).zfill(max_width) if max_width > 1 else str(next_number)


def _archive_dismissal_decision(decision, user=None):
    changed = False
    if not decision.is_archived:
        decision.is_archived = True
        decision.archived_at = timezone.now()
        changed = True
    if user is not None:
        decision.updated_by = user
        changed = True
    if changed:
        decision.save(update_fields=["is_archived", "archived_at", "updated_by", "updated_at"])
    return changed


def _sync_auto_dismissal_decision_for_trainee(program, trainee, user=None):
    """Synchronize the active dismissal decision with the trainee administrative status.

    - removed + removal date + record number: create or restore one active decision for that exact event.
    - not removed or missing date/record: archive active decisions and keep them as administrative history.
    """
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls or trainee is None or not getattr(trainee, "pk", None):
        return {"action": "none", "decision": None, "archived": 0, "created": False, "restored": False}

    is_removed, removal_date, record_number = _dismissal_event_values_from_trainee(trainee)
    ct = ContentType.objects.get_for_model(ModelCls)
    base_qs = DismissalDecision.objects.filter(
        program=program,
        trainee_content_type=ct,
        trainee_object_id=trainee.pk,
    )

    with transaction.atomic():
        base_qs = base_qs.select_for_update()
        archived_count = 0

        if not is_removed or not removal_date or not record_number:
            for decision in list(base_qs.filter(is_archived=False)):
                if _archive_dismissal_decision(decision, user):
                    archived_count += 1
            if hasattr(trainee, "رقم_مقرر_الفصل") and archived_count:
                ModelCls.objects.filter(pk=trainee.pk).update(رقم_مقرر_الفصل="")
                setattr(trainee, "رقم_مقرر_الفصل", "")
            return {"action": "archived" if archived_count else "none", "decision": None, "archived": archived_count, "created": False, "restored": False}

        scope = _dismissal_scope_for_trainee(trainee)
        active_same = base_qs.filter(
            is_archived=False,
            disciplinary_record_number=record_number,
            disciplinary_record_date=removal_date,
        ).order_by("-id").first()

        for decision in list(base_qs.filter(is_archived=False).exclude(pk=getattr(active_same, "pk", None))):
            if _archive_dismissal_decision(decision, user):
                archived_count += 1

        restored = False
        created = False
        decision = active_same
        if decision is None:
            decision = base_qs.filter(
                is_archived=True,
                disciplinary_record_number=record_number,
                disciplinary_record_date=removal_date,
            ).order_by("-id").first()
            if decision is not None:
                decision.is_archived = False
                decision.archived_at = None
                restored = True

        if decision is None:
            decision = DismissalDecision(
                program=program,
                decision_scope=scope,
                trainee_content_type=ct,
                trainee_object_id=trainee.pk,
                trainee_name=getattr(trainee, "اللقب_والاسم", str(trainee) or ""),
                created_by=user,
                updated_by=user,
            )
            created = True

        decision.program = program
        decision.decision_scope = scope
        decision.trainee_content_type = ct
        decision.trainee_object_id = trainee.pk
        decision.sync_snapshot_from_trainee(trainee)
        decision.disciplinary_record_number = record_number
        decision.disciplinary_record_date = removal_date
        decision.dismissal_start_date = decision.dismissal_start_date or removal_date
        decision.decision_date = decision.decision_date or removal_date
        if not decision.decision_number:
            imported_number = str(getattr(trainee, "رقم_مقرر_الفصل", "") or "").strip()
            decision.decision_number = imported_number or _next_dismissal_decision_number(removal_date)
        decision.is_archived = False
        decision.archived_at = None
        if user is not None:
            decision.updated_by = user
            if created and not decision.created_by_id:
                decision.created_by = user
        decision.full_clean()
        decision.save()

        if hasattr(trainee, "رقم_مقرر_الفصل") and decision.decision_number != (getattr(trainee, "رقم_مقرر_الفصل", "") or ""):
            ModelCls.objects.filter(pk=trainee.pk).update(رقم_مقرر_الفصل=decision.decision_number)
            setattr(trainee, "رقم_مقرر_الفصل", decision.decision_number)

        action = "created" if created else ("restored" if restored else "updated")
        return {"action": action, "decision": decision, "archived": archived_count, "created": created, "restored": restored}


def _removed_trainee_filter():
    return (
        Q(الحالة__icontains="مشطوب") |
        Q(الحالة__icontains="شطب") |
        Q(الحالة__icontains="مفصول") |
        Q(الحالة__icontains="فصل") |
        Q(الحالة__icontains="متوقف") |
        Q(الحالة__icontains="موقوف") |
        Q(الحالة__icontains="توقف") |
        Q(الحالة__icontains="منقطع") |
        Q(الحالة__icontains="انسحب")
    )


def _dismissal_base_queryset(program, scope, request=None):
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return None
    graduates = scope == "graduated"
    qs = _get_ordered_rows(ModelCls, program, graduates=graduates).filter(_removed_trainee_filter())
    if request is not None:
        q = _normalized_str(request.GET.get("q"))
        semester = _normalized_str(request.GET.get("semester"))
        year = _normalized_str(request.GET.get("year"))
        promotion_id = _normalized_str(request.GET.get("promotion"))
        specialty = _normalized_str(request.GET.get("specialty"))
        if q:
            qs = qs.filter(
                Q(اللقب__icontains=q) |
                Q(الاسم__icontains=q) |
                Q(رقم_التسجيل__icontains=q) |
                Q(التخصص__icontains=q) |
                Q(رقم_الشطب__icontains=q)
            )
        if semester:
            qs = qs.filter(السداسي=semester)
        if specialty:
            qs = qs.filter(التخصص=specialty)
        if year.isdigit():
            qs = qs.filter(الدفعة__السنة=int(year))
        if promotion_id.isdigit():
            qs = qs.filter(الدفعة_id=int(promotion_id))
    return qs


def _dismissal_decisions_for_trainees(program, scope, model_cls, trainees):
    if not trainees:
        return {}
    ct = ContentType.objects.get_for_model(model_cls)
    ids = [obj.pk for obj in trainees]
    decisions = DismissalDecision.objects.filter(
        decision_scope=scope,
        trainee_content_type=ct,
        trainee_object_id__in=ids,
        is_archived=False,
    )
    decisions = filter_records_by_split_program(decisions, program).order_by("trainee_object_id", "-id")
    result = {}
    for obj in decisions:
        result.setdefault(obj.trainee_object_id, obj)
    return result


def _selected_int_ids(request):
    values = request.POST.getlist("ids") if request.method == "POST" else request.GET.getlist("ids")
    ids = []
    for value in values:
        try:
            value = int(value)
        except (TypeError, ValueError):
            continue
        if value not in ids:
            ids.append(value)
    return ids


def _ensure_dismissal_decisions(program, scope, trainee_ids, user, *, touch_snapshot=True):
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return []
    valid_qs = _dismissal_base_queryset(program, scope).filter(pk__in=trainee_ids)
    trainees = list(valid_qs)
    if not trainees:
        return []
    decisions = []
    for trainee in trainees:
        summary = _sync_auto_dismissal_decision_for_trainee(program, trainee, user)
        decision = summary.get("decision")
        if decision is not None:
            decisions.append(decision)
    decisions.sort(key=lambda obj: (obj.specialty or "", obj.trainee_name or "", obj.pk))
    return decisions


class _DismissalFilterRequest:
    def __init__(self, query_params):
        self.GET = query_params


def _dismissal_ids_from_return_query(program, scope, return_query=""):
    """Return all removed trainees matching the current page filters, not only the selected checkboxes."""
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return []
    params = QueryDict(return_query or "", mutable=True)
    params["scope"] = scope
    params.pop("page", None)
    params.pop("decision_status", None)
    fake_request = _DismissalFilterRequest(params)
    qs = _dismissal_base_queryset(program, scope, request=fake_request)
    return list(qs.values_list("pk", flat=True))


def _ensure_dismissal_decisions_for_scope_query(program, scope, return_query, user):
    trainee_ids = _dismissal_ids_from_return_query(program, scope, return_query)
    return _ensure_dismissal_decisions(program, scope, trainee_ids, user)


def _dismissal_return_url(program, scope, query=""):
    base = reverse("dismissal_decisions", args=[program])
    query = (query or "").strip()
    if query:
        return f"{base}?{query}"
    return f"{base}?scope={scope}"


@login_required
def dismissal_decisions(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    scope = _valid_dismissal_scope(request.GET.get("scope"))
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return HttpResponseBadRequest("نمط غير صالح")

    base_qs = _dismissal_base_queryset(program, scope)
    qs = _dismissal_base_queryset(program, scope, request=request)
    decision_status = (request.GET.get("decision_status") or "").strip()

    paginator = Paginator(qs, 200)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    rows = list(page_obj.object_list)
    rows = _refresh_rows_live_semesters(rows, ModelCls)
    decision_map = _dismissal_decisions_for_trainees(program, scope, ModelCls, rows)
    for row in rows:
        row.dismissal_decision = decision_map.get(row.pk)
    if decision_status:
        rows = [r for r in rows if getattr(getattr(r, "dismissal_decision", None), "status", "") == decision_status]
        page_obj.object_list = rows

    promotion_options = list(
        دفعة.objects.filter(
            id__in=base_qs.exclude(الدفعة_id__isnull=True).values_list("الدفعة_id", flat=True).distinct(),
            مفعلة=True,
        ).order_by("-السنة", "-رقم_الدورة").only("id", "اسم_الدفعة", "السنة")
    )
    specialty_options = build_specialty_options(
        base_qs.order_by().exclude(التخصص__isnull=True).exclude(التخصص="").values_list("التخصص", flat=True)
    )
    semester_options = build_semester_options(
        base_qs.order_by().exclude(السداسي__isnull=True).exclude(السداسي="").values_list("السداسي", flat=True)
    )
    year_options = list(
        base_qs.exclude(الدفعة__السنة__isnull=True).values_list("الدفعة__السنة", flat=True).distinct().order_by("-الدفعة__السنة")
    )
    current_query = request.GET.copy()
    current_query["scope"] = scope
    current_query = current_query.urlencode()
    log_activity(request, "view", program=program, details=f"عرض مقرر الفصل - {_dismissal_scope_label(scope)}")
    return render(request, "trainees/dismissal_decisions.html", {
        "title": f"مقرر الفصل - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "scope": scope,
        "scope_label": _dismissal_scope_label(scope),
        "rows": rows,
        "page_obj": page_obj,
        "promotion_options": promotion_options,
        "specialty_options": specialty_options,
        "semester_options": semester_options,
        "year_options": year_options,
        "decision_status_choices": DismissalDecision.STATUS_CHOICES,
        "selected_decision_status": decision_status,
        "filters": extract_list_filters(request.GET),
        "current_query": current_query,
        "base_current_url": reverse("dismissal_decisions", args=[program]) + "?scope=current",
        "base_graduated_url": reverse("dismissal_decisions", args=[program]) + "?scope=graduated",
        "base_archive_url": reverse("dismissal_decisions_archive", args=[program]) + "?scope=" + scope,
    })


@login_required
@require_POST
def dismissal_decisions_bulk(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        return HttpResponseBadRequest("نمط غير صالح")
    action = (request.POST.get("bulk_action") or "").strip()
    scope = _valid_dismissal_scope(request.POST.get("scope"))
    return_query = (request.POST.get("return_query") or f"scope={scope}").strip()
    trainee_ids = _selected_int_ids(request)

    if action in {"create_filtered", "create_all_scopes"}:
        require_program_permission(request, program, "change")
        scopes = [scope]
        if action == "create_all_scopes":
            scopes = ["current", "graduated"]
        all_decisions = []
        for sc in scopes:
            all_decisions.extend(_ensure_dismissal_decisions_for_scope_query(program, sc, return_query, request.user))
        # إزالة التكرار إذا صادف نفس المقرر في أكثر من نطاق بسبب تغير تاريخ نهاية التكوين.
        unique_decisions = []
        seen = set()
        for obj in all_decisions:
            if obj.pk and obj.pk not in seen:
                seen.add(obj.pk)
                unique_decisions.append(obj)
        created_count = len(unique_decisions)
        if action == "create_all_scopes":
            messages.success(request, f"تم إنشاء/ترقيم {created_count} مقرر فصل للحاليين والمتخرجين حسب الشروط الحالية.")
        else:
            messages.success(request, f"تم إنشاء/ترقيم {created_count} مقرر فصل في قسم {_dismissal_scope_label(scope)} حسب الشروط الحالية.")
        log_activity(request, "create", program=program, details=f"إنشاء/ترقيم جماعي لمقررات الفصل ({created_count})")
        return redirect(_dismissal_return_url(program, scope, return_query))

    if not trainee_ids:
        messages.error(request, "اختر متكونًا واحدًا على الأقل قبل تنفيذ العملية، أو استعمل زر إنشاء/ترقيم الكل.")
        return redirect(_dismissal_return_url(program, scope, return_query))

    if action in {"archive", "delete"}:
        require_program_permission(request, program, "change")
        ModelCls = MODEL_BY_PROGRAM.get(program)
        ct = ContentType.objects.get_for_model(ModelCls)
        qs = DismissalDecision.objects.filter(
            program=program,
            decision_scope=scope,
            trainee_content_type=ct,
            trainee_object_id__in=trainee_ids,
            is_archived=False,
        )
        count = 0
        names = []
        with transaction.atomic():
            for obj in qs.select_for_update():
                names.append(obj.trainee_name)
                if _archive_dismissal_decision(obj, request.user):
                    count += 1
        log_activity(request, "archive", program=program, details=f"أرشفة مقررات فصل ({count}) - {', '.join(names[:20])}")
        messages.success(request, f"تمت أرشفة {count} مقرر/مقررات فصل، ولم يتم حذف أي وثيقة.")
        return redirect(_dismissal_return_url(program, scope, return_query))

    required_perm = "change" if action in {"create", "edit", "preview", "print"} else "view"
    require_program_permission(request, program, required_perm)
    decisions = _ensure_dismissal_decisions(program, scope, trainee_ids, request.user)
    if not decisions:
        messages.error(request, "لم يتم العثور على متكونين مشطوبين مطابقين للاختيار الحالي.")
        return redirect(_dismissal_return_url(program, scope, return_query))
    decision_ids = [obj.pk for obj in decisions]
    query = urlencode([("ids", pk) for pk in decision_ids] + [("return_query", return_query)], doseq=True)
    if action in {"create", "edit"}:
        return redirect(reverse("dismissal_decisions_bulk_edit", args=[program]) + "?" + query)
    if action == "print":
        return redirect(reverse("dismissal_decisions_preview", args=[program]) + "?" + query + "&print=1")
    return redirect(reverse("dismissal_decisions_preview", args=[program]) + "?" + query)


@login_required
def dismissal_decisions_bulk_edit(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        return HttpResponseBadRequest("نمط غير صالح")
    require_program_permission(request, program, "change")
    if request.method == "POST":
        decision_ids = _selected_int_ids(request)
        return_query = (request.POST.get("return_query") or "").strip()
    else:
        decision_ids = _selected_int_ids(request)
        return_query = (request.GET.get("return_query") or "").strip()
    decisions_qs = filter_records_by_split_program(DismissalDecision.objects.filter(pk__in=decision_ids), program)
    decisions = list(decisions_qs.order_by("specialty", "trainee_name", "pk"))
    if not decisions:
        messages.error(request, "لم يتم العثور على مقررات الفصل المحددة.")
        return redirect(reverse("dismissal_decisions", args=[program]))

    if request.method == "POST":
        updated_count = 0
        with transaction.atomic():
            for obj in decisions:
                prefix = f"row_{obj.pk}_"
                obj.decision_number = (request.POST.get(prefix + "decision_number") or "").strip()
                obj.disciplinary_record_number = (request.POST.get(prefix + "disciplinary_record_number") or "").strip()
                obj.group_code = (request.POST.get(prefix + "group_code") or "").strip()
                obj.notes = (request.POST.get(prefix + "notes") or "").strip()
                status = (request.POST.get(prefix + "status") or "draft").strip()
                if status in dict(DismissalDecision.STATUS_CHOICES):
                    obj.status = status
                for field in ("disciplinary_record_date", "dismissal_start_date", "decision_date"):
                    raw = request.POST.get(prefix + field)
                    parsed = parse_bulk_action_date(raw, DATE_INPUT_FORMATS)
                    if raw and parsed is None:
                        messages.error(request, f"تاريخ غير صالح في مقرر: {obj.trainee_name}")
                        return redirect(request.path + "?" + urlencode([("ids", pk) for pk in decision_ids] + ([('return_query', return_query)] if return_query else []), doseq=True))
                    setattr(obj, field, parsed)
                obj.updated_by = request.user
                obj.full_clean()
                obj.save()
                updated_count += 1
        messages.success(request, f"تم حفظ {updated_count} مقرر/مقررات فصل بنجاح.")
        if return_query:
            return redirect(reverse("dismissal_decisions", args=[program]) + "?" + return_query)
        first = decisions[0]
        return redirect(reverse("dismissal_decisions", args=[program]) + f"?scope={first.decision_scope}")

    query_string = urlencode([("ids", obj.pk) for obj in decisions] + ([('return_query', return_query)] if return_query else []), doseq=True)
    return render(request, "trainees/dismissal_decision_bulk_form.html", {
        "title": f"تحرير مقرر الفصل - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "decisions": decisions,
        "return_query": return_query,
        "query_string": query_string,
        "status_choices": DismissalDecision.STATUS_CHOICES,
        "back_url": reverse("dismissal_decisions", args=[program]) + ("?" + return_query if return_query else f"?scope={decisions[0].decision_scope}"),
    })


@login_required
def dismissal_decisions_preview(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    decision_ids = _selected_int_ids(request)
    decisions_qs = filter_records_by_split_program(DismissalDecision.objects.filter(pk__in=decision_ids), program)
    decisions = list(decisions_qs.order_by("specialty", "trainee_name", "pk"))
    if not decisions:
        messages.error(request, "لا توجد مقررات فصل للمعاينة.")
        return redirect(reverse("dismissal_decisions", args=[program]))
    return_query = (request.GET.get("return_query") or "").strip()
    auto_print = (request.GET.get("print") or "") == "1"
    if auto_print:
        DismissalDecision.objects.filter(pk__in=[obj.pk for obj in decisions]).update(status="issued", updated_by=request.user)
        for obj in decisions:
            obj.status = "issued"
    log_activity(request, "view", program=program, details=f"معاينة مقرر الفصل ({len(decisions)})")
    return render(request, "trainees/dismissal_decision_preview.html", {
        "title": f"معاينة مقرر الفصل - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "decisions": decisions,
        "return_query": return_query,
        "auto_print": auto_print,
    })


@login_required
def dismissal_decisions_archive(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    scope = _dismissal_scope_filter_value(request.GET.get("scope"))
    decision_status = (request.GET.get("decision_status") or "").strip()
    specialty = _normalized_str(request.GET.get("specialty"))
    q = _normalized_str(request.GET.get("q"))

    qs = filter_records_by_split_program(DismissalDecision.objects.filter(is_archived=True), program)
    if scope:
        qs = qs.filter(decision_scope=scope)
    if decision_status:
        qs = qs.filter(status=decision_status)
    if specialty:
        qs = qs.filter(specialty=specialty)
    if q:
        qs = qs.filter(
            Q(trainee_name__icontains=q) |
            Q(registration_number__icontains=q) |
            Q(specialty__icontains=q) |
            Q(decision_number__icontains=q) |
            Q(disciplinary_record_number__icontains=q)
        )

    base_qs = filter_records_by_split_program(DismissalDecision.objects.filter(is_archived=True), program)
    specialty_options = build_specialty_options(
        base_qs.order_by().exclude(specialty__isnull=True).exclude(specialty="").values_list("specialty", flat=True)
    )
    paginator = Paginator(qs.order_by("-archived_at", "-updated_at", "-id"), 200)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    current_query = request.GET.copy()
    if scope:
        current_query["scope"] = scope
    current_query = current_query.urlencode()
    log_activity(request, "view", program=program, details="عرض أرشيف مقررات الفصل")
    return render(request, "trainees/dismissal_decisions_archive.html", {
        "title": f"أرشيف مقررات الفصل - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "scope": scope,
        "scope_label": _dismissal_scope_label(scope) if scope else "كل الأقسام",
        "decisions": list(page_obj.object_list),
        "page_obj": page_obj,
        "specialty_options": specialty_options,
        "decision_status_choices": DismissalDecision.STATUS_CHOICES,
        "selected_decision_status": decision_status,
        "filters": extract_list_filters(request.GET),
        "current_query": current_query,
        "active_current_url": reverse("dismissal_decisions", args=[program]) + "?scope=current",
        "active_graduated_url": reverse("dismissal_decisions", args=[program]) + "?scope=graduated",
        "archive_all_url": reverse("dismissal_decisions_archive", args=[program]),
        "archive_current_url": reverse("dismissal_decisions_archive", args=[program]) + "?scope=current",
        "archive_graduated_url": reverse("dismissal_decisions_archive", args=[program]) + "?scope=graduated",
    })


@login_required
@require_POST
def dismissal_decisions_archive_bulk(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        return HttpResponseBadRequest("نمط غير صالح")
    action = (request.POST.get("bulk_action") or "").strip()
    return_query = (request.POST.get("return_query") or "").strip()
    decision_ids = _selected_int_ids(request)
    if not decision_ids:
        messages.error(request, "اختر مقررًا واحدًا على الأقل.")
        url = reverse("dismissal_decisions_archive", args=[program])
        return redirect(url + ("?" + return_query if return_query else ""))

    qs = filter_records_by_split_program(DismissalDecision.objects.filter(pk__in=decision_ids, is_archived=True), program)
    if action == "restore":
        require_program_permission(request, program, "change")
        restored = 0
        with transaction.atomic():
            for obj in qs.select_for_update():
                active_qs = DismissalDecision.objects.filter(
                    program=obj.program,
                    trainee_content_type=obj.trainee_content_type,
                    trainee_object_id=obj.trainee_object_id,
                    is_archived=False,
                ).exclude(pk=obj.pk)
                for active in active_qs.select_for_update():
                    _archive_dismissal_decision(active, request.user)
                obj.is_archived = False
                obj.archived_at = None
                obj.updated_by = request.user
                obj.save(update_fields=["is_archived", "archived_at", "updated_by", "updated_at"])
                restored += 1
        log_activity(request, "restore", program=program, details=f"استرجاع {restored} مقرر فصل من الأرشيف")
        messages.success(request, f"تم استرجاع {restored} مقرر/مقررات فصل من الأرشيف.")
        url = reverse("dismissal_decisions_archive", args=[program])
        return redirect(url + ("?" + return_query if return_query else ""))

    require_program_permission(request, program, "view")
    query = urlencode([("ids", pk) for pk in decision_ids] + ([('return_query', return_query)] if return_query else []), doseq=True)
    if action == "print":
        return redirect(reverse("dismissal_decisions_preview", args=[program]) + "?" + query + "&print=1")
    return redirect(reverse("dismissal_decisions_preview", args=[program]) + "?" + query)


# -----------------------------
# Sanctions / العقوبات
# -----------------------------

def _sanction_scope_label(scope):
    return "المتخرجين" if scope == "graduated" else "الحاليين"


def _valid_sanction_scope(value):
    return "graduated" if str(value or "").strip() == "graduated" else "current"


def _sanction_base_queryset(program, scope, request=None):
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return None
    graduates = scope == "graduated"
    # العقوبات والاستدعاءات والوثائق الإدارية العادية لا تشمل المتكونين المشطوبين/المفصولين.
    # هؤلاء يبقون في مقرر الفصل والأرشيف التاريخي فقط.
    qs = exclude_inactive_trainees(_get_ordered_rows(ModelCls, program, graduates=graduates))
    if request is not None:
        q = _normalized_str(request.GET.get("q"))
        semester = _normalized_str(request.GET.get("semester"))
        year = _normalized_str(request.GET.get("year"))
        promotion_id = _normalized_str(request.GET.get("promotion"))
        specialty = _normalized_str(request.GET.get("specialty"))
        if q:
            qs = qs.filter(
                Q(اللقب__icontains=q) |
                Q(الاسم__icontains=q) |
                Q(رقم_التسجيل__icontains=q) |
                Q(التخصص__icontains=q)
            )
        if semester:
            qs = qs.filter(السداسي=semester)
        if specialty:
            qs = qs.filter(التخصص=specialty)
        if year.isdigit():
            qs = qs.filter(الدفعة__السنة=int(year))
        if promotion_id.isdigit():
            qs = qs.filter(الدفعة_id=int(promotion_id))
    return qs


def _sanction_records_for_trainees(program, scope, model_cls, trainees):
    if not trainees:
        return {}
    ct = ContentType.objects.get_for_model(model_cls)
    ids = [obj.pk for obj in trainees]
    records = SanctionRecord.objects.filter(
        sanction_scope=scope,
        trainee_content_type=ct,
        trainee_object_id__in=ids,
    )
    records = filter_records_by_split_program(records, program)
    return {obj.trainee_object_id: obj for obj in records}


def _ensure_sanction_records(program, scope, trainee_ids, user, *, touch_snapshot=True):
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return []
    valid_qs = _sanction_base_queryset(program, scope).filter(pk__in=trainee_ids)
    trainees = list(valid_qs)
    if not trainees:
        return []
    ct = ContentType.objects.get_for_model(ModelCls)
    records = []
    with transaction.atomic():
        for trainee in trainees:
            record, created = SanctionRecord.objects.get_or_create(
                program=program,
                sanction_scope=scope,
                trainee_content_type=ct,
                trainee_object_id=trainee.pk,
                defaults={
                    "trainee_name": getattr(trainee, "اللقب_والاسم", str(trainee) or ""),
                    "created_by": user,
                    "updated_by": user,
                },
            )
            if touch_snapshot or created:
                record.sync_snapshot_from_trainee(trainee)
            record.updated_by = user
            if created and not record.created_by_id:
                record.created_by = user
            record.save()
            records.append(record)
    records.sort(key=lambda obj: (obj.specialty or "", obj.trainee_name or "", obj.pk))
    return records


def _existing_sanction_records_for_selected_trainees(program, scope, trainee_ids):
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls or not trainee_ids:
        return SanctionRecord.objects.none()
    ct = ContentType.objects.get_for_model(ModelCls)
    qs = SanctionRecord.objects.filter(
        sanction_scope=scope,
        trainee_content_type=ct,
        trainee_object_id__in=trainee_ids,
    )
    return filter_records_by_split_program(qs, program)


def _sanction_return_url(program, scope, query=""):
    base = reverse("sanction_records", args=[program])
    query = (query or "").strip()
    if query:
        return f"{base}?{query}"
    return f"{base}?scope={scope}"


@login_required
def sanction_records(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    scope = _valid_sanction_scope(request.GET.get("scope"))
    ModelCls = MODEL_BY_PROGRAM.get(program)
    if not ModelCls:
        return HttpResponseBadRequest("نمط غير صالح")

    base_qs = _sanction_base_queryset(program, scope)
    qs = _sanction_base_queryset(program, scope, request=request)
    status_filter = (request.GET.get("status") or "").strip()
    archive_state = (request.GET.get("archive_state") or "active").strip()

    paginator = Paginator(qs, 200)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    rows = list(page_obj.object_list)
    rows = _refresh_rows_live_semesters(rows, ModelCls)
    record_map = _sanction_records_for_trainees(program, scope, ModelCls, rows)
    for row in rows:
        row.sanction_record = record_map.get(row.pk)

    if archive_state == "archived":
        rows = [r for r in rows if getattr(getattr(r, "sanction_record", None), "is_archived", False)]
    elif archive_state == "active":
        rows = [r for r in rows if not getattr(getattr(r, "sanction_record", None), "is_archived", False)]

    if status_filter:
        rows = [r for r in rows if getattr(getattr(r, "sanction_record", None), "status", "") == status_filter]
        page_obj.object_list = rows

    promotion_options = list(
        دفعة.objects.filter(
            id__in=base_qs.exclude(الدفعة_id__isnull=True).values_list("الدفعة_id", flat=True).distinct(),
            مفعلة=True,
        ).order_by("-السنة", "-رقم_الدورة").only("id", "اسم_الدفعة", "السنة")
    )
    specialty_options = build_specialty_options(
        base_qs.order_by().exclude(التخصص__isnull=True).exclude(التخصص="").values_list("التخصص", flat=True)
    )
    semester_options = build_semester_options(
        base_qs.order_by().exclude(السداسي__isnull=True).exclude(السداسي="").values_list("السداسي", flat=True)
    )
    year_options = list(
        base_qs.exclude(الدفعة__السنة__isnull=True).values_list("الدفعة__السنة", flat=True).distinct().order_by("-الدفعة__السنة")
    )
    current_query = request.GET.copy()
    current_query["scope"] = scope
    current_query = current_query.urlencode()
    summary = {
        "total": len(rows),
        "missing": sum(1 for row in rows if not getattr(row, "sanction_record", None)),
        "ready": sum(1 for row in rows if getattr(getattr(row, "sanction_record", None), "status", "") == "ready"),
        "issued": sum(1 for row in rows if getattr(getattr(row, "sanction_record", None), "status", "") == "issued"),
        "delivered": sum(1 for row in rows if getattr(getattr(row, "sanction_record", None), "status", "") == "delivered"),
    }
    log_activity(request, "view", program=program, details=f"عرض العقوبات - {_sanction_scope_label(scope)}")
    return render(request, "trainees/sanctions.html", {
        "title": f"العقوبات - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "scope": scope,
        "scope_label": _sanction_scope_label(scope),
        "rows": rows,
        "page_obj": page_obj,
        "summary": summary,
        "promotion_options": promotion_options,
        "specialty_options": specialty_options,
        "semester_options": semester_options,
        "year_options": year_options,
        "status_choices": SanctionRecord.STATUS_CHOICES,
        "archive_state_choices": SanctionRecord.ARCHIVE_STATE_CHOICES,
        "selected_status": status_filter,
        "selected_archive_state": archive_state,
        "filters": extract_list_filters(request.GET),
        "current_query": current_query,
        "base_current_url": reverse("sanction_records", args=[program]) + "?scope=current",
        "base_graduated_url": reverse("sanction_records", args=[program]) + "?scope=graduated",
    })


@login_required
@require_POST
def sanction_records_bulk(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        return HttpResponseBadRequest("نمط غير صالح")
    action = (request.POST.get("bulk_action") or "").strip()
    scope = _valid_sanction_scope(request.POST.get("scope"))
    return_query = (request.POST.get("return_query") or f"scope={scope}").strip()
    trainee_ids = _selected_int_ids(request)
    if not trainee_ids:
        messages.error(request, "اختر متكونًا واحدًا على الأقل قبل تنفيذ العملية.")
        return redirect(_sanction_return_url(program, scope, return_query))

    if action in {"archive", "restore", "delete"}:
        required_perm = "delete" if action == "delete" else "change"
        require_program_permission(request, program, required_perm)
        qs = _existing_sanction_records_for_selected_trainees(program, scope, trainee_ids)
        if action == "archive":
            updated = qs.filter(is_archived=False).update(is_archived=True, archived_at=timezone.now(), updated_by=request.user)
            messages.success(request, f"تم حفظ {updated} عقوبة/عقوبات في الأرشيف.")
            return redirect(_sanction_return_url(program, scope, return_query))
        if action == "restore":
            updated = qs.filter(is_archived=True).update(is_archived=False, archived_at=None, updated_by=request.user)
            messages.success(request, f"تم استرجاع {updated} عقوبة/عقوبات من الأرشيف.")
            return redirect(_sanction_return_url(program, scope, return_query))
        count = qs.count()
        qs.delete()
        log_activity(request, "delete", program=program, details=f"حذف عقوبات ({count})")
        messages.success(request, f"تم حذف {count} عقوبة/عقوبات فقط دون حذف بيانات المتكونين.")
        return redirect(_sanction_return_url(program, scope, return_query))

    required_perm = "change" if action in {"create", "edit", "preview", "print"} else "view"
    require_program_permission(request, program, required_perm)
    records = _ensure_sanction_records(program, scope, trainee_ids, request.user)
    if not records:
        messages.error(request, "لم يتم العثور على متكونين مطابقين للاختيار الحالي.")
        return redirect(_sanction_return_url(program, scope, return_query))
    record_ids = [obj.pk for obj in records]
    query = urlencode([("ids", pk) for pk in record_ids] + [("return_query", return_query)], doseq=True)
    if action in {"create", "edit"}:
        return redirect(reverse("sanction_records_bulk_edit", args=[program]) + "?" + query)
    if action == "print":
        return redirect(reverse("sanction_records_preview", args=[program]) + "?" + query + "&print=1")
    return redirect(reverse("sanction_records_preview", args=[program]) + "?" + query)


@login_required
def sanction_records_bulk_edit(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        return HttpResponseBadRequest("نمط غير صالح")
    require_program_permission(request, program, "change")
    if request.method == "POST":
        record_ids = _selected_int_ids(request)
        return_query = (request.POST.get("return_query") or "").strip()
    else:
        record_ids = _selected_int_ids(request)
        return_query = (request.GET.get("return_query") or "").strip()
    records_qs = filter_records_by_split_program_for_active_trainees(
        SanctionRecord.objects.filter(pk__in=record_ids),
        program,
        MODEL_BY_PROGRAM.get(program),
    )
    records = list(records_qs.order_by("specialty", "trainee_name", "pk"))
    if not records:
        messages.error(request, "لم يتم العثور على العقوبات المحددة.")
        return redirect(reverse("sanction_records", args=[program]))

    if request.method == "POST":
        updated_count = 0
        with transaction.atomic():
            for obj in records:
                prefix = f"row_{obj.pk}_"
                obj.document_number = (request.POST.get(prefix + "document_number") or "").strip()
                obj.sanction_text = (request.POST.get(prefix + "sanction_text") or "").strip()
                obj.disciplinary_record_number = (request.POST.get(prefix + "disciplinary_record_number") or "").strip()
                obj.group_code = (request.POST.get(prefix + "group_code") or "").strip()
                obj.semester = (request.POST.get(prefix + "semester") or "").strip()
                obj.notes = (request.POST.get(prefix + "notes") or "").strip()
                status = (request.POST.get(prefix + "status") or "draft").strip()
                if status in dict(SanctionRecord.STATUS_CHOICES):
                    obj.status = status
                for field in ("disciplinary_record_date", "decision_date"):
                    raw = request.POST.get(prefix + field)
                    parsed = parse_bulk_action_date(raw, DATE_INPUT_FORMATS)
                    if raw and parsed is None:
                        messages.error(request, f"تاريخ غير صالح في عقوبة: {obj.trainee_name}")
                        return redirect(request.path + "?" + urlencode([("ids", pk) for pk in record_ids] + ([("return_query", return_query)] if return_query else []), doseq=True))
                    setattr(obj, field, parsed)
                obj.updated_by = request.user
                obj.full_clean()
                obj.save()
                updated_count += 1
        messages.success(request, f"تم حفظ {updated_count} عقوبة/عقوبات بنجاح.")
        if return_query:
            return redirect(reverse("sanction_records", args=[program]) + "?" + return_query)
        first = records[0]
        return redirect(reverse("sanction_records", args=[program]) + f"?scope={first.sanction_scope}")

    query_string = urlencode([("ids", obj.pk) for obj in records] + ([("return_query", return_query)] if return_query else []), doseq=True)
    return render(request, "trainees/sanction_bulk_form.html", {
        "title": f"تحرير العقوبات - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "records": records,
        "return_query": return_query,
        "query_string": query_string,
        "status_choices": SanctionRecord.STATUS_CHOICES,
        "back_url": reverse("sanction_records", args=[program]) + ("?" + return_query if return_query else f"?scope={records[0].sanction_scope}"),
    })


@login_required
def sanction_records_preview(request, program):
    if program not in ATTENDANCE_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    record_ids = _selected_int_ids(request)
    records_qs = filter_records_by_split_program_for_active_trainees(
        SanctionRecord.objects.filter(pk__in=record_ids),
        program,
        MODEL_BY_PROGRAM.get(program),
    )
    records = list(records_qs.order_by("specialty", "trainee_name", "pk"))
    if not records:
        messages.error(request, "لا توجد عقوبات للمعاينة.")
        return redirect(reverse("sanction_records", args=[program]))
    return_query = (request.GET.get("return_query") or "").strip()
    auto_print = (request.GET.get("print") or "") == "1"
    if auto_print:
        SanctionRecord.objects.filter(pk__in=[obj.pk for obj in records]).update(status="issued", updated_by=request.user)
        for obj in records:
            obj.status = "issued"
    log_activity(request, "view", program=program, details=f"معاينة العقوبات ({len(records)})")
    return render(request, "trainees/sanction_preview.html", {
        "title": f"معاينة العقوبات - {ATTENDANCE_PROGRAMS[program]['label']}",
        "program": program,
        "program_label": ATTENDANCE_PROGRAMS[program]["label"],
        "records": records,
        "return_query": return_query,
        "auto_print": auto_print,
    })


@login_required
def attendance_export(request, program, fmt):
    # القوالب القديمة تستدعي الرابط بصيغة /export/excel/ بينما اسم الملف الحقيقي xlsx.
    # لذلك نقبل excel كاسم صديق ونحوّله داخليًا إلى xlsx بدل إظهار 404.
    fmt = (fmt or "").strip().lower()
    if fmt == "excel":
        fmt = "xlsx"
    if fmt not in {"xlsx", "pdf"}:
        raise Http404()
    require_program_permission(request, program, "view")
    payload = _attendance_table_payload(program, request)
    if not payload.get("sheet"):
        messages.error(request, "اعرض جدول الغيابات أولاً أو أكمل الفلاتر المطلوبة قبل التصدير.")
        return redirect(reverse("attendance_program", args=[program]) + ("?" + request.GET.urlencode() if request.GET else ""))
    if fmt == "xlsx":
        return _attendance_export_excel(program, payload)
    return _attendance_export_pdf(program, payload)



def _audit_compact_json(value, limit=220):
    if not value:
        return ""
    try:
        rendered = json.dumps(value, ensure_ascii=False, sort_keys=True)
    except Exception:
        rendered = str(value)
    if len(rendered) > limit:
        return rendered[:limit].rstrip() + " ..."
    return rendered


def _audit_is_sensitive(item):
    haystack = " ".join([
        item.action or "",
        item.method or "",
        item.screen_name or "",
        item.view_name or "",
        item.model_label or "",
        item.object_repr or "",
        item.details or "",
    ]).lower()
    sensitive_terms = [
        "delete", "حذف", "restore", "استرجاع", "backup", "نسخ احتياطي", "استيراد",
        "import", "export", "permission", "صلاحية", "account", "user", "license", "update",
        "ترخيص", "تحديث", "qr", "image", "صورة", "غياب", "archiv", "أرشفة",
    ]
    return any(term in haystack for term in sensitive_terms) or (item.method or "").upper() in {"DELETE", "PATCH"}


def _build_audit_filtered_queryset(request):
    qs = ComprehensiveAuditLog.objects.select_related("user").all()
    filters = {
        "q": (request.GET.get("q") or "").strip(),
        "username": (request.GET.get("username") or "").strip(),
        "action": (request.GET.get("action") or "").strip(),
        "success": (request.GET.get("success") or "").strip(),
        "method": (request.GET.get("method") or "").strip().upper(),
        "screen": (request.GET.get("screen") or "").strip(),
        "model": (request.GET.get("model") or "").strip(),
        "object": (request.GET.get("object") or "").strip(),
        "date_from": (request.GET.get("date_from") or "").strip(),
        "date_to": (request.GET.get("date_to") or "").strip(),
        "sensitive_only": (request.GET.get("sensitive_only") or "").strip(),
    }

    if filters["q"]:
        q = filters["q"]
        qs = qs.filter(
            Q(details__icontains=q)
            | Q(path__icontains=q)
            | Q(object_repr__icontains=q)
            | Q(object_pk__icontains=q)
            | Q(view_name__icontains=q)
            | Q(screen_name__icontains=q)
            | Q(model_label__icontains=q)
            | Q(username_snapshot__icontains=q)
        )
    if filters["username"]:
        qs = qs.filter(username_snapshot__icontains=filters["username"])
    if filters["action"]:
        qs = qs.filter(action=filters["action"])
    if filters["success"] == "1":
        qs = qs.filter(success=True)
    elif filters["success"] == "0":
        qs = qs.filter(success=False)
    if filters["method"]:
        qs = qs.filter(method=filters["method"])
    if filters["screen"]:
        qs = qs.filter(screen_name__icontains=filters["screen"])
    if filters["model"]:
        qs = qs.filter(model_label__icontains=filters["model"])
    if filters["object"]:
        needle = filters["object"]
        qs = qs.filter(
            Q(object_repr__icontains=needle)
            | Q(object_pk__icontains=needle)
            | Q(details__icontains=needle)
            | Q(before_data__icontains=needle)
            | Q(after_data__icontains=needle)
        )
    if filters["date_from"]:
        qs = qs.filter(created_at__date__gte=filters["date_from"])
    if filters["date_to"]:
        qs = qs.filter(created_at__date__lte=filters["date_to"])
    if filters["sensitive_only"] == "1":
        qs = qs.filter(
            Q(method__in=["DELETE", "PATCH"])
            | Q(details__icontains="حذف")
            | Q(details__icontains="استيراد")
            | Q(details__icontains="تحديث")
            | Q(details__icontains="صلاح")
            | Q(details__icontains="ترخيص")
            | Q(details__icontains="نسخ احتياطي")
            | Q(screen_name__icontains="إدارة")
            | Q(model_label__icontains="user")
        )
    return qs, filters


def _audit_export_rows(qs):
    rows = []
    for item in qs.iterator():
        rows.append([
            timezone.localtime(item.created_at).strftime("%Y-%m-%d %H:%M:%S") if item.created_at else "",
            item.username_snapshot or (item.user.username if item.user else ""),
            item.get_action_display(),
            item.method or "",
            item.screen_name or "",
            item.view_name or "",
            item.model_label or "",
            item.object_pk or "",
            item.object_repr or "",
            "نعم" if item.success else "لا",
            item.status_code or "",
            item.ip_address or "",
            item.path or "",
            item.details or "",
            json.dumps(item.before_data or {}, ensure_ascii=False),
            json.dumps(item.after_data or {}, ensure_ascii=False),
        ])
    return rows


@login_required
def comprehensive_audit_export(request):
    if not can_access_admin_panel(request.user) and not getattr(request.user, "is_superuser", False):
        messages.error(request, "غير مصرح لك بتصدير السجل الشامل.")
        return redirect("dashboard")

    qs, filters = _build_audit_filtered_queryset(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"
    headers = [
        "التاريخ", "المستخدم", "نوع السجل", "الطريقة", "الشاشة", "العرض", "نوع الكيان",
        "المعرف", "وصف السجل", "نجاح", "حالة الاستجابة", "IP", "المسار", "التفاصيل", "قبل", "بعد",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in _audit_export_rows(qs[:5000]):
        ws.append(row)
    from openpyxl.utils import get_column_letter
    widths = [20,18,16,10,24,24,22,12,28,10,12,16,30,40,40,40]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = content_disposition_header("attachment", "comprehensive_audit.xlsx")
    wb.save(response)
    return response


@login_required
def comprehensive_audit_detail(request, pk):
    if not can_access_admin_panel(request.user) and not getattr(request.user, "is_superuser", False):
        messages.error(request, "غير مصرح لك بفتح تفاصيل السجل.")
        return redirect("dashboard")
    item = get_object_or_404(ComprehensiveAuditLog.objects.select_related("user"), pk=pk)
    context = {
        "title": "تفاصيل سجل العملية",
        "item": item,
        "before_json": json.dumps(item.before_data or {}, ensure_ascii=False, indent=2),
        "after_json": json.dumps(item.after_data or {}, ensure_ascii=False, indent=2),
        "is_sensitive": _audit_is_sensitive(item),
    }
    return render(request, "trainees/comprehensive_audit_detail.html", context)


@login_required
def comprehensive_audit_dashboard(request):
    if not can_access_admin_panel(request.user) and not getattr(request.user, "is_superuser", False):
        messages.error(request, "غير مصرح لك بفتح السجل الشامل للعمليات.")
        return redirect("dashboard")

    qs, filters = _build_audit_filtered_queryset(request)

    action_choices = list(ComprehensiveAuditLog.ACTION_CHOICES)
    method_choices = [m for m in qs.order_by().values_list("method", flat=True).distinct() if m]

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    summary = qs.aggregate(
        total=Count("id"),
        success_count=Count("id", filter=Q(success=True)),
        failed_count=Count("id", filter=Q(success=False)),
    )

    items = []
    for item in page_obj.object_list:
        item.before_data_compact = _audit_compact_json(item.before_data)
        item.after_data_compact = _audit_compact_json(item.after_data)
        item.details_compact = (item.details or "")[:260] + (" ..." if item.details and len(item.details) > 260 else "")
        item.is_sensitive = _audit_is_sensitive(item)
        items.append(item)

    base_qs = request.GET.copy()
    if "page" in base_qs:
        del base_qs["page"]

    return render(request, "trainees/comprehensive_audit_dashboard.html", {
        "title": "السجل الشامل للعمليات",
        "items": items,
        "page_obj": page_obj,
        "filters": filters,
        "action_choices": action_choices,
        "method_choices": method_choices,
        "summary": summary,
        "query_string": base_qs.urlencode(),
    })
