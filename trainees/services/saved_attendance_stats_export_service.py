from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak


def _write_saved_attendance_stats_excel_sheet(
    ws,
    context,
    detail_rows,
    *,
    safe_sheet_title,
    apply_official_stats_excel_header,
    specialty_label="",
    sheet_title=None,
):
    detail_scope = context["detail_scope"]
    show_specialty_column = not bool(detail_scope.get("specialty")) and not specialty_label
    semester_label = detail_scope.get("semester") or ""
    batch = detail_scope.get("batch")
    batch_display = f"الدفعة: {batch.السنة} / {batch.رقم_الدورة}" if batch else ""

    ws.title = safe_sheet_title(sheet_title or specialty_label or "الإحصائيات المحفوظة")
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A15"

    latest_saved_at = max((item.updated_at for item in detail_rows), default=None)
    average_absence_rate = round(sum(item.absence_rate for item in detail_rows) / len(detail_rows), 2) if detail_rows else 0
    summary_rows = [
        ["عدد المتكونين", len(detail_rows)],
        ["متوسط نسبة الغياب", f"{average_absence_rate}%"],
        ["آخر حفظ", latest_saved_at.strftime("%Y-%m-%d %H:%M") if latest_saved_at else ""],
    ]

    headers = ["الاسم واللقب"]
    if show_specialty_column:
        headers.append("التخصص")
    headers.extend(["عدد الحضور", "عدد الغيابات", "غياب بعذر", "عدد التأخرات", "المجموع المسجل", "نسبة الغياب", "تاريخ الحفظ"])

    total_columns = len(headers)
    official_specialty = specialty_label or (detail_scope.get("specialty_display") if detail_scope.get("specialty") else "") or ""
    apply_official_stats_excel_header(
        ws,
        total_columns,
        detail_scope["program"],
        detail_scope,
        specialty_label=official_specialty,
        semester_label=semester_label,
        batch_display=batch_display,
    )

    summary_start_row = 11
    for offset, item in enumerate(summary_rows):
        row_num = summary_start_row + offset
        ws.cell(row=row_num, column=1, value=item[0])
        ws.cell(row=row_num, column=2, value=item[1])

    header_row = summary_start_row + len(summary_rows) + 1
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)

    data_start_row = header_row + 1
    for r, row in enumerate(detail_rows, start=0):
        values = [row.trainee_name]
        if show_specialty_column:
            values.append(row.trainee_specialty or "")
        values.extend([
            row.present_count,
            row.absent_count,
            row.excused_count,
            row.late_count,
            row.total_recorded,
            row.absence_rate,
            row.updated_at.strftime("%Y-%m-%d %H:%M"),
        ])
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=data_start_row + r, column=col_idx, value=value)

    final_row = data_start_row + len(detail_rows)
    if detail_rows:
        total_row = ["الإجمالي"]
        if show_specialty_column:
            total_row.append("")
        total_row.extend([
            sum(item.present_count for item in detail_rows),
            sum(item.absent_count for item in detail_rows),
            sum(item.excused_count for item in detail_rows),
            sum(item.late_count for item in detail_rows),
            sum(item.total_recorded for item in detail_rows),
            average_absence_rate,
            "",
        ])
        for col_idx, value in enumerate(total_row, start=1):
            ws.cell(row=final_row, column=col_idx, value=value)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9E2F3")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for row_num in range(summary_start_row, summary_start_row + len(summary_rows)):
        ws.row_dimensions[row_num].height = 23
        for col_idx in range(1, 3):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=1).fill = subtitle_fill
        ws.cell(row=row_num, column=2).font = Font(bold=True)

    for col_idx in range(1, total_columns + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    end_data_row = final_row if detail_rows else data_start_row - 1
    for row_cells in ws.iter_rows(min_row=data_start_row, max_row=end_data_row, min_col=1, max_col=total_columns):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    if detail_rows:
        for cell in ws[final_row]:
            cell.font = Font(bold=True)
            cell.fill = subtitle_fill

    from openpyxl.utils import get_column_letter

    for col_idx in range(1, total_columns + 1):
        if col_idx == 1:
            width = 28
        elif show_specialty_column and col_idx == 2:
            width = 24
        else:
            width = 16
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 18)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 26)



def build_saved_attendance_stats_excel_response(
    context,
    *,
    safe_sheet_title,
    apply_official_stats_excel_header,
    group_items_by_specialty,
    attendance_export_filename,
    finalize_workbook_response,
):
    detail_scope = context["detail_scope"]
    detail_rows = context["detail_rows"]

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    specialty = (detail_scope.get("specialty") or "").strip()
    if specialty:
        ws = wb.create_sheet(title=safe_sheet_title(specialty))
        _write_saved_attendance_stats_excel_sheet(
            ws,
            context,
            detail_rows,
            safe_sheet_title=safe_sheet_title,
            apply_official_stats_excel_header=apply_official_stats_excel_header,
            sheet_title=specialty,
            specialty_label=specialty,
        )
    else:
        groups = group_items_by_specialty(detail_rows, lambda row: row.trainee_specialty or "بدون تخصص")
        for idx, (specialty_name, specialty_rows) in enumerate(groups):
            ws = wb.create_sheet(title=safe_sheet_title(specialty_name or f"تخصص {idx + 1}"))
            _write_saved_attendance_stats_excel_sheet(
                ws,
                context,
                specialty_rows,
                safe_sheet_title=safe_sheet_title,
                apply_official_stats_excel_header=apply_official_stats_excel_header,
                sheet_title=specialty_name,
                specialty_label=specialty_name,
            )

    filename = attendance_export_filename("saved_stats", detail_scope["program"], detail_scope, specialty=specialty, ext="xlsx")
    return finalize_workbook_response(wb, filename)



def build_saved_attendance_stats_pdf_response(
    context,
    *,
    register_pdf_font,
    pdf_text,
    header_lines,
    attendance_stats_template_title,
    attendance_stats_scope_subtitle,
    pdf_row,
    attendance_export_filename,
    set_download_filename,
    group_items_by_specialty,
):
    detail_scope = context["detail_scope"]
    detail_rows = context["detail_rows"]
    detail_summary = context["detail_summary"]
    show_specialty_column = not bool(detail_scope.get("specialty"))

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=16, bottomMargin=16)
    font_name = register_pdf_font()

    batch = detail_scope.get("batch")
    batch_display = f"الدفعة: {batch.السنة} / {batch.رقم_الدورة}" if batch else ""

    grouped_rows = [(detail_scope.get("specialty") or "كل التخصصات", detail_rows)]
    if not detail_scope.get("specialty"):
        grouped_rows = group_items_by_specialty(detail_rows, lambda row: row.trainee_specialty or "بدون تخصص")

    dark = colors.HexColor("#1F4E78")
    light = colors.HexColor("#D9E2F3")
    white = colors.white
    story = []

    for idx2, (specialty_name, specialty_rows) in enumerate(grouped_rows, start=1):
        average_absence_rate = round(sum(item.absence_rate for item in specialty_rows) / len(specialty_rows), 2) if specialty_rows else 0
        latest_saved_at = max((item.updated_at for item in specialty_rows), default=detail_summary.get("latest_saved_at"))
        semester_label = detail_scope.get("semester") or ""

        header_table = Table([[pdf_text(line)] for line in header_lines], colWidths=[doc.width], hAlign="CENTER")
        header_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 1), 12),
            ("FONTSIZE", (0, 2), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 6))

        official_specialty = specialty_name if specialty_name != "كل التخصصات" else (detail_scope.get("specialty") or "")
        title_table = Table([[pdf_text(attendance_stats_template_title(detail_scope["program"], official_specialty))]], colWidths=[doc.width], hAlign="CENTER")
        title_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), dark),
            ("TEXTCOLOR", (0, 0), (-1, -1), white),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 15),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(title_table)

        subtitle_table = Table([[pdf_text(attendance_stats_scope_subtitle(detail_scope, specialty_label="", semester_label=semester_label, batch_display=batch_display))]], colWidths=[doc.width], hAlign="CENTER")
        subtitle_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), light),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(subtitle_table)
        story.append(Spacer(1, 8))

        summary_data = [
            pdf_row(["عدد المتكونين", str(len(specialty_rows))]),
            pdf_row(["متوسط نسبة الغياب", f"{average_absence_rate}%"]),
            pdf_row(["آخر حفظ", latest_saved_at.strftime("%Y-%m-%d %H:%M") if latest_saved_at else ""]),
        ]
        summary_table = Table(summary_data, colWidths=[150, 260], hAlign="RIGHT")
        summary_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (0, -1), light),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 10))

        headers = ["الاسم واللقب"]
        if show_specialty_column:
            headers.append("التخصص")
        headers.extend(["عدد الحضور", "عدد الغيابات", "غياب بعذر", "عدد التأخرات", "المجموع المسجل", "نسبة الغياب", "تاريخ الحفظ"])
        data = [pdf_row(headers)]

        for row in specialty_rows:
            line = [row.trainee_name]
            if show_specialty_column:
                line.append(row.trainee_specialty or "")
            line.extend([
                row.present_count,
                row.absent_count,
                row.excused_count,
                row.late_count,
                row.total_recorded,
                f"{row.absence_rate}%",
                row.updated_at.strftime("%Y-%m-%d %H:%M"),
            ])
            data.append(pdf_row(line))

        if specialty_rows:
            total_line = ["الإجمالي"]
            if show_specialty_column:
                total_line.append("")
            total_line.extend([
                sum(item.present_count for item in specialty_rows),
                sum(item.absent_count for item in specialty_rows),
                sum(item.excused_count for item in specialty_rows),
                sum(item.late_count for item in specialty_rows),
                sum(item.total_recorded for item in specialty_rows),
                f"{average_absence_rate}%",
                "",
            ])
            data.append(pdf_row(total_line))

        col_widths = [150]
        if show_specialty_column:
            col_widths.append(95)
        col_widths.extend([58, 58, 58, 58, 62, 62, 95])
        table = Table(data, repeatRows=1, colWidths=col_widths, hAlign="CENTER")
        style_cmds = [
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), dark),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#F4F7FB")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        if specialty_rows:
            style_cmds.append(("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), light))
        table.setStyle(TableStyle(style_cmds))
        story.append(table)

        if idx2 < len(grouped_rows):
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    specialty = detail_scope.get("specialty") or ""
    filename = attendance_export_filename("saved_stats", detail_scope["program"], detail_scope, specialty=specialty, ext="pdf")
    return set_download_filename(response, filename)
