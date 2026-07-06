from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    import arabic_reshaper
except Exception:  # pragma: no cover
    arabic_reshaper = None

try:
    from bidi.algorithm import get_display
except Exception:  # pragma: no cover
    get_display = None

PDF_ALLOWED_FIELDS = {
    "الرقم_التعريفي",
    "اللقب",
    "الاسم",
    "التخصص",
    "رقم_التسجيل",
    "السداسي",
    "الحالة",
    "تاريخ_بداية_التكوين",
    "تاريخ_نهاية_التكوين",
}


def register_pdf_font():
    candidates = [
        Path(settings.BASE_DIR) / "fonts" / "Amiri-Regular.ttf",
        Path(settings.BASE_DIR) / "fonts" / "NotoNaskhArabic-Regular.ttf",
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/tahoma.ttf"),
        Path("C:/Windows/Fonts/tradbdo.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
    ]
    for candidate in candidates:
        try:
            if not candidate.exists():
                continue
            if "ArabicUI" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("ArabicUI", str(candidate)))
            return "ArabicUI"
        except Exception:
            continue
    return "Helvetica"


def pdf_text(value):
    if value is None:
        return ""
    text = str(value)
    if not text:
        return ""
    try:
        if arabic_reshaper is not None and get_display is not None and any("؀" <= ch <= "ۿ" for ch in text):
            return get_display(arabic_reshaper.reshape(text))
    except Exception:
        pass
    return text


def pdf_row(values):
    return [pdf_text(v) if isinstance(v, str) else v for v in values]


def export_excel_response(title, cols, rows, visible_value):
    wb = Workbook()
    ws = wb.active
    ws.title = "البيانات"
    headers = [label for field, label in cols if field != "__actions__"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for obj in rows:
        ws.append([visible_value(obj, field) for field, _ in cols if field != "__actions__"])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 30)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{title}.xlsx"'
    return response


def export_pdf_response(title, cols, rows, visible_value):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=20, bottomMargin=20)
    font_name = register_pdf_font()
    styles = getSampleStyleSheet()
    styles["Normal"].fontName = font_name
    styles["Title"].fontName = font_name
    pdf_cols = [(field, label) for field, label in cols if field in PDF_ALLOWED_FIELDS]
    data = [[label for _, label in pdf_cols]]
    for obj in rows[:1500]:
        data.append([str(visible_value(obj, field)) for field, _ in pdf_cols])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#EAF1F8")]),
    ]))
    story = [Paragraph(title, styles["Title"]), Spacer(1, 10), table]
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title}.pdf"'
    return response
