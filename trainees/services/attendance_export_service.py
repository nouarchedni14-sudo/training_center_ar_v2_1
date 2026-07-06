from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import PageBreak, SimpleDocTemplate, Spacer, Table, TableStyle


class AttendanceExportService:
    """يبني ملفات تصدير الغيابات بعد أن تكون payload جاهزة داخل views."""

    def __init__(
        self,
        *,
        status_label_map,
        group_items_by_specialty,
        safe_sheet_title,
        attendance_rows_semester_label,
        apply_official_header,
        register_pdf_font,
        pdf_text,
        pdf_row,
        attendance_template_title,
        attendance_scope_subtitle,
        export_header_lines,
    ):
        self.status_label_map = status_label_map
        self.group_items_by_specialty = group_items_by_specialty
        self.safe_sheet_title = safe_sheet_title
        self.attendance_rows_semester_label = attendance_rows_semester_label
        self.apply_official_header = apply_official_header
        self.register_pdf_font = register_pdf_font
        self.pdf_text = pdf_text
        self.pdf_row = pdf_row
        self.attendance_template_title = attendance_template_title
        self.attendance_scope_subtitle = attendance_scope_subtitle
        self.export_header_lines = export_header_lines

    def build_workbook(self, program, payload):
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        rows = payload["rows"]
        scope = payload["scope"]
        slot_count = payload.get("slot_count", 1)
        columns = payload["columns"]
        label_map = self.status_label_map(program)
        show_all_specialties = payload["show_all_specialties"]
        specialty_value = (scope.get("specialty") or "").strip()

        grouped_rows = [(specialty_value or "كل التخصصات", rows)]
        if show_all_specialties:
            grouped_rows = self.group_items_by_specialty(rows, lambda row: getattr(row["trainee"], "التخصص", "") or "بدون تخصص")

        for idx, (specialty_name, specialty_rows) in enumerate(grouped_rows, start=1):
            ws = wb.create_sheet(title=self.safe_sheet_title(specialty_name or f"تخصص {idx}"))
            ws.sheet_view.rightToLeft = True
            ws.freeze_panes = "C12"

            show_specialty_column = bool(show_all_specialties)
            base_header = ["الرقم", "الاسم و اللقب"]
            if show_specialty_column:
                base_header.append("التخصص")
            base_col_count = len(base_header)
            total_columns = base_col_count + (len(columns) * slot_count)

            semester_label = self.attendance_rows_semester_label(specialty_rows)
            self.apply_official_header(ws, total_columns, program, scope, specialty_label=specialty_name, semester_label=semester_label)

            header_row_1 = 10
            header_row_2 = 11
            data_start_row = 12

            dark_fill = PatternFill("solid", fgColor="1F4E78")
            light_fill = PatternFill("solid", fgColor="D9E2F3")
            border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

            for idx_col, title in enumerate(base_header, start=1):
                ws.merge_cells(start_row=header_row_1, start_column=idx_col, end_row=header_row_2, end_column=idx_col)
                cell = ws.cell(row=header_row_1, column=idx_col)
                cell.value = title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = dark_fill

            current_col = base_col_count + 1
            for col in columns:
                if slot_count > 1:
                    ws.merge_cells(start_row=header_row_1, start_column=current_col, end_row=header_row_1, end_column=current_col + slot_count - 1)
                    ws.merge_cells(start_row=header_row_2, start_column=current_col, end_row=header_row_2, end_column=current_col + slot_count - 1)
                ws.cell(row=header_row_1, column=current_col).value = col["weekday_label"]
                ws.cell(row=header_row_2, column=current_col).value = str(col["day_num"]).zfill(2)
                for offset in range(slot_count):
                    for row_num in (header_row_1, header_row_2):
                        cell = ws.cell(row=row_num, column=current_col + offset)
                        cell.font = Font(bold=True, color="FFFFFF" if row_num == header_row_1 else "000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = dark_fill if row_num == header_row_1 else light_fill
                        cell.border = border
                current_col += slot_count

            for row_num in (header_row_1, header_row_2):
                for col_num in range(1, total_columns + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border
                    if col_num <= base_col_count:
                        cell.fill = dark_fill
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif row_num == header_row_2:
                        cell.fill = light_fill
                        cell.font = Font(bold=True, color="000000")

            for row_idx, row in enumerate(specialty_rows, start=data_start_row):
                values = [str(row["index"]).zfill(2), f"{row['trainee'].اللقب} {row['trainee'].الاسم}"]
                if show_specialty_column:
                    values.append(getattr(row["trainee"], "التخصص", "") or "")
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border

                current_col = base_col_count + 1
                for day_cell in row["cells"]:
                    if program == "apprentice":
                        slots = day_cell.get("slots", [])
                        for slot_index in range(slot_count):
                            status = slots[slot_index].get("status", "") if slot_index < len(slots) else ""
                            cell = ws.cell(row=row_idx, column=current_col)
                            cell.value = label_map.get(status, "")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = border
                            current_col += 1
                    else:
                        cell = ws.cell(row=row_idx, column=current_col)
                        cell.value = label_map.get(day_cell.get("status", ""), "")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border
                        current_col += 1
                ws.row_dimensions[row_idx].height = 19

            from openpyxl.utils import get_column_letter

            ws.column_dimensions["A"].width = 5.0
            ws.column_dimensions["B"].width = 26.0
            next_col = 3
            if show_specialty_column:
                ws.column_dimensions["C"].width = 24.0
                next_col = 4
            for col_num in range(next_col, total_columns + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 4.5 if slot_count > 1 else 5.6

            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = 8 if slot_count > 1 else 9
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins.left = 0.24
            ws.page_margins.right = 0.24
            ws.page_margins.top = 0.2
            ws.page_margins.bottom = 0.2
            ws.print_options.horizontalCentered = True
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_title_rows = "$1:$11"
            ws.print_area = f"$A$1:${get_column_letter(total_columns)}${max(data_start_row, ws.max_row)}"

        return wb

    def build_pdf_bytes(self, program, payload):
        font_name = self.register_pdf_font()
        styles = getSampleStyleSheet()
        for style_name in ("Normal", "Title", "Heading2"):
            styles[style_name].fontName = font_name
            styles[style_name].alignment = 1

        rows = payload["rows"]
        columns = payload["columns"]
        scope = payload["scope"]
        slot_count = payload.get("slot_count", 1)
        label_map = self.status_label_map(program)
        show_all_specialties = payload["show_all_specialties"]
        specialty_value = (scope.get("specialty") or "").strip()

        grouped_rows = [(specialty_value or "كل التخصصات", rows)]
        if show_all_specialties:
            grouped_rows = self.group_items_by_specialty(rows, lambda row: getattr(row["trainee"], "التخصص", "") or "بدون تخصص")

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A3),
            rightMargin=18,
            leftMargin=18,
            topMargin=16,
            bottomMargin=16,
        )

        story = []
        dark = colors.HexColor("#1F4E78")
        light = colors.HexColor("#D9E2F3")
        white = colors.white

        for index, (specialty_name, specialty_rows) in enumerate(grouped_rows, start=1):
            semester_label = self.attendance_rows_semester_label(specialty_rows)

            header_table = Table([[self.pdf_text(line)] for line in self.export_header_lines], colWidths=[doc.width], hAlign="CENTER")
            header_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 1), 12),
                ("FONTSIZE", (0, 2), (-1, -1), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 6))

            title_table = Table([[self.pdf_text(self.attendance_template_title(program, specialty_name))]], colWidths=[doc.width], hAlign="CENTER")
            title_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), dark),
                ("TEXTCOLOR", (0, 0), (-1, -1), white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 15),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(title_table)

            subtitle = self.attendance_scope_subtitle(scope, specialty_label=specialty_name, semester_label=semester_label)
            subtitle_table = Table([[self.pdf_text(subtitle)]], colWidths=[doc.width], hAlign="CENTER")
            subtitle_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), light),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(subtitle_table)
            story.append(Spacer(1, 10))

            header_row_1 = ["الرقم", "الاسم و اللقب"]
            header_row_2 = ["", ""]
            col_widths = [34, 140]
            show_specialty_column = bool(show_all_specialties)
            if show_specialty_column:
                header_row_1.append("التخصص")
                header_row_2.append("")
                col_widths.append(115)

            for col in columns:
                if slot_count > 1:
                    for _slot_index in range(slot_count):
                        header_row_1.append(self.pdf_text(col["weekday_label"]))
                        header_row_2.append(str(col["day_num"]).zfill(2))
                        col_widths.append(20)
                else:
                    header_row_1.append(self.pdf_text(col["weekday_label"]))
                    header_row_2.append(str(col["day_num"]).zfill(2))
                    col_widths.append(30)

            table_data = [self.pdf_row(header_row_1), self.pdf_row(header_row_2)]
            for row in specialty_rows:
                values = [str(row["index"]).zfill(2), f"{row['trainee'].اللقب} {row['trainee'].الاسم}"]
                if show_specialty_column:
                    values.append(getattr(row["trainee"], "التخصص", "") or "")
                for day_cell in row["cells"]:
                    if program == "apprentice":
                        slots = day_cell.get("slots", [])
                        for slot in slots[:slot_count]:
                            values.append(label_map.get(slot.get("status", ""), ""))
                        if len(slots) < slot_count:
                            values.extend([""] * (slot_count - len(slots)))
                    else:
                        values.append(label_map.get(day_cell.get("status", ""), ""))
                table_data.append(self.pdf_row(values))

            table = Table(table_data, colWidths=col_widths, repeatRows=2, hAlign="CENTER")
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), dark),
                ("TEXTCOLOR", (0, 0), (-1, 0), white),
                ("BACKGROUND", (0, 1), (-1, 1), light),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 1), 8),
                ("FONTSIZE", (0, 2), (-1, -1), 7),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.whitesmoke, colors.HexColor("#F4F7FB")]),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(table)
            if index < len(grouped_rows):
                story.append(PageBreak())

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
