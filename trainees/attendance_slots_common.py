from __future__ import annotations

import calendar
from collections import Counter
from datetime import date
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import content_disposition_header

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .attendance_slots_models import AttendanceSlotCell, AttendanceSlotSheet
from .models import حضوري_أولي, تمهين, مسائي_ومعابر, دفعة
from .permissions import has_program_permission, require_program_permission
from .evening_training_type import EVENING_TRAINING_TYPE_EVENING, EVENING_TRAINING_TYPE_CROSSING


SLOT_COUNT_PER_DAY = 4
VALID_SLOT_STATUSES = {"present", "absent"}

AR_WEEKDAYS = {
    0: "الإثنين",
    1: "الثلاثاء",
    2: "الأربعاء",
    3: "الخميس",
    4: "الجمعة",
    5: "السبت",
    6: "الأحد",
}

# أيام الاختيار اليدوي: من السبت إلى الخميس فقط، بدون الجمعة.
CUSTOM_ATTENDANCE_WEEKDAY_VALUES = [5, 6, 0, 1, 2, 3]

MONTH_CHOICES = [
    (1, "جانفي 01"),
    (2, "فيفري 02"),
    (3, "مارس 03"),
    (4, "أفريل 04"),
    (5, "ماي 05"),
    (6, "جوان 06"),
    (7, "جويلية 07"),
    (8, "أوت 08"),
    (9, "سبتمبر 09"),
    (10, "أكتوبر 10"),
    (11, "نوفمبر 11"),
    (12, "ديسمبر 12"),
]

SLOT_PROGRAMS = {
    "initial": {
        "label": "الحضوري الأولي بالحـصص",
        "short_label": "الحضوري الأولي",
        "description": "نظام جديد مستقل: كل يوم دراسي مقسم إلى 4 حصص، والحساب حسب الحصص البيداغوجية.",
        "model": حضوري_أولي,
        "weekdays": [6, 0, 1, 2, 3],  # الأحد إلى الخميس
        "table_url_name": "attendance_initial_slots",
        "stats_url_name": "attendance_initial_slots_stats",
        "export_url_name": "attendance_initial_slots_export",
        "sync_url_name": "attendance_initial_slots_sync_actions",
    },
    "apprentice": {
        "label": "التمهين بالحـصص",
        "short_label": "التمهين",
        "description": "نظام جديد مستقل: اختر يومي الدراسة مع إمكانية إضافة يوم ثالث، وكل يوم مقسم إلى 4 حصص.",
        "model": تمهين,
        "weekdays": [],
        "table_url_name": "attendance_apprentice_slots",
        "stats_url_name": "attendance_apprentice_slots_stats",
        "export_url_name": "attendance_apprentice_slots_export",
        "sync_url_name": "attendance_apprentice_slots_sync_actions",
    },
    "evening": {
        "label": "الدروس المسائية بالحـصص",
        "short_label": "الدروس المسائية",
        "description": "نظام مستقل للحصص: اختر أيام الدراسة الخاصة بالفوج، وكل يوم مقسم إلى 4 حصص.",
        "model": مسائي_ومعابر,
        "weekdays": [],
        "table_url_name": "attendance_evening_slots",
        "stats_url_name": "attendance_evening_slots_stats",
        "export_url_name": "attendance_evening_slots_export",
        "sync_url_name": "attendance_evening_slots_sync_actions",
    },
    "crossing": {
        "label": "المعابر بالحـصص",
        "short_label": "المعابر",
        "description": "نظام جديد مستقل للمعابر: اختر أيام الدراسة، وكل يوم مقسم إلى 4 حصص.",
        "model": مسائي_ومعابر,
        "weekdays": [],
        "table_url_name": "attendance_crossing_slots",
        "stats_url_name": "attendance_crossing_slots_stats",
        "export_url_name": "attendance_crossing_slots_export",
        "sync_url_name": "attendance_crossing_slots_sync_actions",
    },
}


def _month_label(month: int) -> str:
    return dict(MONTH_CHOICES).get(int(month), str(month))


def _is_bridge_specialty(specialty: str) -> bool:
    return "معابر" in ((specialty or "").strip())


def _evening_type_for_program(program: str):
    if program == "crossing":
        return EVENING_TRAINING_TYPE_CROSSING
    if program == "evening":
        return EVENING_TRAINING_TYPE_EVENING
    return None


def _filter_queryset_by_program(qs, program: str):
    training_type = _evening_type_for_program(program)
    if training_type:
        return qs.filter(نوع_التكوين=training_type)
    return qs


def _requires_custom_weekdays(program: str, specialty: str = "") -> bool:
    return program in {"apprentice", "evening", "crossing"}


def _allows_third_weekday(program: str) -> bool:
    return program in {"apprentice", "evening", "crossing"}


def _program_max_custom_weekdays(program: str) -> int:
    if program == "apprentice":
        return 3
    if program in {"evening", "crossing"}:
        return 5
    return len(SLOT_PROGRAMS.get(program, {}).get("weekdays", []))


def _custom_weekday_keys(program: str):
    max_days = _program_max_custom_weekdays(program)
    return tuple(f"weekday{i}" for i in range(1, max_days + 1))


def _weekday_choices_for_program(program: str):
    if _requires_custom_weekdays(program):
        return [{"value": value, "label": AR_WEEKDAYS[value]} for value in CUSTOM_ATTENDANCE_WEEKDAY_VALUES]
    return [{"value": v, "label": l} for v, l in AttendanceSlotSheet.WEEKDAY_CHOICES]


def _unique_clean_values(values):
    seen = set()
    result = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _active_trainees_qs(model_cls, program=None):
    today = timezone.localdate()
    return _filter_queryset_by_program(model_cls.objects, program or "").filter(
        Q(تاريخ_نهاية_التكوين__isnull=True) | Q(تاريخ_نهاية_التكوين__gt=today)
    ).exclude(
        Q(الحالة__icontains="مشطوب") |
        Q(الحالة__icontains="شطب") |
        Q(الحالة__icontains="مفصول") |
        Q(الحالة__icontains="منقطع") |
        Q(الحالة__icontains="متوقف") |
        Q(الحالة__icontains="موقوف") |
        Q(الحالة__icontains="انسحب")
    )


def _promotion_options(model_cls, program=None):
    promotion_ids = (
        _active_trainees_qs(model_cls, program=program)
        .exclude(الدفعة__isnull=True)
        .values_list("الدفعة_id", flat=True)
        .distinct()
    )
    return دفعة.objects.filter(id__in=promotion_ids, مفعلة=True).order_by("-السنة", "-رقم_الدورة")


def _specialty_options(model_cls, promotion=None, program=None):
    qs = _active_trainees_qs(model_cls, program=program)
    if promotion:
        qs = qs.filter(الدفعة=promotion)
    return sorted(_unique_clean_values(
        qs.order_by()
        .exclude(التخصص__isnull=True)
        .exclude(التخصص__exact="")
        .values_list("التخصص", flat=True)
    ))


def _parse_scope(program: str, request):
    today = timezone.localdate()

    def _int_value(name: str, default: int) -> int:
        raw = request.GET.get(name) or request.POST.get(name) or default
        try:
            return int(raw)
        except Exception:
            return int(default)

    month = _int_value("month", today.month)
    year = _int_value("year", today.year)
    if month < 1 or month > 12:
        month = today.month
    if year < 2000 or year > 2100:
        year = today.year

    specialty = (request.GET.get("specialty") or request.POST.get("specialty") or "").strip()
    promotion_id = (request.GET.get("promotion") or request.POST.get("promotion") or "").strip()
    promotion = None
    if promotion_id:
        try:
            promotion = دفعة.objects.get(pk=promotion_id)
        except Exception:
            promotion = None

    return {
        "month": month,
        "year": year,
        "specialty": specialty,
        "promotion_id": promotion_id,
        "promotion": promotion,
        "promotion_obj": promotion,
        "weekday1": request.GET.get("weekday1") or request.POST.get("weekday1") or "",
        "weekday2": request.GET.get("weekday2") or request.POST.get("weekday2") or "",
        "weekday3": request.GET.get("weekday3") or request.POST.get("weekday3") or "",
        "weekday4": request.GET.get("weekday4") or request.POST.get("weekday4") or "",
        "weekday5": request.GET.get("weekday5") or request.POST.get("weekday5") or "",
        "show_table": (request.GET.get("show_table") or request.POST.get("show_table") or "") == "1",
        "action": (request.GET.get("action") or request.POST.get("action") or "").strip(),
    }


def _selected_weekdays(program: str, scope: dict):
    program_conf = SLOT_PROGRAMS[program]
    if _requires_custom_weekdays(program, scope.get("specialty", "")):
        weekdays = []
        for key in _custom_weekday_keys(program):
            value = scope.get(key)
            if value in (None, ""):
                continue
            try:
                num = int(value)
            except Exception:
                continue
            if num not in weekdays:
                weekdays.append(num)
        return weekdays
    return list(program_conf["weekdays"])


def _weekday_validation(program: str, scope: dict):
    if not _requires_custom_weekdays(program, scope.get("specialty", "")):
        return {"is_valid": True, "message": "", "weekdays": list(SLOT_PROGRAMS[program]["weekdays"])}

    raw_keys = list(_custom_weekday_keys(program))

    provided = []
    cleaned = []
    invalid = False
    for key in raw_keys:
        value = scope.get(key)
        if value in (None, ""):
            continue
        provided.append(value)
        try:
            num = int(value)
        except Exception:
            invalid = True
            continue
        if num not in cleaned:
            cleaned.append(num)

    if invalid:
        return {"is_valid": False, "message": "قيمة يوم الدراسة غير صحيحة.", "weekdays": cleaned}
    if len(provided) != len(cleaned):
        return {"is_valid": False, "message": "لا يمكن اختيار نفس اليوم أكثر من مرة.", "weekdays": cleaned}
    if len(cleaned) < 2:
        return {"is_valid": False, "message": "اختر أيام الدراسة قبل عرض الجدول.", "weekdays": cleaned}
    if len(cleaned) > _program_max_custom_weekdays(program):
        return {"is_valid": False, "message": "عدد أيام الدراسة المحدد أكبر من المسموح.", "weekdays": cleaned}
    return {"is_valid": True, "message": "", "weekdays": cleaned}


def _month_dates(year: int, month: int, weekday_numbers):
    result = []
    for week in calendar.monthcalendar(year, month):
        for weekday_idx, day_num in enumerate(week):
            if not day_num:
                continue
            if weekday_idx in weekday_numbers:
                current = date(year, month, day_num)
                result.append({
                    "date": current,
                    "day_num": day_num,
                    "weekday": weekday_idx,
                    "weekday_label": AR_WEEKDAYS.get(weekday_idx, str(weekday_idx)),
                    "iso": current.isoformat(),
                })
    return result


def _trainee_queryset(program: str, scope: dict):
    model_cls = SLOT_PROGRAMS[program]["model"]
    qs = _active_trainees_qs(model_cls, program=program)
    if scope["promotion"]:
        qs = qs.filter(الدفعة=scope["promotion"])
    if scope["specialty"]:
        qs = qs.filter(التخصص=scope["specialty"])
    return model_cls, qs.only("id", "الاسم", "اللقب", "التخصص", "الدفعة_id", "تاريخ_نهاية_التكوين", "الحالة").order_by("التخصص", "اللقب", "الاسم")


def _status_choices():
    return [("present", "ح"), ("absent", "غ")]


def _status_label_map():
    return dict(_status_choices())


def _get_or_create_sheet(program: str, scope: dict, weekdays: list[int], user):
    filters = {
        "البرنامج": program,
        "الدفعة": scope["promotion"],
        "التخصص": scope["specialty"],
        "الشهر": scope["month"],
        "السنة": scope["year"],
        "يوم_الدراسة_1": weekdays[0] if len(weekdays) > 0 else None,
        "يوم_الدراسة_2": weekdays[1] if len(weekdays) > 1 else None,
        "يوم_الدراسة_3": weekdays[2] if len(weekdays) > 2 else None,
    }
    for idx in range(4, 6):
        field_name = f"يوم_الدراسة_{idx}"
        if hasattr(AttendanceSlotSheet, field_name):
            filters[field_name] = weekdays[idx - 1] if len(weekdays) >= idx else None
    sheet = AttendanceSlotSheet.objects.filter(**filters).first()
    if sheet:
        return sheet
    return AttendanceSlotSheet.objects.create(**filters, created_by=user if getattr(user, "is_authenticated", False) else None)


def _table_payload(program: str, request):
    if program not in SLOT_PROGRAMS:
        raise Http404()

    scope = _parse_scope(program, request)
    action = scope.get("action") or ""
    show_all_specialties = action == "all"
    if show_all_specialties:
        scope["specialty"] = ""

    model_cls, trainee_qs = _trainee_queryset(program, scope)
    promotion_options = _promotion_options(model_cls, program=program)
    specialty_options = _specialty_options(model_cls, scope["promotion"], program=program)

    validation = _weekday_validation(program, scope)
    weekdays = validation["weekdays"] if validation["is_valid"] else _selected_weekdays(program, scope)

    can_prepare = False
    if show_all_specialties:
        can_prepare = True
    elif scope["promotion"] and scope["specialty"]:
        can_prepare = True
    if _requires_custom_weekdays(program, scope.get("specialty", "")):
        can_prepare = can_prepare and validation["is_valid"]

    show_table = bool(scope.get("show_table")) and can_prepare
    columns = _month_dates(scope["year"], scope["month"], weekdays) if show_table else []
    sheet = _get_or_create_sheet(program, scope, weekdays, request.user) if show_table else None
    trainees = list(trainee_qs) if columns else []

    entry_map = {}
    if sheet and trainees and columns:
        trainee_ids = [t.pk for t in trainees]
        dates = [c["date"] for c in columns]
        for entry in AttendanceSlotCell.objects.filter(الكشف=sheet, trainee_id__in=trainee_ids, التاريخ__in=dates).only("id", "trainee_id", "التاريخ", "رقم_الحصة", "الحالة"):
            entry_map[(entry.trainee_id, entry.التاريخ.isoformat(), entry.رقم_الحصة)] = entry

    label_map = _status_label_map()
    rows = []
    for idx, trainee in enumerate(trainees, start=1):
        cells = []
        for col in columns:
            slots = []
            for slot_no in range(1, SLOT_COUNT_PER_DAY + 1):
                entry = entry_map.get((trainee.pk, col["iso"], slot_no))
                status = entry.الحالة if entry else ""
                slots.append({
                    "slot": slot_no,
                    "status": status,
                    "display_label": label_map.get(status, ""),
                })
            cells.append({
                "date": col["date"],
                "iso": col["iso"],
                "slots": slots,
            })
        rows.append({"index": idx, "trainee": trainee, "cells": cells})

    scheduled_slots_per_trainee = len(columns) * SLOT_COUNT_PER_DAY
    return {
        "scope": scope,
        "action": action,
        "program": program,
        "program_label": SLOT_PROGRAMS[program]["label"],
        "program_short_label": SLOT_PROGRAMS[program]["short_label"],
        "program_description": SLOT_PROGRAMS[program]["description"],
        "model_cls": model_cls,
        "promotion_options": promotion_options,
        "specialty_options": specialty_options,
        "weekday_choices": _weekday_choices_for_program(program),
        "month_choices": [{"value": v, "label": l} for v, l in MONTH_CHOICES],
        "weekday_validation": validation,
        "is_bridge_specialty": _requires_custom_weekdays(program, scope.get("specialty", "")),
        "requires_custom_weekdays": _requires_custom_weekdays(program, scope.get("specialty", "")),
        "allow_third_weekday": _allows_third_weekday(program),
        "allow_fourth_weekday": program in {"evening", "crossing"},
        "allow_fifth_weekday": program in {"evening", "crossing"},
        "show_all_specialties": show_all_specialties,
        "show_table": show_table,
        "columns": columns,
        "sheet": sheet,
        "rows": rows,
        "slot_count": SLOT_COUNT_PER_DAY,
        "display_columns_count": len(columns) * SLOT_COUNT_PER_DAY,
        "scheduled_slots_per_trainee": scheduled_slots_per_trainee,
        "total_pedagogical_slots": len(rows) * scheduled_slots_per_trainee,
    }


def _apply_save(request, program: str, payload: dict):
    sheet = payload.get("sheet")
    rows = payload.get("rows") or []
    columns = payload.get("columns") or []
    if not sheet:
        messages.error(request, "اعرض الجدول أولًا قبل الحفظ.")
        return 0

    trainee_ids = [row["trainee"].pk for row in rows]
    dates = [col["date"] for col in columns]
    existing = {
        (e.trainee_id, e.التاريخ.isoformat(), e.رقم_الحصة): e
        for e in AttendanceSlotCell.objects.filter(الكشف=sheet, trainee_id__in=trainee_ids, التاريخ__in=dates)
    }

    to_create = []
    to_update = []
    to_delete = []
    for row in rows:
        trainee = row["trainee"]
        for cell in row["cells"]:
            for slot in cell.get("slots", []):
                slot_no = int(slot.get("slot") or 1)
                raw = (request.POST.get(f"status__{trainee.pk}__{cell['iso']}__{slot_no}") or "").strip()
                status = raw if raw in VALID_SLOT_STATUSES else ""
                entry = existing.get((trainee.pk, cell["iso"], slot_no))
                if not status:
                    if entry:
                        to_delete.append(entry.pk)
                    continue
                if entry is None:
                    to_create.append(AttendanceSlotCell(
                        الكشف=sheet,
                        trainee_id=trainee.pk,
                        التاريخ=cell["date"],
                        رقم_الحصة=slot_no,
                        الحالة=status,
                        recorded_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                    ))
                elif entry.الحالة != status or entry.recorded_by_id != getattr(request.user, "id", None):
                    entry.الحالة = status
                    entry.recorded_by = request.user if getattr(request.user, "is_authenticated", False) else None
                    to_update.append(entry)

    with transaction.atomic():
        if to_delete:
            AttendanceSlotCell.objects.filter(pk__in=to_delete).delete()
        if to_create:
            AttendanceSlotCell.objects.bulk_create(to_create, batch_size=1000)
        if to_update:
            AttendanceSlotCell.objects.bulk_update(to_update, ["الحالة", "recorded_by", "updated_at"], batch_size=1000)
    return len(to_create) + len(to_update)


def _apply_delete(payload: dict):
    sheet = payload.get("sheet")
    rows = payload.get("rows") or []
    columns = payload.get("columns") or []
    if not sheet:
        return 0
    trainee_ids = [row["trainee"].pk for row in rows]
    dates = [col["date"] for col in columns]
    deleted, _ = AttendanceSlotCell.objects.filter(الكشف=sheet, trainee_id__in=trainee_ids, التاريخ__in=dates).delete()
    return deleted


def _preserved_query_from_post(post):
    keep = {}
    for key in ("month", "year", "promotion", "specialty", "weekday1", "weekday2", "weekday3", "weekday4", "weekday5", "show_table", "action"):
        value = post.get(key)
        if value is not None and value != "":
            keep[key] = value
    return urlencode(keep)




def _auto_sync_slot_actions(request, program: str, payload: dict) -> dict:
    """تسجيل الإعذارات والاستدعاءات تلقائيًا بعد حفظ جدول الحصص.

    هذا يجعل نظام الحصص الجديد يعمل مثل النظام القديم: بمجرد الحفظ يتم
    تحليل الغيابات وإنشاء/تحديث الإعذارات والاستدعاءات المناسبة.
    """
    from .services.attendance_slot_action_sync_service import sync_slot_attendance_actions
    return sync_slot_attendance_actions(program, payload, request.user)


def _sync_result_message(result: dict) -> str:
    return (
        f"الإعذارات المنشأة: {result.get('created_excuses', 0)}، "
        f"الإعذارات المحدثة: {result.get('updated_excuses', 0)}، "
        f"الإعذارات المؤرشفة: {result.get('archived_excuses', 0)}، "
        f"استدعاءات الغيابات المتذبذبة/المتكررة المنشأة: {result.get('created_summons', 0)}، "
        f"استدعاءات الغيابات المتذبذبة/المتكررة المحدثة/الملغاة: {result.get('updated_summons', 0)}."
    )

@login_required
def attendance_slots_program(request, program: str):
    if program not in SLOT_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    payload = _table_payload(program, request)

    if request.method == "POST":
        post_action = (request.POST.get("post_action") or "save").strip()
        if post_action in {"save", "delete_saved", "save_and_stats", "sync_actions"}:
            require_program_permission(request, program, "change")
            query = _preserved_query_from_post(request.POST)
            if post_action == "delete_saved":
                deleted = _apply_delete(payload)
                sync_result = _auto_sync_slot_actions(request, program, payload)
                messages.success(
                    request,
                    f"تم حذف {deleted} خلية محفوظة من جدول الحصص الجديد، وتم تحديث الإعذارات والاستدعاءات تلقائيًا. "
                    + _sync_result_message(sync_result),
                )
                return redirect(request.path + (f"?{query}" if query else ""))

            saved = _apply_save(request, program, payload)
            sync_result = _auto_sync_slot_actions(request, program, payload)

            if post_action == "save_and_stats":
                if saved:
                    messages.success(
                        request,
                        f"تم حفظ {saved} تغيير، وتم تسجيل الإعذارات والاستدعاءات تلقائيًا، ثم فتح صفحة الحساب. "
                        + _sync_result_message(sync_result),
                    )
                else:
                    messages.info(
                        request,
                        "تم تحديث الإعذارات والاستدعاءات تلقائيًا ثم فتح صفحة الحساب. "
                        + _sync_result_message(sync_result),
                    )
                stats_url = reverse(SLOT_PROGRAMS[program]["stats_url_name"])
                return redirect(stats_url + (f"?{query}" if query else ""))

            messages.success(
                request,
                f"تم حفظ {saved} تغيير في جدول الحصص الجديد، وتم تسجيل الإعذارات والاستدعاءات تلقائيًا. "
                + _sync_result_message(sync_result),
            )
            return redirect(request.path + (f"?{query}" if query else ""))

    return render(request, "trainees/attendance_slots_grid.html", {
        **payload,
        "title": payload["program_label"],
        "can_change": has_program_permission(request.user, program, "change"),
        "status_choices": _status_choices(),
        "status_display_map": _status_label_map(),
        "current_query": request.GET.urlencode(),
        "table_url_name": SLOT_PROGRAMS[program]["table_url_name"],
        "stats_url_name": SLOT_PROGRAMS[program]["stats_url_name"],
        "export_url_name": SLOT_PROGRAMS[program]["export_url_name"],
        "sync_url_name": SLOT_PROGRAMS[program]["sync_url_name"],
    })



@login_required
def attendance_slots_sync_actions(request, program: str):
    if program not in SLOT_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "change")
    if request.method != "POST":
        return redirect(f"{reverse(SLOT_PROGRAMS[program]['table_url_name'])}?{request.GET.urlencode()}")

    payload = _table_payload(program, request)
    saved = _apply_save(request, program, payload)

    result = _auto_sync_slot_actions(request, program, payload)

    query = _preserved_query_from_post(request.POST)
    messages.success(
        request,
        "تم تحليل جدول الحصص وتسجيل الإعذارات والاستدعاءات تلقائيًا. "
        f"تم فحص {result.get('checked', 0)} متكون. "
        + _sync_result_message(result)
        + f" تغييرات الجدول المحفوظة أولًا: {saved}.",
    )
    return redirect(f"{reverse(SLOT_PROGRAMS[program]['table_url_name'])}?{query}")


def _stats_payload(program: str, request):
    payload = _table_payload(program, request)
    rows = payload["rows"]
    columns = payload["columns"]
    scheduled_slots_per_trainee = len(columns) * SLOT_COUNT_PER_DAY

    stats_rows = []
    totals = Counter()
    total_pedagogical_slots = len(rows) * scheduled_slots_per_trainee

    for row in rows:
        counter = Counter()
        for cell in row.get("cells", []):
            for slot in cell.get("slots", []):
                status = (slot.get("status") or "").strip()
                if status in VALID_SLOT_STATUSES:
                    counter[status] += 1
        absent_count = counter["absent"]
        official_present = max(scheduled_slots_per_trainee - absent_count, 0)
        absence_rate = round((absent_count / scheduled_slots_per_trainee) * 100, 2) if scheduled_slots_per_trainee else 0
        presence_rate = round(100 - absence_rate, 2) if scheduled_slots_per_trainee else 0
        stats_rows.append({
            "index": row["index"],
            "trainee": row["trainee"],
            "recorded_present_count": counter["present"],
            "official_present_count": official_present,
            "absent_count": absent_count,
            "scheduled_slots": scheduled_slots_per_trainee,
            "absence_rate": absence_rate,
            "presence_rate": presence_rate,
        })
        totals["present"] += official_present
        totals["recorded_present"] += counter["present"]
        totals["absent"] += absent_count
        totals["scheduled"] += scheduled_slots_per_trainee

    stats_rows.sort(key=lambda item: (-item["absence_rate"], -item["absent_count"], getattr(item["trainee"], "التخصص", "") or "", getattr(item["trainee"], "اللقب", "") or "", getattr(item["trainee"], "الاسم", "") or ""))
    for index, row in enumerate(stats_rows, start=1):
        row["display_index"] = index

    official_absence_rate = round((totals["absent"] / totals["scheduled"]) * 100, 2) if totals["scheduled"] else 0
    payload.update({
        "stats_rows": stats_rows,
        "stats_totals": {
            "official_present_count": totals["present"],
            "recorded_present_count": totals["recorded_present"],
            "absent_count": totals["absent"],
            "scheduled_slots": totals["scheduled"],
        },
        "trainee_count": len(rows),
        "displayed_days_count": len(columns),
        "official_absence_rate": official_absence_rate,
        "official_presence_rate": round(100 - official_absence_rate, 2) if totals["scheduled"] else 0,
        "total_pedagogical_slots": total_pedagogical_slots,
        "scheduled_slots_per_trainee": scheduled_slots_per_trainee,
    })
    return payload


@login_required
def attendance_slots_stats(request, program: str):
    if program not in SLOT_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")

    current_query = request.GET.urlencode()

    # عند الضغط على زر "حفظ ثم حساب النسبة الرسمية" من صفحة الجدول،
    # نأتي إلى صفحة الحساب عبر POST، لذلك يجب حفظ الخانات أولاً ثم حساب النسبة.
    # هذا يمنع اختفاء القيم أو بقاء الصفحة في الجدول بدون حساب.
    if request.method == "POST":
        require_program_permission(request, program, "change")
        table_payload = _table_payload(program, request)
        saved = _apply_save(request, program, table_payload)
        sync_result = _auto_sync_slot_actions(request, program, table_payload)
        current_query = _preserved_query_from_post(request.POST)
        if saved:
            messages.success(
                request,
                f"تم حفظ {saved} تغيير، وتم تحديث الإعذارات والاستدعاءات تلقائيًا، ثم حساب النسبة الرسمية. "
                + _sync_result_message(sync_result),
            )
        else:
            messages.info(
                request,
                "تم تحديث الإعذارات والاستدعاءات تلقائيًا، ثم حساب النسبة الرسمية. "
                + _sync_result_message(sync_result),
            )

    payload = _stats_payload(program, request)
    return render(request, "trainees/attendance_slots_stats.html", {
        **payload,
        "title": f"{payload['program_label']} - حساب النسبة الرسمية",
        "current_query": current_query,
        "table_url_name": SLOT_PROGRAMS[program]["table_url_name"],
        "export_url_name": SLOT_PROGRAMS[program]["export_url_name"],
        "sync_url_name": SLOT_PROGRAMS[program]["sync_url_name"],
    })


def _safe_sheet_title(value: str) -> str:
    value = (value or "غيابات").strip()
    for ch in '[]:*?/\\':
        value = value.replace(ch, ' ')
    return ' '.join(value.split())[:31] or "غيابات"


def _set_filename(response, filename: str):
    response["Content-Disposition"] = content_disposition_header(True, filename)
    return response


def _write_table_excel(wb, program: str, payload: dict):
    ws = wb.active
    ws.title = _safe_sheet_title("جدول الحصص")
    ws.sheet_view.rightToLeft = True
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    light_fill = PatternFill("solid", fgColor="D9E2F3")

    columns = payload["columns"]
    rows = payload["rows"]
    show_specialty = bool(payload["show_all_specialties"])
    total_cols = 2 + (1 if show_specialty else 0) + len(columns) * SLOT_COUNT_PER_DAY
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1).value = f"جدول الغيابات بالحصة - {payload['program_label']}"
    ws.cell(1, 1).font = Font(bold=True, size=15)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(2, 1).value = f"{_month_label(payload['scope']['month'])} {payload['scope']['year']} - كل يوم = 4 حصص"
    ws.cell(2, 1).alignment = Alignment(horizontal="center")

    base = ["الرقم", "الاسم واللقب"]
    if show_specialty:
        base.append("التخصص")
    for col_idx, title in enumerate(base, start=1):
        ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
        c = ws.cell(4, col_idx)
        c.value = title
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    current_col = len(base) + 1
    for col in columns:
        ws.merge_cells(start_row=4, start_column=current_col, end_row=4, end_column=current_col + SLOT_COUNT_PER_DAY - 1)
        ws.cell(4, current_col).value = f"{col['weekday_label']} {col['day_num']:02d}"
        ws.cell(4, current_col).font = Font(bold=True, color="FFFFFF")
        ws.cell(4, current_col).fill = header_fill
        ws.cell(4, current_col).alignment = Alignment(horizontal="center")
        for offset in range(SLOT_COUNT_PER_DAY):
            c = ws.cell(5, current_col + offset)
            c.value = f"ح{offset+1}"
            c.font = Font(bold=True)
            c.fill = light_fill
            c.alignment = Alignment(horizontal="center")
        current_col += SLOT_COUNT_PER_DAY

    label_map = _status_label_map()
    for r, row in enumerate(rows, start=6):
        ws.cell(r, 1).value = row["index"]
        ws.cell(r, 2).value = f"{row['trainee'].اللقب} {row['trainee'].الاسم}"
        col_pos = 3
        if show_specialty:
            ws.cell(r, 3).value = getattr(row["trainee"], "التخصص", "") or ""
            col_pos = 4
        for cell in row["cells"]:
            for slot in cell["slots"]:
                ws.cell(r, col_pos).value = label_map.get(slot.get("status") or "", "")
                ws.cell(r, col_pos).alignment = Alignment(horizontal="center")
                col_pos += 1

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    if show_specialty:
        ws.column_dimensions["C"].width = 25
    ws.freeze_panes = "C6" if not show_specialty else "D6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _write_stats_excel(wb, payload: dict):
    ws = wb.active
    ws.title = _safe_sheet_title("النسبة الرسمية")
    ws.sheet_view.rightToLeft = True
    headers = ["الترتيب", "الاسم واللقب", "التخصص", "الحضور الرسمي", "الغياب بالحصة", "الحصص البيداغوجية", "نسبة الغياب", "نسبة الحضور"]
    ws.append([f"إحصائيات الغياب بالحصة - {payload['program_label']}"])
    ws.append([f"القاعدة: الغيابات بالحصة ÷ عدد المتكونين × عدد الأيام الظاهرة × 4"])
    ws.append([""])
    ws.append(headers)
    for row in payload.get("stats_rows", []):
        trainee = row["trainee"]
        ws.append([
            row["display_index"],
            f"{trainee.اللقب} {trainee.الاسم}",
            getattr(trainee, "التخصص", "") or "",
            row["official_present_count"],
            row["absent_count"],
            row["scheduled_slots"],
            f"{row['absence_rate']}%",
            f"{row['presence_rate']}%",
        ])
    ws.append(["", "الإجمالي", "", payload["stats_totals"]["official_present_count"], payload["stats_totals"]["absent_count"], payload["stats_totals"]["scheduled_slots"], f"{payload['official_absence_rate']}%", f"{payload['official_presence_rate']}%"])
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14


@login_required
def attendance_slots_export(request, program: str, fmt: str):
    if program not in SLOT_PROGRAMS:
        raise Http404()
    require_program_permission(request, program, "view")
    fmt = (fmt or "").lower().strip()
    if fmt == "excel":
        fmt = "xlsx"
    if fmt != "xlsx":
        messages.error(request, "التصدير المتاح حاليًا لنظام الحصص الجديد هو Excel فقط.")
        return redirect(reverse(SLOT_PROGRAMS[program]["table_url_name"]) + ("?" + request.GET.urlencode() if request.GET else ""))

    export_type = (request.GET.get("type") or "table").strip()
    payload = _stats_payload(program, request) if export_type == "stats" else _table_payload(program, request)
    if not payload.get("columns"):
        messages.error(request, "اعرض الجدول أولًا قبل التصدير.")
        return redirect(reverse(SLOT_PROGRAMS[program]["table_url_name"]) + ("?" + request.GET.urlencode() if request.GET else ""))

    wb = Workbook()
    if export_type == "stats":
        _write_stats_excel(wb, payload)
        filename = f"إحصائيات الغياب بالحصة - {SLOT_PROGRAMS[program]['short_label']} - {_month_label(payload['scope']['month'])} {payload['scope']['year']}.xlsx"
    else:
        _write_table_excel(wb, program, payload)
        filename = f"جدول الغياب بالحصة - {SLOT_PROGRAMS[program]['short_label']} - {_month_label(payload['scope']['month'])} {payload['scope']['year']}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return _set_filename(response, filename)
