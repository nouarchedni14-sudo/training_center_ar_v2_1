import unittest
from datetime import datetime
from types import SimpleNamespace

from openpyxl import Workbook

from trainees.services.attendance_stats_export_service import (
    attendance_stats_export_excel,
    saved_attendance_stats_export_excel,
    write_attendance_stats_excel_sheet,
)


class AttendanceStatsExportServiceTests(unittest.TestCase):
    def _dummy_scope(self):
        return {"year": 2026, "month": 4, "specialty": "", "batch": None}

    def test_write_attendance_stats_excel_sheet_populates_sheet(self):
        wb = Workbook()
        ws = wb.active
        trainee = SimpleNamespace(اللقب="بن", الاسم="علي", التخصص="إعلام")
        payload = {"show_all_specialties": True, "scope": self._dummy_scope(), "displayed_days_count": 20, "slot_count": 1}
        rows = [{"trainee": trainee, "present_count": 15, "absent_count": 3, "excused_count": 1, "late_count": 1, "total_recorded": 20, "absence_rate": 15.0}]

        write_attendance_stats_excel_sheet(
            ws=ws, program="initial", payload=payload, stats_rows=rows, specialty_label="",
            apply_official_stats_excel_header=lambda *args, **kwargs: None,
            attendance_rows_semester_label=lambda _rows: "س1",
            safe_sheet_title=lambda value: value[:31],
        )

        self.assertEqual(ws.cell(17, 1).value, 1)
        self.assertIn("الاسم", ws.cell(16, 2).value)

    def test_attendance_stats_export_excel_returns_workbook_response(self):
        trainee = SimpleNamespace(اللقب="بن", الاسم="علي", التخصص="إعلام")
        payload = {
            "show_all_specialties": False,
            "scope": {"year": 2026, "month": 4, "specialty": "إعلام"},
            "stats_rows": [{"trainee": trainee, "present_count": 10, "absent_count": 2, "excused_count": 0, "late_count": 0, "total_recorded": 12, "absence_rate": 16.7}],
            "displayed_days_count": 12,
            "slot_count": 1,
        }
        def write_sheet(ws, program, payload, stats_rows, sheet_title=None, specialty_label=""):
            ws.cell(1,1,value="ok")
        response = attendance_stats_export_excel(
            program="initial", payload=payload, safe_sheet_title=lambda v: v[:31],
            group_items_by_specialty=lambda rows, fn: [("إعلام", rows)],
            write_sheet=write_sheet,
            attendance_export_filename=lambda kind, program, scope, specialty="", ext="xlsx": f"{kind}.{ext}",
            finalize_workbook_response=lambda wb, filename: filename,
        )
        self.assertEqual(response, "stats.xlsx")

    def test_saved_attendance_stats_export_excel_groups_by_specialty(self):
        row = SimpleNamespace(
            trainee_name="بن علي", trainee_specialty="إعلام", present_count=10, absent_count=2, excused_count=0, late_count=0, total_recorded=12, absence_rate=16.7, updated_at=datetime(2026,4,5,12,0)
        )
        context = {
            "detail_scope": {"program": "initial", "specialty": "", "year": 2026, "month": 4},
            "detail_rows": [row],
        }
        filenames = []
        saved_attendance_stats_export_excel(
            context=context,
            attendance_export_filename=lambda kind, program, scope, specialty="", ext="xlsx": f"{kind}.{ext}",
            finalize_workbook_response=lambda wb, filename: filenames.append(filename) or filename,
            group_items_by_specialty=lambda rows, fn: [("إعلام", rows)],
            safe_sheet_title=lambda v: v[:31],
            write_saved_sheet=lambda ws, context, detail_rows, sheet_title=None, specialty_label="": ws.cell(1,1,value="ok"),
        )
        self.assertEqual(filenames[0], "saved_stats.xlsx")


if __name__ == "__main__":
    unittest.main()
